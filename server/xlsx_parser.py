"""Honey 가 업로드한 .xlsx 에서 summary / yield / issue_table 추출.

클라이언트 report generator 산출물 레이아웃 전제(실측 기준):
  - 표는 A열을 비우고 B열부터 시작, 제목 배너 A1, 표 헤더 3행, 데이터 4행~.
  - summary 는 B4="DEVICE"(Device Feature 헤더), B7("2. Yield") + B15("3. Evaluation").
  - yield 는 B3="bin" 헤더행, 데이터 4행~.
  - issue_table 은 B3="Category" / C3="Bin" 헤더행.

산출물:
  1. sheet_data dict  — 순수 텍스트 데이터(스타일 없음). DB 저장·웹 렌더링용.
  2. summary dict     — 의미 단위 (기존 report_analysis_summary 매핑·검색 유지용).
  3. yield_rows list  — report_analysis_summary 저장용.
  4. issue_rows list  — comment 제외 legacy(하위호환).
  5. issue_images     — Distribution PNG 추출 (ISSUE_IMAGES_ENABLED=True 시).

앵커 탐색: 지정 셀 우선 → 키워드 전체 스캔 폴백 (_find_anchor).
"""
from io import BytesIO

# Issue_table 행별 임베드 PNG 추출 활성화 플래그. 골격 단계에선 False(빈 리스트).
ISSUE_IMAGES_ENABLED = False


def parse_report_xlsx(xlsx_bytes: bytes) -> dict:
    """xlsx 바이트열을 받아 sheet_data + legacy 의미 dict + 이미지 훅 결과를 반환."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed; pip install openpyxl") from exc

    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=False)

    sm = _get_sheet(wb, "summary")
    yl = _get_sheet(wb, "yield")
    it = _get_sheet(wb, "issue_table")

    summary = _extract_summary_sheet(sm)
    yield_rows = _rows_with_header(yl)
    issue_rows = [
        r for r in _rows_with_header(it, drop_cols={"Distribution"})
        if _stringify(r.get("bin")) != ""
    ]

    # 순수 텍스트 데이터 (스타일 없음) — DB 저장용
    summary_blocks = _extract_summary_blocks(sm)
    issue_full = _rows_with_header(it, drop_cols={"Distribution"})

    sheet_data = {}
    if summary_blocks:
        sheet_data["summary"] = summary_blocks
    if yield_rows:
        sheet_data["yield"] = yield_rows
    if issue_full is not None:
        sheet_data["issue_table"] = issue_full

    issue_hdr = _find_issue_header_row(it)
    issue_images = _extract_issue_images(it, issue_hdr, enabled=ISSUE_IMAGES_ENABLED)

    return {
        "summary": summary,
        "yield_rows": yield_rows,
        "issue_rows": issue_rows,
        "sheet_data": sheet_data,
        "issue_images": issue_images,
    }


# ── Summary 3-block 순수 텍스트 추출 ─────────────────────────────────────────

def _extract_summary_blocks(ws) -> dict | None:
    """Summary 시트 → {"blocks": [...]} 순수 텍스트 데이터.

    고정 앵커 탐색 순서:
      Block 1 "Device Feature" : B4="DEVICE" → 그 행이 헤더, 다음 행이 데이터
      Block 2 "2. Yield"       : B7/B8 사이 섹션 — 헤더행을 B8 기준 탐색
      Block 3 "3. Evaluation"  : B15/B16 사이 — 헤더행을 B16 기준 탐색
    각 블록은 헤더행부터 완전 빈 행 직전까지 포함.
    못 찾으면 None.
    """
    if ws is None:
        return None

    blocks = []

    # ── Block 1: Device Feature ──────────────────────────────────────────────
    a1 = _find_anchor(ws, "DEVICE", "B4")
    if a1:
        hdr_row = a1[0]
        reg = _table_region(ws, hdr_row)
        if reg:
            blk = _region_to_block(ws, reg, "1. Device Feature")
            if blk:
                blocks.append(blk)

    # ── Block 2: 2. Yield (Summary 시트 내) ──────────────────────────────────
    # 헤더행 탐색: B8 우선("Lot NO"), 없으면 "Yield" 키워드
    a2 = _find_anchor(ws, "Lot NO", "B8")
    if not a2:
        a2 = _find_anchor(ws, "Yield", None)
    if a2:
        hdr_row = a2[0]
        # 같은 헤더행에 비어있지 않은 셀이 1개뿐이면 다음 행 탐색
        row_cells = [ws.cell(hdr_row, c).value for c in range(1, (ws.max_column or 8) + 1)]
        non_empty = sum(1 for v in row_cells if _stringify(v) != "")
        if non_empty < 2:
            hdr_row += 1
        # Device Feature 블록과 겹치지 않도록
        if a1 and hdr_row <= a1[0]:
            a2 = None
        if a2:
            reg = _table_region(ws, hdr_row)
            if reg:
                blk = _region_to_block(ws, reg, "2. Yield")
                if blk:
                    blocks.append(blk)

    # ── Block 3: 3. Evaluation Summary ───────────────────────────────────────
    a3 = _find_anchor(ws, "Category", "B16")
    if not a3:
        a3 = _find_anchor(ws, "Evaluation", None)
        if a3:
            # 섹션 제목이 아니라 헤더행을 원하므로 한 행 아래
            a3 = (a3[0] + 1, a3[1], a3[2])
    if a3:
        hdr_row = a3[0]
        # Yield 블록과 겹치지 않도록
        if a2 and hdr_row <= (a2[0] if a2 else 0):
            a3 = None
        if a3:
            reg = _table_region(ws, hdr_row)
            if reg:
                blk = _region_to_block(ws, reg, "3. Evaluation Summary")
                if blk:
                    blocks.append(blk)

    if not blocks:
        return None
    return {"blocks": blocks}


def _region_to_block(ws, region, label: str) -> dict | None:
    """(r0,c0,r1,c1) 영역을 {"label","headers","rows"} 블록으로 변환.
    헤더행이 없거나 컬럼명 있는 셀이 1개 이하면 None 반환."""
    r0, c0, r1, c1 = region
    raw_headers = [_stringify(ws.cell(r0, c).value) for c in range(c0, c1 + 1)]
    named = sum(1 for h in raw_headers if h)
    if named < 2:
        return None
    rows = []
    for r in range(r0 + 1, r1 + 1):
        row = [_normalize(ws.cell(r, c).value) for c in range(c0, c1 + 1)]
        if any(_stringify(v) != "" for v in row):
            rows.append(row)
    return {"label": label, "headers": raw_headers, "rows": rows}


def _find_issue_header_row(ws):
    """issue_table 헤더행(1-based). 없으면 None."""
    if ws is None:
        return None
    a = _find_anchor(ws, "bin", "C3")
    return a[0] if a else _guess_header_row(ws)


# ── 앵커 탐색 ─────────────────────────────────────────────────────────────────

def _find_anchor(ws, text, expect_cell):
    """앵커 텍스트 위치 (row, col, by) 탐색. 1-based.

    1) expect_cell 이 None 이 아니고 해당 셀값이 text 와 대소문자 무시 일치하면 그 위치.
    2) 아니면 시트 전체를 스캔해 첫 일치 셀.
    3) 못 찾으면 None.
    """
    if ws is None:
        return None
    target = text.strip().lower()
    if expect_cell:
        try:
            cell = ws[expect_cell]
            if _stringify(cell.value).lower() == target:
                return (cell.row, cell.column, "position")
        except (KeyError, ValueError, AttributeError):
            pass
    for row in ws.iter_rows():
        for cell in row:
            if _stringify(cell.value).lower() == target:
                return (cell.row, cell.column, "keyword")
    return None


def _guess_header_row(ws):
    """비어있지 않은 셀이 2개 이상인 첫 행(1-based) — 배너(셀1개) 건너뜀."""
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row or 1)):
        if sum(1 for c in row if _stringify(c.value) != "") >= 2:
            return row[0].row
    return 1


def _table_region(ws, header_row):
    """헤더행 기준 (r0,c0,r1,c1). 열=헤더행 비어있지 않은 최소~최대 col,
    행=헤더행부터 [c0..c1] 전부 빈 행 직전까지."""
    if not header_row:
        return None
    cols = [c.column for c in ws[header_row] if _stringify(c.value) != ""]
    if not cols:
        return None
    c0, c1 = min(cols), max(cols)
    r1 = header_row
    r = header_row + 1
    last = ws.max_row or header_row
    while r <= last:
        if any(_stringify(ws.cell(r, c).value) != "" for c in range(c0, c1 + 1)):
            r1 = r
            r += 1
        else:
            break
    return (header_row, c0, r1, c1)


# ── Issue_table 행별 임베드 PNG 추출 (골격) ──────────────────────────────────

def _extract_issue_images(ws, header_row, enabled=False):
    """Issue_table 의 행별 임베드 PNG 추출. 골격 단계(enabled=False)에선 빈 리스트."""
    out = []
    if not enabled or ws is None or not header_row:
        return out
    for img in (getattr(ws, "_images", None) or []):
        try:
            frm = img.anchor._from
            grid_row = (int(frm.row) + 1) - header_row
            ref = img.ref
            data = ref.getvalue() if hasattr(ref, "getvalue") else ref
            if data:
                out.append({"row": grid_row, "png": bytes(data)})
        except (AttributeError, TypeError, ValueError):
            continue
    return out


def _get_sheet(wb, name):
    """시트 이름이 정확히 일치하지 않을 때 대소문자/공백 무시 fallback."""
    if name in wb.sheetnames:
        return wb[name]
    key = name.strip().lower()
    for s in wb.sheetnames:
        if s.strip().lower() == key:
            return wb[s]
    return None


# ── summary (legacy 의미 dict — report_analysis_summary 매핑용) ───────────────

def _extract_summary_sheet(ws) -> dict:
    """summary 시트에서 Device Feature / Yield / Major Fail / Evaluation 추출."""
    if ws is None:
        return {}
    out = {}
    rows = list(ws.iter_rows(values_only=True))
    out["title"] = _stringify(rows[0][0]) if rows and rows[0] else ""

    anchors = _find_anchors_2d(rows, {"1. Device Feature", "2. Yield",
                                      "Major Fail Bins", "3. Evaluation Summary"})

    if "1. Device Feature" in anchors:
        r, _c = anchors["1. Device Feature"]
        out["feature"] = _read_kv_2rows(rows, r + 1, r + 2)

    if "2. Yield" in anchors:
        r, _c = anchors["2. Yield"]
        out["yield_summary"] = _read_kv_2rows(rows, r + 1, r + 2,
                                              only_keys={"Lot NO", "Yield"})

    if "Major Fail Bins" in anchors:
        r, c = anchors["Major Fail Bins"]
        out["major_fail_bins"] = _read_major_fail(rows, r + 1, c)

    if "3. Evaluation Summary" in anchors:
        r, c = anchors["3. Evaluation Summary"]
        out["evaluation"] = _read_table_2d(rows, r + 1, c)

    if not any(k in out for k in ("feature", "yield_summary", "major_fail_bins", "evaluation")):
        out["raw_rows"] = [
            [_stringify(x) for x in row if x is not None]
            for row in rows if any(x is not None for x in row)
        ]
    return out


def _find_anchors_2d(rows, names):
    result = {}
    for i, row in enumerate(rows):
        if not row:
            continue
        for j, cell in enumerate(row):
            val = _stringify(cell)
            if not val:
                continue
            for name in names:
                if name in result:
                    continue
                if val == name or val.startswith(name):
                    result[name] = (i, j)
    return result


def _read_kv_2rows(rows, hdr_i, val_i, only_keys=None):
    if hdr_i >= len(rows) or val_i >= len(rows):
        return {}
    header = [_stringify(c) for c in (rows[hdr_i] or ())]
    values = list(rows[val_i] or ())
    out = {}
    for k, v in zip(header, values):
        if not k:
            continue
        if only_keys is not None and k not in only_keys:
            continue
        out[k] = _normalize(v)
    return out


def _read_major_fail(rows, start_i, lc, max_n=10):
    out = []
    for i in range(start_i, min(start_i + max_n, len(rows))):
        row = rows[i]
        if not row or lc >= len(row):
            break
        label = _stringify(row[lc])
        if not label:
            break
        out.append({
            "rank": label,
            "subject": _normalize(row[lc + 1]) if lc + 1 < len(row) else None,
            "ratio": _normalize(row[lc + 2]) if lc + 2 < len(row) else None,
        })
    return out


def _read_table_2d(rows, hdr_i, c, max_n=20):
    if hdr_i >= len(rows):
        return []
    header = [_stringify(x) for x in (rows[hdr_i] or ())[c:]]
    out = []
    for i in range(hdr_i + 1, min(hdr_i + 1 + max_n, len(rows))):
        row = rows[i]
        cells = list((row or ())[c:])
        if not cells or all(x is None or _stringify(x) == "" for x in cells):
            break
        d = {h: _normalize(v) for h, v in zip(header, cells) if h}
        if d:
            out.append(d)
    return out


# ── yield / issue_table (헤더 행 기반) ───────────────────────────────────────

def _rows_with_header(ws, drop_cols=None) -> list:
    """헤더 행을 찾아 그 아래 행들을 list[dict] 로 변환.

    drop_cols: 무시할 헤더명 집합 (예: 이미지가 들어있는 'Distribution' 컬럼).
    """
    if ws is None:
        return []
    drop = set(drop_cols or ())
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdr_i = next(
        (i for i, row in enumerate(rows[:10])
         if row and sum(1 for c in row if _stringify(c) != "") >= 2),
        0,
    )
    header = [_stringify(c) for c in rows[hdr_i]]
    keep_idx = [i for i, h in enumerate(header) if h and h not in drop]
    out = []
    for row in rows[hdr_i + 1:]:
        if not row or all(c is None for c in row):
            continue
        d = {}
        for i in keep_idx:
            if i < len(row):
                d[header[i]] = _normalize(row[i])
        if d:
            out.append(d)
    return out


def _stringify(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize(v):
    """primitive 만 유지 (datetime → ISO 문자열)."""
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    try:
        return v.isoformat()
    except AttributeError:
        return str(v)
