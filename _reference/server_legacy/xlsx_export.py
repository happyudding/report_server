"""
Excel report exporter — builds a 7-sheet .xlsx from a dataset.

Sheets (in order):
  1. summary      — Feature table, yield bins, evaluation, web link
  2. yield        — Full yield table with comments
  3. cpk          — CPK table with color coding and comments
  4. fail_data    — Fail item table with per-subject PNG thumbnails
  5. fail_values  — Per-student per-subject raw fail records (source/call/grade/class/type/subject/value/limits)
  6. issue_table  — Issue table with distribution PNG + editable fields
  7. distribution — Combined thumbnail grid PNG + link to web page

Thumbnail strategy (no live server calls, reads disk directly):
  - Try cairosvg  : SVG thumbs/  → PNG  (fast, install: pip install cairosvg)
  - Fallback PIL  : resize any pre-cached fail_pngs/ entry
  - Fallback none : cell left blank, image skipped

Combined distribution PNG requires pillow (pip install pillow).
"""

from __future__ import annotations

import json
import re
import threading
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import DATASETS_DIR, SERVER_BASE_URL
from analysis.table_builder import read_table_json, get_fail_values

# ── private helpers duplicated from dash_dashboard to avoid Dash import ────────

def _read_json_file(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _read_yield_comments(dataset_id: str) -> dict:
    return _read_json_file(DATASETS_DIR / dataset_id / "tables" / "yield_comments.json") or {}


def _read_cpk_comments(dataset_id: str) -> dict:
    return _read_json_file(DATASETS_DIR / dataset_id / "tables" / "cpk_comments.json") or {}


def _read_issue_comments(dataset_id: str) -> dict:
    return _read_json_file(DATASETS_DIR / dataset_id / "tables" / "issue_comments.json") or {}


def _read_summary_comments(dataset_id: str) -> dict:
    return _read_json_file(DATASETS_DIR / dataset_id / "tables" / "summary_yield_comments.json") or {}


def _read_summary_eval(dataset_id: str) -> dict:
    return _read_json_file(DATASETS_DIR / dataset_id / "tables" / "summary_eval.json") or {}


def _cpk_comment_key(subject, source) -> str:
    return f"{(subject or '').strip()}|{(source or '').strip()}"


def _merge_cpk_subject(rows: list) -> list:
    merged = []
    prev = None
    for row in rows:
        row = dict(row)
        cur = row.get("subject")
        if cur == prev:
            row["subject"] = ""
            row["lower_limit"] = ""
            row["upper_limit"] = ""
            row["units"] = ""
        else:
            prev = cur
        merged.append(row)
    return merged


def _yield_sort_key(r: dict):
    st = str(r.get("bin", "")).strip()
    is_pass = 0 if st == "1" else 1
    try:
        avg = float(r.get("avg") or 0)
    except (TypeError, ValueError):
        avg = 0.0
    return (is_pass, -avg)


def _build_issue_rows(fail_items: dict, sources: list, issue_comments: dict) -> list:
    rows = []
    for r in (fail_items or {}).get("rows", []) or []:
        st = str(r.get("bin", "")).strip()
        fail_subjects = r.get("fail_subjects") or []
        is_pass = st == "1"
        if is_pass or not fail_subjects:
            subject = "Pass" if is_pass else "N/A"
            subject_id = None
        else:
            top = fail_subjects[0]
            subject = top.get("subject", "N/A")
            subject_id = top.get("subject_id")
        saved = issue_comments.get(st) or {}
        if not isinstance(saved, dict):
            saved = {}
        row = {
            "bin": st,
            "subject": subject,
            "subject_id": subject_id,
            "avg": r.get("avg"),
            "issue_point": saved.get("issue_point", ""),
            "issue_comment": saved.get("comment", ""),
            "dev_comment": saved.get("dev_comment", ""),
            "pte_comment": saved.get("pte_comment", ""),
        }
        for src in sources or []:
            row[f"portion_{src}"] = r.get(f"portion_{src}")
        rows.append(row)
    return rows


# ── Style constants ─────────────────────────────────────────────────────────────

_FILL_HDR     = PatternFill("solid", fgColor="F6F7F9")
_FILL_SECTION = PatternFill("solid", fgColor="D0DFF0")
_FILL_COMMENT = PatternFill("solid", fgColor="FFFDF3")
_FILL_TOTAL   = PatternFill("solid", fgColor="EEF4FB")
_FILL_LOW_CPK = PatternFill("solid", fgColor="FFF3BF")
_FILL_PASS    = PatternFill("solid", fgColor="F0F8F0")

_FONT_TITLE   = Font(bold=True, size=14)
_FONT_SECTION = Font(bold=True, size=12, color="1F3D6B")
_FONT_HDR     = Font(bold=True, size=11)
_FONT_LINK    = Font(bold=True, size=12, color="2369B3", underline="single")
_FONT_SMALL   = Font(size=9, color="666666")
_FONT_BOLD    = Font(bold=True)

_THIN   = Side(style="thin",   color="CCCCCC")
_MEDIUM = Side(style="medium", color="B8C4D4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_LT = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

# thumbnail dimensions embedded in xlsx cells
THUMB_W_PX  = 120
THUMB_H_PX  = 83
THUMB_ROW_H = 65   # row height in points
THUMB_COL_W = 18   # column width in chars

# ── Cell / row helpers ──────────────────────────────────────────────────────────

def _c(ws, row, col, value=None, *, font=None, fill=None, align=None,
       border=_BORDER, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font   is not None: cell.font      = font
    if fill   is not None: cell.fill      = fill
    if align  is not None: cell.alignment = align
    if border is not None: cell.border    = border
    if num_fmt:            cell.number_format = num_fmt
    return cell


def _header_row(ws, row_num: int, headers: list[str], widths: list[float] | None = None):
    for i, h in enumerate(headers, 1):
        _c(ws, row_num, i, h, font=_FONT_HDR, fill=_FILL_HDR, align=_ALIGN_C)
    if widths:
        for i, w in enumerate(widths, 1):
            if w:
                ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row_num].height = 20


def _section_row(ws, row_num: int, text: str, span: int = 8):
    _c(ws, row_num, 1, text, font=_FONT_SECTION, fill=_FILL_SECTION,
       align=_ALIGN_L, border=None)
    if span > 1:
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num,   end_column=span,
        )
    ws.row_dimensions[row_num].height = 22


def _blank_row(ws, row_num: int, height: float = 6):
    ws.row_dimensions[row_num].height = height


# ── PNG helpers (disk-only, no server requests) ─────────────────────────────────

_png_lock = threading.Lock()


def _try_cairosvg(svg_path: Path, w: int, h: int) -> bytes | None:
    try:
        import cairosvg  # optional
        return cairosvg.svg2png(url=str(svg_path), output_width=w, output_height=h)
    except Exception:
        return None


def _try_cairosvg_bytes(svg_text: str, w: int, h: int) -> bytes | None:
    try:
        import cairosvg
        return cairosvg.svg2png(bytestring=svg_text.encode("utf-8"),
                                output_width=w, output_height=h)
    except Exception:
        return None


def _build_compact_svg_text(dataset_id: str, subject_id: int) -> str | None:
    """charts/<sid>.json + tables/meta.json 으로부터 compact SVG 를 즉석 생성.

    issue_table 용 — 디스크에 캐시된 일반 SVG 는 title/subtitle 여백이 커서
    썸네일 크기에서 산포가 거의 안 보임. compact 모드는 여백 축소 + 마커 굵게.
    """
    chart_path = DATASETS_DIR / dataset_id / "charts" / f"{subject_id}.json"
    meta_path  = DATASETS_DIR / dataset_id / "tables" / "meta.json"
    if not (chart_path.exists() and meta_path.exists()):
        return None
    try:
        chart = json.loads(chart_path.read_text(encoding="utf-8"))
        meta  = json.loads(meta_path.read_text(encoding="utf-8"))
        subj = next(
            (s for s in (meta.get("subjects") or []) if s.get("subject_id") == subject_id),
            None,
        )
        if not subj:
            return None
        # plotly trace 형식 → svg_builder 가 기대하는 {school, color, xs, ys}
        traces = []
        for t in (chart.get("data") or []):
            color = (t.get("marker") or {}).get("color", "#000000")
            traces.append({
                "school": t.get("name", ""),
                "color":  color,
                "xs":     t.get("x", []),
                "ys":     t.get("y", []),
            })
        from analysis.svg_builder import build_subject_svg
        return build_subject_svg(
            subject_id,
            subj.get("subject", ""),
            subj.get("units", ""),
            subj.get("lower_limit"),
            subj.get("upper_limit"),
            traces,
            chart.get("layout") or {},
            compact=True,
        )
    except Exception:
        return None


def _compactify_plotly_layout(layout: dict) -> dict:
    """plotly layout 을 썸네일용 compact 로 수정 (margin / font 축소)."""
    out = json.loads(json.dumps(layout))
    out["margin"] = {"l": 35, "r": 10, "t": 22, "b": 30}
    out.setdefault("font", {})["size"] = 8
    if "title" in out:
        out["title"].setdefault("font", {})["size"] = 10
    for axis_key in ("xaxis", "yaxis"):
        if axis_key in out:
            out[axis_key].setdefault("tickfont", {})["size"] = 7
            ax_title = out[axis_key].get("title")
            if isinstance(ax_title, str):
                # 문자열 → dict 로 승격 후 font 적용
                out[axis_key]["title"] = {"text": ax_title, "font": {"size": 7}}
            elif isinstance(ax_title, dict):
                ax_title.setdefault("font", {})["size"] = 7
    return out


def _try_kaleido_compact(dataset_id: str, subject_id: int, w: int, h: int) -> bytes | None:
    """kaleido 로 layout 만 compact 하게 수정해서 PNG 생성."""
    chart_path = DATASETS_DIR / dataset_id / "charts" / f"{subject_id}.json"
    if not chart_path.exists():
        return None
    try:
        import plotly.io as pio
        payload = json.loads(chart_path.read_text(encoding="utf-8"))
        layout = _compactify_plotly_layout(payload.get("layout") or {})
        # 마커도 좀 더 크게 (썸네일 사이즈에서 가시성 확보)
        data = []
        for t in (payload.get("data") or []):
            t_new = dict(t)
            if "marker" in t_new:
                m = dict(t_new["marker"])
                m["size"] = max(4, int(m.get("size", 5) * 0.9))
                t_new["marker"] = m
            data.append(t_new)
        return pio.to_image(
            {"data": data, "layout": layout},
            format="png", width=w, height=h, scale=2,
        )
    except Exception:
        return None


def _get_compact_thumb_png(dataset_id: str, subject_id: int,
                           w: int = None, h: int = None) -> bytes | None:
    """issue_table 용 압축 썸네일 PNG. fail_pngs/<sid>_c.png 에 캐시.

    우선순위:
      1. cairosvg ← compact SVG (libcairo 있을 때만)
      2. kaleido  ← compact layout (plotly, 항상 동작)
      3. 일반 _get_thumb_png (최후 폴백)
    """
    w = w or THUMB_W_PX
    h = h or THUMB_H_PX
    cache_dir  = DATASETS_DIR / dataset_id / "fail_pngs"
    cache_path = cache_dir / f"{subject_id}_c.png"

    use_cache = (w == THUMB_W_PX and h == THUMB_H_PX)
    if use_cache:
        with _png_lock:
            if cache_path.exists():
                return cache_path.read_bytes()

    png = None

    # 1. cairosvg + compact SVG
    svg_text = _build_compact_svg_text(dataset_id, subject_id)
    if svg_text:
        png = _try_cairosvg_bytes(svg_text, w, h)

    # 2. kaleido + compact layout
    if png is None:
        png = _try_kaleido_compact(dataset_id, subject_id, w, h)

    # 3. 마지막 폴백
    if png is None:
        png = _get_thumb_png(dataset_id, subject_id, w, h)

    if png is not None and use_cache:
        with _png_lock:
            cache_dir.mkdir(parents=True, exist_ok=True)
            cache_path.write_bytes(png)
    return png


def _try_pil_resize(png_bytes: bytes, w: int, h: int) -> bytes | None:
    try:
        from PIL import Image as PILImage
        img = PILImage.open(BytesIO(png_bytes)).convert("RGB")
        img = img.resize((w, h), PILImage.LANCZOS)
        buf = BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return None


def _try_kaleido(dataset_id: str, subject_id: int, w: int, h: int) -> bytes | None:
    """Render chart JSON → PNG via kaleido (already a project dependency)."""
    chart_path = DATASETS_DIR / dataset_id / "charts" / f"{subject_id}.json"
    if not chart_path.exists():
        return None
    try:
        import plotly.io as pio
        payload = json.loads(chart_path.read_text(encoding="utf-8"))
        raw = pio.to_image(
            {"data": payload["data"], "layout": payload["layout"]},
            format="png", width=w, height=h, scale=1,
        )
        return raw
    except Exception:
        return None


def _get_thumb_png(dataset_id: str, subject_id: int,
                   w: int = THUMB_W_PX, h: int = THUMB_H_PX,
                   allow_kaleido: bool = True) -> bytes | None:
    """
    Return PNG bytes for a subject thumbnail.
    Priority:
      1. Disk cache        (fail_pngs/<sid>_s.png  for default size)
      2. cairosvg          from thumbs/<sid>.svg    (needs cairo system lib)
      3. PIL resize        of any pre-cached kaleido PNG in fail_pngs/<sid>.png
      4. kaleido           from charts/<sid>.json   (always works, cached on disk)
    Results at default size are cached to fail_pngs/<sid>_s.png.
    Pass allow_kaleido=False to skip step 4 (e.g. for large batch operations).
    """
    use_cache  = (w == THUMB_W_PX and h == THUMB_H_PX)
    cache_dir  = DATASETS_DIR / dataset_id / "fail_pngs"
    cache_path = cache_dir / f"{subject_id}_s.png"

    if use_cache:
        with _png_lock:
            if cache_path.exists():
                return cache_path.read_bytes()

    svg_path = DATASETS_DIR / dataset_id / "thumbs"    / f"{subject_id}.svg"
    full_png = DATASETS_DIR / dataset_id / "fail_pngs" / f"{subject_id}.png"

    png: bytes | None = None

    # cairosvg — fast, lossless; requires libcairo system library
    if svg_path.exists():
        png = _try_cairosvg(svg_path, w, h)

    # PIL resize of an already-cached full-size kaleido PNG
    if png is None and full_png.exists():
        png = _try_pil_resize(full_png.read_bytes(), w, h)

    # kaleido from chart JSON — reliable fallback, result is cached so cost is once-per-subject
    if png is None and allow_kaleido:
        png = _try_kaleido(dataset_id, subject_id, w, h)

    if png is not None and use_cache:
        with _png_lock:
            cache_dir.mkdir(parents=True, exist_ok=True)
            cache_path.write_bytes(png)

    return png


def _add_image(ws, png_bytes: bytes, anchor: str, w_px: int, h_px: int):
    if not png_bytes:
        return
    try:
        img = XLImage(BytesIO(png_bytes))
        img.width  = w_px
        img.height = h_px
        ws.add_image(img, anchor)
    except Exception:
        pass


# ── Sheet 1: summary ────────────────────────────────────────────────────────────

def _sheet_summary(wb, dataset_id, meta, yield_rows, fail_items,
                   summary_comments, eval_data):
    ws = wb.create_sheet("summary")
    ws.sheet_view.showGridLines = False

    sources  = meta.get("sources")  or []
    subjects = meta.get("subjects") or []

    pass_row     = next((r for r in yield_rows if str(r.get("bin",""))=="1"), {})
    avg_val      = pass_row.get("avg")
    pass_yield   = f"{avg_val:.2f}%" if avg_val is not None else "-"
    non_pass     = [r for r in (fail_items.get("rows") or [])
                    if str(r.get("bin","")) != "1"]

    ROW = 1

    # Title
    _c(ws, ROW, 1, f"Plotly Data Dashboard — {dataset_id}", font=_FONT_TITLE, border=None)
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    # Web link — summary/dashboard
    dash_url = f"{SERVER_BASE_URL}/dash/{dataset_id}"
    lnk = _c(ws, ROW, 1, "🌐  대시보드 웹으로 열기 (서버 실행 중일 때 클릭)", font=_FONT_LINK, align=_ALIGN_L, border=None)
    lnk.hyperlink = dash_url
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws.row_dimensions[ROW].height = 22
    ROW += 1

    _blank_row(ws, ROW); ROW += 1

    # ── Feature ──────────────────────────────────────────────────────────────
    _section_row(ws, ROW, "Feature", span=6); ROW += 1

    feat_headers = ["Dataset ID", "Total DUT", "Pass (type 1)", "Fail Types", "Sources", "Subjects"]
    feat_widths  = [22, 10, 14, 12, 10, 10]
    _header_row(ws, ROW, feat_headers, feat_widths); ROW += 1

    feat_vals = [
        dataset_id,
        meta.get("row_count", "-"),
        pass_row.get("count", "-"),
        len(non_pass),
        len(sources),
        len(subjects),
    ]
    for i, v in enumerate(feat_vals, 1):
        _c(ws, ROW, i, v, align=_ALIGN_C)
    ws.row_dimensions[ROW].height = 18; ROW += 1

    _blank_row(ws, ROW); ROW += 1

    # ── Yield Summary ─────────────────────────────────────────────────────────
    _section_row(ws, ROW, "Yield Summary", span=6); ROW += 1

    _c(ws, ROW, 1,
       f"Overall Pass Yield (Bin 1):  {pass_yield}",
       font=Font(bold=True, size=12, color="1F4D8C"), align=_ALIGN_L, border=None)
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws.row_dimensions[ROW].height = 22; ROW += 1

    _c(ws, ROW, 1, "Major Fail Bins (top 5)",
       font=Font(bold=True, size=11, color="444444"), align=_ALIGN_L, border=None)
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws.row_dimensions[ROW].height = 18; ROW += 1

    ys_headers = ["Rank", "Fail Type", "Main Fail Subject", "Fail Ratio", "Comment"]
    ys_widths  = [6, 10, 34, 12, 40]
    _header_row(ws, ROW, ys_headers, ys_widths); ROW += 1

    for rank, r in enumerate(non_pass[:5], 1):
        st           = str(r.get("bin",""))
        fail_subs    = r.get("fail_subjects") or []
        main_fail    = fail_subs[0].get("subject","N/A") if fail_subs else "N/A"
        avg_r        = r.get("avg")
        fail_ratio   = f"{avg_r:.2f}%" if avg_r is not None else "-"
        comment      = summary_comments.get(st, "")
        row_data     = [rank, st, main_fail, fail_ratio, comment]
        aligns       = [_ALIGN_C, _ALIGN_C, _ALIGN_L, _ALIGN_C, _ALIGN_LT]
        fills        = [None, None, None, None, _FILL_COMMENT]
        for i, (v, al, fi) in enumerate(zip(row_data, aligns, fills), 1):
            _c(ws, ROW, i, v, align=al, fill=fi)
        ws.row_dimensions[ROW].height = 18; ROW += 1

    _blank_row(ws, ROW); ROW += 1

    # ── Evaluation Summary ────────────────────────────────────────────────────
    _section_row(ws, ROW, "Evaluation Summary", span=2); ROW += 1
    _header_row(ws, ROW, ["Category", "Result"], [14, 52]); ROW += 1

    for cat, key in [("Yield","yield"), ("CPK","cpk"), ("Temp","temp"), ("ETC","etc")]:
        _c(ws, ROW, 1, cat, font=_FONT_BOLD, fill=_FILL_HDR, align=_ALIGN_C)
        _c(ws, ROW, 2, eval_data.get(key,""), fill=_FILL_COMMENT, align=_ALIGN_LT)
        ws.row_dimensions[ROW].height = 18; ROW += 1

    # column A width
    ws.column_dimensions["A"].width = 22


# ── Sheet 2: yield ──────────────────────────────────────────────────────────────

def _sheet_yield(wb, yield_rows, sources, yield_comments):
    ws = wb.create_sheet("yield")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["bin", "count"]
    widths  = [12, 8]
    for src in sources:
        headers.append(str(src))
        widths.append(10)
    headers += ["avg", "Main Fail Subject", "comment"]
    widths  += [10, 30, 40]

    _header_row(ws, 1, headers, widths)

    merged = []
    for row in yield_rows:
        row = dict(row)
        key = str(row.get("bin",""))
        row["comment"] = yield_comments.get(key, row.get("comment","") or "")
        merged.append(row)
    merged.sort(key=_yield_sort_key)

    for r_idx, row in enumerate(merged, 2):
        is_pass = str(row.get("bin","")) == "1"
        bg = _FILL_PASS if is_pass else None

        col = 1
        _c(ws, r_idx, col, row.get("bin"), fill=bg, align=_ALIGN_C); col += 1
        _c(ws, r_idx, col, row.get("count"),         fill=bg, align=_ALIGN_C); col += 1
        for src in sources:
            _c(ws, r_idx, col, row.get(f"portion_{src}"), fill=bg, align=_ALIGN_C, num_fmt="0.00")
            col += 1
        _c(ws, r_idx, col, row.get("avg"),              fill=_FILL_TOTAL, align=_ALIGN_C, num_fmt="0.00"); col += 1
        _c(ws, r_idx, col, row.get("Main Fail subject"), fill=bg,          align=_ALIGN_L); col += 1
        _c(ws, r_idx, col, row.get("comment",""),         fill=_FILL_COMMENT, align=_ALIGN_LT)
        ws.row_dimensions[r_idx].height = 18


# ── Sheet 3: cpk ────────────────────────────────────────────────────────────────

def _sheet_cpk(wb, cpk_rows, cpk_comments):
    ws = wb.create_sheet("cpk")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["subject", "lower_limit", "upper_limit", "units", "source",
               "min", "median", "max", "average", "stdev",
               "cpl", "cpu", "cp", "cpk", "comment"]
    widths  = [28, 9, 9, 8, 14, 9, 9, 9, 9, 9, 9, 9, 9, 9, 40]
    _header_row(ws, 1, headers, widths)

    for r in cpk_rows:
        r["comment"] = cpk_comments.get(
            _cpk_comment_key(r.get("subject"), r.get("source")), "")
    rows = _merge_cpk_subject(cpk_rows)

    numeric_cols = {"lower_limit","upper_limit","min","median","max","average","stdev","cpl","cpu","cp","cpk"}
    center_cols  = numeric_cols | {"units","source"}

    for r_idx, row in enumerate(rows, 2):
        is_total = row.get("source") == "total"
        is_new   = bool(row.get("subject"))  # non-empty → new subject group

        try:
            cpk_v = float(row.get("cpk",""))
            low   = cpk_v < 1.33
        except (TypeError, ValueError):
            low = False

        top_side = _MEDIUM if is_new else _THIN
        grp_border = Border(left=_THIN, right=_THIN, bottom=_THIN, top=top_side)

        for i, key in enumerate(headers, 1):
            v    = row.get(key, "")
            fill = (_FILL_COMMENT if key == "comment"
                    else _FILL_LOW_CPK if (key == "cpk" and low)
                    else _FILL_TOTAL   if is_total
                    else None)
            font = Font(bold=True, color="5C4400") if (key == "cpk" and low) else None
            al   = _ALIGN_C if key in center_cols else _ALIGN_L
            nf   = "0.000" if key in numeric_cols else None
            c    = _c(ws, r_idx, i, v, fill=fill, align=al, border=grp_border, num_fmt=nf)
            if font:
                c.font = font

        ws.row_dimensions[r_idx].height = 16


# ── Sheet 4: fail_data ──────────────────────────────────────────────────────────

MAX_FAIL_SUBS = 5   # how many fail subjects to show per row


def _sheet_fail_data(wb, dataset_id, fail_items=None):
    """Fail Item 은 그림 수가 많아 SVG→PNG 변환 비용이 크므로 sheet 에는 링크만 둔다.
    클릭하면 세션의 Fail Item 탭으로 이동해서 모든 fail subject 썸네일을 확인 가능.

    fail_items 인자는 시그니처 호환을 위해 받지만 더 이상 사용하지 않는다.
    """
    _sheet_link_only(wb, "fail_data", dataset_id, "Fail Item")


# ── Sheet 5: fail_values ────────────────────────────────────────────────────────

_FILL_LO = PatternFill("solid", fgColor="DBEAFE")   # blue-100
_FILL_HI = PatternFill("solid", fgColor="FEE2E2")   # red-100
_FONT_LO = Font(bold=True, color="1E40AF")           # blue-800
_FONT_HI = Font(bold=True, color="991B1B")           # red-800


def _sheet_fail_values(wb, dataset_id: str):
    ws = wb.create_sheet("fail_values")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Source (Sheet)", "DUT", "XCoord", "YCoord", "Bin",
               "Subject", "Value", "Lower Limit", "Upper Limit", "Fail"]
    ids     = ["source", "dut", "x_coord", "y_coord", "bin",
               "subject", "value", "lower_limit", "upper_limit", "fail"]
    widths  = [22, 8, 8, 8, 8, 28, 12, 12, 12, 9]

    _header_row(ws, 1, headers, widths)

    try:
        rows = get_fail_values(dataset_id)
    except Exception:
        rows = []

    if not rows:
        _c(ws, 2, 1, "데이터 없음 (non-pass DUT가 없거나 fail 기준 미설정)",
           font=Font(size=11, color="888888"), border=None)
        return

    center_cols = {"dut", "x_coord", "y_coord", "bin", "value",
                   "lower_limit", "upper_limit", "fail"}

    for r_idx, row in enumerate(rows, 2):
        is_lo  = row.get("fail") == "< lo"
        is_hi  = row.get("fail") == "> hi"
        v_fill = _FILL_LO if is_lo else (_FILL_HI if is_hi else None)
        v_font = _FONT_LO if is_lo else (_FONT_HI if is_hi else None)
        f_fill = v_fill
        f_font = v_font

        for col_idx, key in enumerate(ids, 1):
            val  = row.get(key, "")
            al   = _ALIGN_C if key in center_cols else _ALIGN_L
            fill = (v_fill if key in ("value", "fail") else None)
            c    = _c(ws, r_idx, col_idx, val, align=al, fill=fill)
            if key in ("value", "fail") and (v_font is not None):
                c.font = v_font

        ws.row_dimensions[r_idx].height = 16

    ws.auto_filter.ref = f"A1:{get_column_letter(len(ids))}1"


# ── Sheet 6: issue_table ────────────────────────────────────────────────────────

def _sheet_issue_table(wb, dataset_id, fail_items, sources, issue_comments):
    ws = wb.create_sheet("issue_table")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["bin", "subject", "average"]
    widths  = [12, 26, 10]
    for src in sources:
        headers.append(str(src))
        widths.append(10)
    headers += ["Distribution", "Issue Point", "Comment", "개발팀 1차 Comment", "PTE 1차 comment"]
    widths  += [THUMB_COL_W, 32, 32, 32, 32]

    _header_row(ws, 1, headers, widths)

    rows = _build_issue_rows(fail_items, sources, issue_comments)
    rows.sort(key=_yield_sort_key)

    dist_col = 4 + len(sources)  # 1-based column index of "Distribution"

    for r_idx, row in enumerate(rows, 2):
        st      = str(row.get("bin",""))
        is_pass = st == "1"

        _c(ws, r_idx, 1, st,               align=_ALIGN_C)
        _c(ws, r_idx, 2, row.get("subject",""), align=_ALIGN_L)
        _c(ws, r_idx, 3, row.get("avg"),   align=_ALIGN_C, fill=_FILL_TOTAL, num_fmt="0.0000")

        col = 4
        for src in sources:
            _c(ws, r_idx, col, row.get(f"portion_{src}"), align=_ALIGN_C, num_fmt="0.00")
            col += 1

        # Distribution cell (image will overlay)
        _c(ws, r_idx, dist_col, "", align=_ALIGN_C)

        # Comment fields
        for off, key in enumerate(
            ["issue_point","issue_comment","dev_comment","pte_comment"], 1
        ):
            _c(ws, r_idx, dist_col + off, row.get(key,""),
               fill=_FILL_COMMENT, align=_ALIGN_LT)

        if is_pass:
            ws.row_dimensions[r_idx].height = 20
            continue

        ws.row_dimensions[r_idx].height = THUMB_ROW_H
        sid = row.get("subject_id")
        # 산포 그림 1장은 무조건 노출되도록 보강.
        # 1) 정상 변환 시 PNG 임베드
        # 2) 변환 실패 시 셀에 thumbnail 직접 URL 하이퍼링크 + 안내 텍스트
        if sid is not None:
            # issue_table 은 산포가 잘 보이도록 compact SVG 로 별도 렌더링
            png = _get_compact_thumb_png(dataset_id, sid)
            if png:
                _add_image(ws, png, f"{get_column_letter(dist_col)}{r_idx}",
                           THUMB_W_PX, THUMB_H_PX)
            else:
                fallback_url = f"{SERVER_BASE_URL}/api/{dataset_id}/thumb/{sid}"
                cell = _c(
                    ws, r_idx, dist_col,
                    f"산포 보기 (subject {sid})",
                    align=_ALIGN_C, border=_BORDER,
                )
                cell.hyperlink = fallback_url
                cell.font = Font(size=10, color="2369B3", underline="single", bold=True)
        else:
            # subject_id 자체가 없는 경우 (예: 단일 fail subject 없음) — 세션 페이지 링크.
            cell = _c(
                ws, r_idx, dist_col,
                "세션에서 산포 보기",
                align=_ALIGN_C, border=_BORDER,
            )
            cell.hyperlink = f"{SERVER_BASE_URL}/dash/{dataset_id}"
            cell.font = Font(size=10, color="2369B3", underline="single")


# ── Sheet 6 / 7: distribution & histogram (link-only) ──────────────────────────
#
# 그래프 자체는 워크북에 박지 않고 세션 페이지(`/dash/<dataset_id>`) 로 연결.
# 그 페이지에서 Summary~Histogram 까지의 모든 탭에 접근할 수 있다.

def _sheet_link_only(wb, sheet_name: str, dataset_id: str, tab_label: str):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    dash_url = f"{SERVER_BASE_URL}/dash/{dataset_id}"

    _c(ws, 1, 1, f"{tab_label} — {dataset_id}",
       font=_FONT_TITLE, border=None)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 26

    lnk = _c(ws, 3, 1,
             f"🌐  세션 페이지 열기 (전체 탭: Summary → Histogram)",
             font=_FONT_LINK, align=_ALIGN_L, border=None)
    lnk.hyperlink = dash_url
    ws.merge_cells("A3:F3")
    ws.row_dimensions[3].height = 24

    _c(ws, 5, 1, dash_url,
       font=Font(size=10, color="666666"), align=_ALIGN_L, border=None)
    ws.merge_cells("A5:F5")
    ws.row_dimensions[5].height = 16

    _c(ws, 7, 1,
       f"이 시트에는 {tab_label} 그래프를 직접 포함하지 않습니다. "
       "위 링크를 클릭하면 서버에서 동일 세션의 모든 탭(Summary, Yield, CPK, "
       "Fail Item, Issue Table, Distribution, Histogram)이 열립니다.",
       font=Font(size=10, color="888888"), align=_ALIGN_LT, border=None)
    ws.merge_cells("A7:F7")
    ws.row_dimensions[7].height = 48

    ws.column_dimensions["A"].width = 80


def _sheet_distribution(wb, dataset_id):
    _sheet_link_only(wb, "distribution", dataset_id, "Distribution")


def _sheet_histogram(wb, dataset_id):
    _sheet_link_only(wb, "histogram", dataset_id, "Histogram")


# ── RAW data sheets (옵션) ─────────────────────────────────────────────────────

def _append_raw_sheets(wb, dataset_id: str, progress_cb=None):
    """build_raw_xlsx 와 동일한 데이터를 같은 workbook 에 raw_<source> 시트로 추가."""
    from analysis.data_loader import load_table
    from analysis.table_builder import _subject_columns, _fmt_type
    import pandas as pd

    input_dir = DATASETS_DIR / dataset_id / "input"
    if not input_dir.exists():
        return

    csv_paths = sorted(input_dir.glob("*.csv"))
    if not csv_paths:
        return

    used = set(wb.sheetnames)
    total = len(csv_paths)
    for idx, p in enumerate(csv_paths, 1):
        source_name = p.stem
        table = load_table(p)
        meta = table.meta.reset_index(drop=True).copy()
        scores = table.scores.reset_index(drop=True).copy()
        scores.columns = _subject_columns(table)
        frame = pd.concat([meta, scores], axis=1)
        frame["Bin"] = frame["Bin"].map(_fmt_type)

        raw_name = f"raw_{source_name}"[:31] or f"raw_{idx}"
        base = raw_name
        i = 1
        while raw_name in used:
            suffix = f"_{i}"
            raw_name = (base[: 31 - len(suffix)] + suffix)
            i += 1
        used.add(raw_name)

        ws = wb.create_sheet(raw_name)
        ws.append([str(c) for c in frame.columns])
        for row_vals in frame.itertuples(index=False, name=None):
            ws.append([
                (None if (isinstance(v, float) and pd.isna(v)) else v)
                for v in row_vals
            ])

        if progress_cb:
            try:
                progress_cb(idx, total, source_name)
            except Exception:
                pass


# ── Public entry point ──────────────────────────────────────────────────────────

def build_report_xlsx(dataset_id: str, *, include_raw: bool = False,
                       progress_cb=None) -> bytes:
    """Report XLSX 생성.

    Args:
      include_raw: True 면 원본 CSV 들을 raw_<source> 시트로 추가.
      progress_cb: callable(percent:int, stage:str) — 진행률 알림용.
    """
    def _emit(pct, stage):
        if progress_cb is None:
            return
        try:
            progress_cb(max(0, min(100, int(pct))), stage)
        except Exception:
            pass

    _emit(2, "테이블 로딩")
    meta     = read_table_json(dataset_id, "meta")       or {}
    yield_r  = read_table_json(dataset_id, "yield")      or []
    cpk_r    = read_table_json(dataset_id, "cpk")        or []
    fail_i   = read_table_json(dataset_id, "fail_items") or {"rows": []}

    sources  = meta.get("sources")  or []

    _emit(6, "주석 로딩")
    ycomm  = _read_yield_comments(dataset_id)
    ccomm  = _read_cpk_comments(dataset_id)
    icomm  = _read_issue_comments(dataset_id)
    scomm  = _read_summary_comments(dataset_id)
    edata  = _read_summary_eval(dataset_id)

    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    _emit(10, "Summary 시트 작성 중")
    _sheet_summary(wb, dataset_id, meta, yield_r, fail_i, scomm, edata)

    _emit(20, "Yield 시트 작성 중")
    _sheet_yield(wb, yield_r, sources, ycomm)

    _emit(30, "CPK 시트 작성 중")
    _sheet_cpk(wb, cpk_r, ccomm)

    _emit(40, "Fail Item 시트 작성 중 (SVG 썸네일 변환)")
    _sheet_fail_data(wb, dataset_id, fail_i)

    _emit(55, "Fail Values 시트 작성 중")
    _sheet_fail_values(wb, dataset_id)

    _emit(70, "Issue Table 시트 작성 중 (SVG 썸네일 변환)")
    _sheet_issue_table(wb, dataset_id, fail_i, sources, icomm)

    _emit(80, "Distribution 시트 (링크)")
    _sheet_distribution(wb, dataset_id)

    _emit(82, "Histogram 시트 (링크)")
    _sheet_histogram(wb, dataset_id)

    if include_raw:
        _emit(85, "RAW 데이터 시트 추가 중")
        def _raw_cb(idx, total, name):
            base = 85
            span = 10  # 85 → 95
            pct = base + (idx / max(1, total)) * span
            _emit(pct, f"RAW 시트 {idx}/{total}: {name}")
        _append_raw_sheets(wb, dataset_id, progress_cb=_raw_cb)

    _emit(96, "파일 저장 중")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    _emit(100, "완료")
    return buf.getvalue()
