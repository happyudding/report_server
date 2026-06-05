"""Honey 클라이언트 (PyQt5).

UI 레이아웃은 .ui (Qt Designer 편집 가능) 에 정의, 런타임에 uic.loadUi 로 로드.
- honey_main.ui   : 메인 화면 (d1_storage 검색 → 분석 → 자동 저장 → 업로드)
- upload_dialog.ui: 서버 업로드용 메타(Product Type 라디오/Product/LOT/Revision/PW) 팝업
- d1_browser.ui   : d1_storage(가상 서버 스토리지) 파일 검색/선택 팝업

워크플로우: d1_storage 에서 CSV 검색·선택 → 출력 시트 선택 → '분석 실행' 시
입력 폴더에 xlsx 자동 저장(xlwings) → '서버에 업로드' 클릭 시 메타 팝업 입력 후 전송.
"""
import concurrent.futures
import contextlib
import datetime
import os
import queue
import re
import sys
import tempfile
import time
import traceback
import zipfile
from pathlib import Path

import requests

from PyQt5 import uic
from PyQt5.QtCore import Qt, QTimer, QEvent
from PyQt5.QtGui import QColor, QFont, QIntValidator
from PyQt5.QtWidgets import (
    QAbstractItemView, QApplication, QColorDialog, QDialog, QDialogButtonBox, QFileDialog,
    QGridLayout, QHBoxLayout, QHeaderView, QInputDialog, QLabel,
    QListWidgetItem, QMainWindow, QMessageBox, QPushButton,
    QTableWidgetItem, QVBoxLayout,
)

from config import D1_STORAGE_DIR
from transport.config import CURRENT_VERSION, SERVER_BASE_URL
from transport import chart_export, updater, uploader, version_check
import app_settings
import chart_colors

# 로컬 리포트 엔진 (pandas/xlwings 의존). 미설치 시 화면 비활성.
try:
    import report_generator as rg
    from report_generator import xlsx_writer
    _RG_IMPORT_ERROR = None
except Exception as exc:  # noqa: BLE001
    rg = None
    xlsx_writer = None
    _RG_IMPORT_ERROR = exc

SHEET_OPTIONS = ["summary", "yield", "cpk", "fail_item", "issue_table", "distribution"]
PRODUCT_TYPES = ["MD", "PD", "PM", "SE"]
_FLOW_PROFILE_ON = bool(os.environ.get("HONEY_FLOW_PROFILE"))
# 파일명 끝 시간 접미사 패턴 (_YYMMDD_HHMM) — 중복 부착 방지용
_TS_RE = re.compile(r"_\d{6}_\d{4}$")

# 프리징(onedir) 시 _MEIPASS, 아니면 스크립트 폴더에서 .ui 탐색
_BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
UI_PATH = os.path.join(_BASE_DIR, "honey_main.ui")
UPLOAD_UI_PATH = os.path.join(_BASE_DIR, "upload_dialog.ui")
D1_UI_PATH = os.path.join(_BASE_DIR, "d1_browser.ui")
ORDER_UI_PATH = os.path.join(_BASE_DIR, "file_order.ui")
SETTINGS_UI_PATH = os.path.join(_BASE_DIR, "report_settings.ui")


@contextlib.contextmanager
def _flow_time(label):
    if not _FLOW_PROFILE_ON:
        yield
        return
    t0 = time.perf_counter()
    try:
        yield
    finally:
        elapsed = time.perf_counter() - t0
        print(f"[flow-profile] honey_main.{label}: {elapsed:.3f}s", file=sys.stderr, flush=True)


def _validate_meta(product, lot_id, password):
    """공통 메타/PIN 검증. 문제 메시지(str) 반환, 정상이면 None.
    password 는 빈 문자열 허용 — 미설정 시 웹에서 비밀번호 없이 수정/삭제 가능.
    """
    if not product or not lot_id:
        return "Product 와 LOT ID 를 모두 입력하세요."
    if password and (len(password) != 4 or not password.isdigit()):
        return "비밀번호는 숫자 4자리 또는 빈칸(미설정)으로 입력하세요."
    return None


class _ElapsedProgress:
    """메인 UI Status progress bar 의 elapsed 시간을 계속 갱신한다."""

    def __init__(self, bar, label, status_cb=None, busy=True, minimum=0, maximum=100):
        self.bar = bar
        self.status_cb = status_cb
        self.busy = busy
        self.started = time.monotonic()
        self.label = label
        self.status = None
        self._last_secs = -1
        self._last_rendered = None
        self.token = int(self.bar.property("_honey_progress_token") or 0) + 1
        self.bar.setProperty("_honey_progress_token", self.token)
        self.bar.setRange(0, 0) if busy and maximum == 0 else self.bar.setRange(minimum, maximum)
        self.bar.setValue(minimum)
        self.bar.setFormat("")
        self.bar.show()
        self.update(force=True)

    def _elapsed(self):
        secs = int(time.monotonic() - self.started)
        return secs, f"{secs // 60:02d}:{secs % 60:02d}"

    def set(self, label=None, value=None, status=None, busy=None):
        if label is not None:
            self.label = label
        if busy is not None:
            self.busy = busy
        if value is not None:
            self.bar.setValue(value)
        if status is not None:
            self.status = status
            if self.status_cb is not None:
                self.status_cb(status)
        self.update(force=True)

    def value(self):
        return self.bar.value()

    def maximum(self):
        return self.bar.maximum()

    def set_maximum(self, value):
        self.bar.setMaximum(value)

    def update(self, force=False):
        secs, elapsed = self._elapsed()
        if not force and secs == self._last_secs:
            return
        suffix = " (진행중)" if self.busy else ""
        text = f"{self.label}  [{elapsed}]{suffix}"
        if force or text != self._last_rendered:
            self.bar.setFormat(text)
            self._last_rendered = text
        self._last_secs = secs
        QApplication.processEvents()

    def success(self, text, value=None, hide_ms=5000):
        was_indeterminate = self.bar.minimum() == 0 and self.bar.maximum() == 0
        if value is None:
            value = 100 if was_indeterminate else self.bar.maximum()
        if was_indeterminate:
            self.bar.setRange(0, 100)
        self.busy = False
        self.bar.setValue(value)
        self.label = text
        self.bar.setFormat(text)
        if self.status_cb is not None:
            self.status_cb(text)
        self._hide_later(hide_ms)
        QApplication.processEvents()

    def fail(self, text, hide_ms=8000):
        if self.bar.minimum() == 0 and self.bar.maximum() == 0:
            self.bar.setRange(0, 100)
        self.busy = False
        self.bar.setValue(0)
        self.label = text
        self.bar.setFormat(text)
        if self.status_cb is not None:
            self.status_cb(text)
        self._hide_later(hide_ms)
        QApplication.processEvents()

    def _hide_later(self, ms):
        token = self.token

        def _hide_if_current():
            if int(self.bar.property("_honey_progress_token") or 0) != token:
                return
            self.bar.hide()
            self.bar.setFormat("")

        QTimer.singleShot(ms, _hide_if_current)


def _init_com_for_worker():
    """Excel COM/xlwings 를 워커 스레드에서 쓸 수 있으면 초기화한다."""
    try:
        import pythoncom  # type: ignore
    except Exception:
        return None
    try:
        pythoncom.CoInitialize()
        return pythoncom
    except Exception:
        return None


def _co_uninitialize(com_module):
    if com_module is None:
        return
    try:
        com_module.CoUninitialize()
    except Exception:
        pass


def _wait_for_future(future, progress, poll_cb=None, timeout=0.1):
    while True:
        if poll_cb is not None:
            poll_cb()
        progress.update()
        done, _ = concurrent.futures.wait([future], timeout=timeout)
        if done:
            if poll_cb is not None:
                poll_cb()
            return future.result()


def _timestamp():
    """파일명용 현재 시각: 260601_0949 (YYMMDD_HHMM)."""
    return datetime.datetime.now().strftime("%y%m%d_%H%M")


def _suggest_base_name(csv_paths, group=None):
    """저장 파일명 base 를 **첫 입력 파일의 FileName** 기준으로 생성.

    group 이 있으면 통일된 첫 FileName(group.names()[0], 설정 rename 반영분)을,
    없으면(로드 직후 프리뷰) 첫 파일 stem 을 쓴다. 여러 파일명 join 으로 길어지는
    것을 피하고 재계산/지연 없이 단순하게 첫 파일 이름만 사용한다.
    """
    if group is not None and group.names():
        base = group.names()[0]
    elif csv_paths:
        base = Path(csv_paths[0]).stem
    else:
        base = "report"
    base = base.strip(" _-") or "report"
    return f"{base}_report_{_timestamp()}"


def _build_output_path(out_dir, base):
    """base 이름으로 최종 저장 경로 생성. 시간 접미사가 없으면 현재 시각을 붙인다."""
    base = base.strip()
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    base = base.strip(" _-") or "report"
    if not _TS_RE.search(base):
        base = f"{base}_{_timestamp()}"
    return str(Path(out_dir) / f"{base}.xlsx")


# ───────────────────────────────────────────────────────────────────────────
# 서버 업로드 메타 입력 팝업 (Product Type 라디오 상단 + Product/LOT/Revision/PW)

class UploadDialog(QDialog):
    def __init__(self, parent=None, defaults=None):
        super().__init__(parent)
        uic.loadUi(UPLOAD_UI_PATH, self)
        # validator 제거 — 빈칸(미설정)도 허용. 검증은 _on_ok 에서 수행.
        self._pt_radios = {
            "MD": self.rb_pt_MD, "PD": self.rb_pt_PD,
            "PM": self.rb_pt_PM, "SE": self.rb_pt_SE,
        }
        self.buttonBox.accepted.connect(self._on_ok)
        self.buttonBox.rejected.connect(self.reject)
        if defaults:
            self._pt_radios.get(defaults.get("product_type", "MD"),
                                self.rb_pt_MD).setChecked(True)
            self.le_product.setText(defaults.get("product", ""))
            self.le_lot_id.setText(defaults.get("lot_id", ""))
            self.le_revision.setText(defaults.get("revision", ""))

    def product_type(self):
        for key, rb in self._pt_radios.items():
            if rb.isChecked():
                return key
        return "MD"

    def _on_ok(self):
        err = _validate_meta(self.le_product.text().strip(),
                             self.le_lot_id.text().strip(),
                             self.le_password.text().strip())
        if err:
            QMessageBox.warning(self, "입력 오류", err)
            return
        self.accept()

    def values(self):
        return {
            "product_type": self.product_type(),
            "product": self.le_product.text().strip(),
            "lot_id": self.le_lot_id.text().strip(),
            "revision": self.le_revision.text().strip(),
            "password": self.le_password.text().strip(),
        }


# ───────────────────────────────────────────────────────────────────────────
# distribution 차트 색 편집 팝업 (8x6 = 48색, 클릭 시 팔레트로 변경)

def _is_light(hex_color):
    s = str(hex_color).lstrip("#")
    try:
        r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
    except Exception:
        return True
    return (0.299 * r + 0.587 * g + 0.114 * b) > 150


class ColorEditorDialog(QDialog):
    COLS, ROWS = 8, 6

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Chart 색 편집 (Legend 1~48)")
        self._colors = chart_colors.load_colors()

        root = QVBoxLayout(self)
        info = QLabel(
            "각 색을 클릭하면 팔레트가 열립니다. 번호는 distribution Legend(소스) 순서와 같습니다.")
        info.setWordWrap(True)
        info.setStyleSheet("color:#555;")
        root.addWidget(info)

        grid = QGridLayout()
        grid.setSpacing(6)
        self._btns = []
        for i in range(chart_colors.N_COLORS):
            b = QPushButton(str(i + 1))
            b.setFixedSize(60, 40)
            b.clicked.connect(lambda _c, idx=i: self._pick(idx))
            self._btns.append(b)
            grid.addWidget(b, i // self.COLS, i % self.COLS)
        root.addLayout(grid)

        row = QHBoxLayout()
        btn_reset = QPushButton("기본값 복원")
        btn_reset.clicked.connect(self._reset)
        row.addWidget(btn_reset)
        row.addStretch(1)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.accepted.connect(self._on_ok)
        bb.rejected.connect(self.reject)
        row.addWidget(bb)
        root.addLayout(row)

        self._refresh()

    def _refresh(self):
        for i, b in enumerate(self._btns):
            c = self._colors[i]
            fg = "#000" if _is_light(c) else "#fff"
            b.setStyleSheet(
                f"background-color:{c}; color:{fg}; font-weight:600;"
                "border:1px solid #999; border-radius:4px;")

    def _pick(self, idx):
        col = QColorDialog.getColor(QColor(self._colors[idx]), self,
                                    f"{idx + 1}번 색 선택")
        if col.isValid():
            self._colors[idx] = col.name().upper()
            self._refresh()

    def _reset(self):
        self._colors = chart_colors.generate_default_colors()
        self._refresh()

    def _on_ok(self):
        try:
            chart_colors.save_colors(self._colors)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "저장 실패", f"색 저장에 실패했습니다:\n{exc}")
            return
        self.accept()


# ───────────────────────────────────────────────────────────────────────────
# 파일 순서 지정 팝업 (열기/검색 후, 메인 창에 로드하기 전에 순서 확정)

class FileOrderDialog(QDialog):
    def __init__(self, parent, paths):
        super().__init__(parent)
        uic.loadUi(ORDER_UI_PATH, self)
        for p in paths:
            it = QListWidgetItem(Path(p).name)
            it.setData(Qt.UserRole, p)
            it.setToolTip(p)
            self.list_order.addItem(it)
        self.list_order.setCurrentRow(0)
        self.btn_up.clicked.connect(lambda: self._move(-1))
        self.btn_down.clicked.connect(lambda: self._move(1))
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)

    def _move(self, delta):
        row = self.list_order.currentRow()
        new = row + delta
        if row < 0 or not (0 <= new < self.list_order.count()):
            return
        it = self.list_order.takeItem(row)
        self.list_order.insertItem(new, it)
        self.list_order.setCurrentRow(new)

    def ordered_paths(self):
        return [self.list_order.item(i).data(Qt.UserRole)
                for i in range(self.list_order.count())]


# ───────────────────────────────────────────────────────────────────────────
# d1_storage 파일 검색/선택 팝업

class D1BrowserDialog(QDialog):
    def __init__(self, parent, storage_dir):
        super().__init__(parent)
        uic.loadUi(D1_UI_PATH, self)
        self.storage_dir = storage_dir
        self.lbl_path.setText(f"D1: {storage_dir}")
        # 검색 전에는 비어 있다가, 키워드 입력 후 [검색] 시에만 조회
        self.btn_refresh.clicked.connect(self._search)
        self.le_search.returnPressed.connect(self._search)
        self.list_files.itemDoubleClicked.connect(lambda _i: self.accept())
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        self.le_search.setFocus()

    def _scan(self):
        p = Path(self.storage_dir)
        if not p.exists():
            return []
        files = [f for f in p.rglob("*")
                 if f.is_file() and f.suffix.lower() in (".csv", ".xlsx")]
        return sorted(files, key=lambda x: str(x).lower())

    def _search(self):
        """현재 키워드로 D1 을 조회해 결과를 채운다 (매번 디스크 재스캔)."""
        q = self.le_search.text().strip().lower()
        self.list_files.clear()
        for f in self._scan():
            rel = str(f.relative_to(self.storage_dir))
            if q and q not in rel.lower():
                continue
            it = QListWidgetItem(rel)
            it.setData(Qt.UserRole, str(f))
            self.list_files.addItem(it)
        if self.list_files.count() == 0:
            self.lbl_hint.setText(f"'{self.le_search.text().strip()}' 검색 결과가 없습니다.")
        else:
            self.lbl_hint.setText(
                f"{self.list_files.count()}개 결과 — Ctrl/Shift 로 여러 파일 선택 가능")

    def selected_paths(self):
        return [it.data(Qt.UserRole) for it in self.list_files.selectedItems()]


# ───────────────────────────────────────────────────────────────────────────
# 리포트 설정 팝업 (Start 시 전처리 끝난 그룹으로 열림):
# 분석 항목(Select Items) · 출력 옵션(Option) · 차트 색 · Auto Upload · Confirm

class ReportSettingsDialog(QDialog):
    def __init__(self, parent, group, csv_count):
        super().__init__(parent)
        uic.loadUi(SETTINGS_UI_PATH, self)
        self.group = group
        self.csv_count = csv_count
        # Filename(legend) 사용자 지정값. None 이면 미변경(기존 파일명 사용).
        self._filename_overrides = None
        self.sheet_checks = {
            name: getattr(self, f"cb_sheet_{name}") for name in SHEET_OPTIONS
        }

        # 분석 항목: 좌(제외) ↔ 우(선택) 이동
        self.btn_all_right.clicked.connect(self._move_all_right)
        self.btn_sel_right.clicked.connect(self._move_selected_right)
        self.btn_sel_left.clicked.connect(self._move_selected_left)
        self.btn_all_left.clicked.connect(self._move_all_left)
        self.btn_sel_fail.clicked.connect(self._select_fail_only)
        self.list_items_avail.itemDoubleClicked.connect(
            lambda it: self._move(self.list_items_avail, self.list_items_sel, [it]))
        self.list_items_sel.itemDoubleClicked.connect(
            lambda it: self._move(self.list_items_sel, self.list_items_avail, [it]))
        # yield 미선택 시 fail_item / issue_table 도 선택 불가
        self.cb_sheet_yield.toggled.connect(self._sync_yield_dependents)
        self.btn_filename_change.clicked.connect(self.on_edit_filenames)
        self.btn_chart_colors.clicked.connect(self.on_edit_chart_colors)
        # Confirm 만 노출 — 취소는 ESC / 창 닫기(X)로 QDialog.reject() 자동 처리.
        self.btn_confirm.clicked.connect(self._on_confirm)
        self.btn_confirm.setMinimumHeight(36)
        self.btn_confirm.setDefault(True)
        self.btn_confirm.setStyleSheet(
            "QPushButton { font-size: 16pt; font-weight: 700; "
            "padding: 7px 21px; background: #2f7de1; color: white; "
            "border: 1px solid #1f62b8; border-radius: 6px; }"
            "QPushButton:hover { background: #236cc7; }"
        )

        self._populate_items()
        self._sync_yield_dependents()
        self._update_dut_mode_availability()

    # ── 분석 항목 선택 ────────────────────────────────────────────────────────
    def _make_item(self, idx, text):
        it = QListWidgetItem(text)
        it.setData(Qt.UserRole, idx)   # 원본 subject 순서 보존용
        return it

    def _populate_items(self):
        self.list_items_avail.clear()
        self.list_items_sel.clear()
        # default: 전체 선택 → 모두 우측(선택)
        for idx, subj in enumerate(self.group.subjects()):
            self.list_items_sel.addItem(self._make_item(idx, subj))

    def _resort(self, lw):
        """원본 subject 순서(UserRole)대로 리스트 정렬."""
        pairs = sorted((lw.item(i).data(Qt.UserRole), lw.item(i).text())
                       for i in range(lw.count()))
        lw.clear()
        for idx, text in pairs:
            lw.addItem(self._make_item(idx, text))

    def _move(self, src, dst, items):
        for it in items:
            dst.addItem(src.takeItem(src.row(it)))
        self._resort(dst)
        self._resort(src)

    def _move_all_right(self):
        self._move(self.list_items_avail, self.list_items_sel,
                   [self.list_items_avail.item(i) for i in range(self.list_items_avail.count())])

    def _move_all_left(self):
        self._move(self.list_items_sel, self.list_items_avail,
                   [self.list_items_sel.item(i) for i in range(self.list_items_sel.count())])

    def _move_selected_right(self):
        self._move(self.list_items_avail, self.list_items_sel,
                   list(self.list_items_avail.selectedItems()))

    def _move_selected_left(self):
        self._move(self.list_items_sel, self.list_items_avail,
                   list(self.list_items_sel.selectedItems()))

    def _select_fail_only(self):
        """Fail 발생 항목만 우측(선택), 나머지는 좌측(제외)."""
        if self.group is None:
            return
        subjects = self.group.subjects()
        fail_names = set(self.group.fail_subject_names())
        self.list_items_avail.clear()
        self.list_items_sel.clear()
        for idx, subj in enumerate(subjects):
            target = self.list_items_sel if subj in fail_names else self.list_items_avail
            target.addItem(self._make_item(idx, subj))

    # ── 출력 옵션 ─────────────────────────────────────────────────────────────
    def _sync_yield_dependents(self, *_):
        """yield 시트 미선택이면 fail_item / issue_table 선택 불가(해제+비활성)."""
        enabled = self.cb_sheet_yield.isChecked()
        for name in ("fail_item", "issue_table"):
            cb = self.sheet_checks[name]
            if not enabled:
                cb.setChecked(False)
            cb.setEnabled(enabled)

    def _update_dut_mode_availability(self):
        """DUT 정리는 입력 파일이 정확히 1개일 때만 가능."""
        ok = self.csv_count == 1
        if not ok:
            self.cb_mode_dut.setChecked(False)
        self.cb_mode_dut.setEnabled(ok)

    # ── Filename(legend) 편집 ─────────────────────────────────────────────────
    def _current_filenames(self):
        """현재 적용될 Filename(legend) 목록 — 사용자 지정값 우선, 없으면 파일명."""
        if self._filename_overrides is not None:
            return list(self._filename_overrides)
        return self.group.names() if self.group is not None else []

    def on_edit_filenames(self):
        """입력 파일별 legend 명(Filename)을 콤마로 구분해 한 줄로 편집."""
        names = self._current_filenames()
        if not names:
            QMessageBox.information(self, "Filename", "입력 파일이 없습니다.")
            return
        text, ok = QInputDialog.getText(
            self, "Name Change",
            "각 입력 파일의 Filename(legend)을 콤마(,)로 구분해 입력하세요.\n"
            f"(파일 {len(names)}개)",
            text=",".join(names))
        if not ok:
            return
        parts = [p.strip() for p in text.split(",")]
        # 개수 안내(부족/초과해도 적용은 앞에서부터, 빈칸은 기존명 유지)
        if len(parts) != len(names):
            QMessageBox.warning(
                self, "개수 불일치",
                f"입력 파일은 {len(names)}개인데 {len(parts)}개를 입력했습니다.\n"
                "앞에서부터 매칭하며, 빈 항목은 기존 파일명을 사용합니다.")
        self._filename_overrides = [
            (parts[i] if i < len(parts) else "") or names[i]
            for i in range(len(names))
        ]

    def filename_overrides(self):
        """사용자가 지정한 Filename(legend) 목록. 미변경이면 None."""
        return self._filename_overrides

    # ── 차트 색 편집 ──────────────────────────────────────────────────────────
    def on_edit_chart_colors(self):
        dlg = ColorEditorDialog(self)
        dlg.exec_()

    # ── 외부에서 읽는 값 ──────────────────────────────────────────────────────
    def selected_items(self):
        return [self.list_items_sel.item(i).text()
                for i in range(self.list_items_sel.count())]

    def selected_sheets(self):
        return [n for n, cb in self.sheet_checks.items() if cb.isChecked()]

    def mode_bin1(self):
        return self.cb_mode_bin1.isChecked()

    def mode_dut(self):
        return self.cb_mode_dut.isChecked()

    def auto_upload(self):
        return self.cb_auto_upload.isChecked()

    def raw_data(self):
        """체크 시 df_honey 원본 데이터를 'Raw Data' 시트로 추가."""
        return self.cb_raw_data.isChecked()

    def _on_confirm(self):
        if not self.selected_items():
            QMessageBox.warning(self, "항목 누락", "분석할 항목을 1개 이상 선택하세요.")
            return
        if not self.selected_sheets():
            QMessageBox.warning(self, "시트 누락", "출력할 시트를 1개 이상 선택하세요.")
            return
        self.accept()


def _png_from_com_shape(shape):
    """Excel COM Shape 를 클립보드 PNG 포맷으로 추출. 실패 시 1회 재시도 후 None.

    shape.Copy() 후 클립보드 반영까지 약간의 지연이 필요하므로 sleep(0.05) 한다.
    """
    import win32clipboard
    fmt = win32clipboard.RegisterClipboardFormat("PNG")
    for _ in range(2):  # 원샷 + 재시도 1회
        try:
            shape.Copy()
            time.sleep(0.05)  # 클립보드 반영 대기
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(fmt):
                    data = win32clipboard.GetClipboardData(fmt)
                    if data:
                        return bytes(data)
            finally:
                win32clipboard.CloseClipboard()
        except Exception:  # noqa: BLE001
            pass
    return None


def _normalize_grid(value):
    """win32com Range.Value (단일셀=scalar, 다중셀=tuple-of-tuples) 를 2D 리스트로 정규화.

    날짜 셀은 pywintypes 시간객체로 오므로 파이썬 datetime 으로 변환한다
    (xlwings 동작 보존 + openpyxl 저장 깨짐 방지).
    """
    def _conv(v):
        # pywintypes time 류: year 등 속성을 갖지만 숫자/문자열은 아님
        if hasattr(v, "year") and not isinstance(v, (int, float, str)):
            try:
                return datetime.datetime(
                    int(v.year), int(v.month), int(v.day),
                    int(v.hour), int(v.minute), int(v.second))
            except Exception:  # noqa: BLE001
                return v
        return v

    if not isinstance(value, (tuple, list)):           # 단일 셀 (scalar)
        return [[_conv(value)]]
    if not value or not isinstance(value[0], (tuple, list)):  # 1D
        return [[_conv(x) for x in value]]
    return [[_conv(x) for x in row] for row in value]  # 2D


def _extract_via_excel_com(src_path, header_row=3):
    """xlsx 를 win32com(Excel COM)으로 직접 열어 (DRM 자동 복호화) 시트 값을 openpyxl 로
    위치 보존 재구성. distribution/_dist* 시트는 제외, issue_table 이미지는 추출.

    xlwings 의 gencache.EnsureDispatch 가 PyInstaller(frozen) 환경에서 실패하므로,
    late-binding DispatchEx 로 Excel 을 직접 제어한다 (gen_py 캐시 불필요).

    반환: (tmp_path, issue_imgs). 실패 시 예외 raise (caller 가 fallback 처리).
    """
    import openpyxl
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(src_path, UpdateLinks=0, ReadOnly=True)

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)          # 기본 빈 시트 제거
        issue_imgs = []
        for sht in wb.Worksheets:
            name = sht.Name
            low = name.lower()
            if low == "distribution" or low.startswith("_dist"):
                continue
            ws = out_wb.create_sheet(title=name)
            ur = sht.UsedRange
            r0, c0 = ur.Row, ur.Column        # 1-based 시작 위치(위치 보존용)
            values = _normalize_grid(ur.Value)  # 항상 2D 로 정규화
            for i, row_vals in enumerate(values):
                for j, val in enumerate(row_vals):
                    if val is not None:
                        ws.cell(row=r0 + i, column=c0 + j, value=val)
            if low == "issue_table":
                for shape in sht.Shapes:
                    ri = int(shape.TopLeftCell.Row) - (header_row + 1)  # 0-based
                    if ri < 0:
                        continue
                    png = _png_from_com_shape(shape)
                    if png:
                        issue_imgs.append({"row": ri, "png": png})
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        out_wb.save(tmp.name)
        return tmp.name, sorted(issue_imgs, key=lambda x: x["row"])
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:  # noqa: BLE001
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:  # noqa: BLE001
            pass
        pythoncom.CoUninitialize()


def _prepare_upload_xlsx(src_path: str) -> tuple:
    """업로드용 xlsx 전처리. 반환: (upload_path, is_tmp, issue_imgs).

    1순위: win32com(Excel COM) — DRM 자동 복호화 + distribution 제외 + 이미지 추출.
    2순위(fallback): zip 검증 + openpyxl distribution 제거 (일반 xlsx 전용, 이미지 없음).
    is_tmp=True 이면 호출자가 upload 후 upload_path 를 삭제해야 한다.
    """
    # 1순위: Excel COM 으로 재구성 (DRM·일반 모두 처리)
    com_error = None
    try:
        tmp_path, issue_imgs = _extract_via_excel_com(src_path)
        return tmp_path, True, issue_imgs
    except Exception:
        # Excel 미설치/COM 실패 → openpyxl fallback 시도.
        # 진짜 원인을 보존해 fallback ValueError 메시지에 노출(오진 방지).
        com_error = traceback.format_exc()

    # 2순위(fallback): zip 검증 + distribution 제거
    try:
        with zipfile.ZipFile(src_path):
            pass
    except zipfile.BadZipFile:
        raise ValueError(
            "선택한 파일을 처리할 수 없습니다.\n"
            "DRM(NASCA) 파일은 Excel 이 설치된 PC 에서만 업로드할 수 있고,\n"
            "그 외에는 Excel 에서 일반 xlsx 로 다시 저장한 뒤 시도하세요.\n\n"
            f"[Excel 처리 실패 원인]\n{com_error}")

    try:
        import openpyxl
    except ImportError:
        return src_path, False, []

    wb = openpyxl.load_workbook(src_path)
    dist_names = [s for s in wb.sheetnames if s.lower() == "distribution"]
    if not dist_names:
        wb.close()
        return src_path, False, []

    for name in dist_names:
        del wb[name]
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    wb.close()
    return tmp.name, True, []


# ───────────────────────────────────────────────────────────────────────────

class HoneyMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi(UI_PATH, self)
        self.status = self.statusbar
        self.setWindowTitle(f"Honey  v{CURRENT_VERSION}")
        self.status.showMessage(f"Server: {SERVER_BASE_URL}")
        self.progress_status.hide()
        self.txt_summary.setReadOnly(True)
        self.txt_summary.setUndoRedoEnabled(False)

        self.csv_paths = []
        self.group = None          # df_honey_group
        self.last_result = None    # AnalysisResult
        self.out_path = None       # 생성된 xlsx 경로
        self._last_upload = None   # 마지막 업로드 메타 (팝업 프리필용)

        self._pt_radios = {
            "MD": self.rb_pt_MD, "PD": self.rb_pt_PD,
            "PM": self.rb_pt_PM, "SE": self.rb_pt_SE,
        }
        # 지난 실행에서 고른 Product Type 복원 (사용자별 settings.json)
        saved_pt = app_settings.get_setting("product_type")
        if saved_pt in self._pt_radios:
            self._pt_radios[saved_pt].setChecked(True)
        self._setup_csv_table()
        self._connect_signals()
        self.btn_open_local.setText("LOCAL FILE OPEN")

        if rg is None:
            self._disable_engine()
        QTimer.singleShot(500, self.check_for_update)

    def _init_run_log(self, title):
        self._run_log_started = time.perf_counter()
        self._run_log_step = 0
        self._run_log_total = 0
        self.txt_summary.clear()
        self._append_run_log(title)

    def _set_run_log_total(self, total):
        self._run_log_total = max(int(total or 0), 0)

    def _elapsed_run_log(self):
        started = getattr(self, "_run_log_started", None)
        secs = int(time.perf_counter() - started) if started is not None else 0
        return f"{secs // 60:02d}:{secs % 60:02d}"

    def _append_run_log(self, message, advance=False):
        if advance:
            self._run_log_step = int(getattr(self, "_run_log_step", 0)) + 1
        step = int(getattr(self, "_run_log_step", 0))
        total = int(getattr(self, "_run_log_total", 0))
        if total:
            prefix = f"[{self._elapsed_run_log()}] [{step:02d}/{total:02d}]"
        else:
            prefix = f"[{self._elapsed_run_log()}]"
        self.txt_summary.append(f"{prefix} {message}")
        bar = self.txt_summary.verticalScrollBar()
        bar.setValue(bar.maximum())

    def _log_profile_event(self, event):
        label = str(event.get("label") or "")
        skip = {
            "select_items",
            "split_for_diff",
            "subjects_meta",
            "subjects_meta_common",
            "build_yield",
            "build_fail_items",
            "build_issue_summary",
            "build_summary_rows",
            "build_major_fail_subjects",
            "build_cpk",
            "build_cpk_common",
            "build_distributions",
            "build_distributions_common",
            "combined_df_yield",
            "fill_cpk",
            "fill_fail_item",
            "fail_values.title",
            "fail_values.borders",
            "fill_fail_item.style",
            "normalize_sheet_names",
            "zoom_gridlines",
        }
        if label in skip:
            return
        status = event.get("status")
        elapsed = event.get("elapsed")
        error = event.get("error")
        if status == "start":
            return
        elif status == "done":
            self._append_run_log(f"{label} done: {elapsed:.2f}s" if elapsed is not None
                                 else f"{label} done", advance=True)
        elif status == "error":
            msg = f"{label} ERROR"
            if elapsed is not None:
                msg += f" after {elapsed:.2f}s"
            if error:
                msg += f" - {error}"
            self._append_run_log(msg, advance=True)

    def _estimate_run_log_steps(self, work_group, sheets, raw_data):
        sources = len(work_group.names()) if work_group is not None else 0
        table_sheets = {"summary", "yield", "cpk", "fail_item", "issue_table"}
        selected_tables = [s for s in sheets if s in table_sheets]
        steps = 0
        if raw_data:
            steps += 1  # raw_frames
        steps += 1 + sources  # analysis table builders + fail_detail per source
        steps += 1  # workbook_init
        steps += sum(1 for s in selected_tables if s != "fail_item")
        if "cpk" in selected_tables:
            steps += 4  # fill_cpk expands into five substeps
        if "fail_item" in selected_tables:
            steps += 2 + sources  # top table + FAIL_VALUES + source chunks
        if raw_data:
            steps += sources
        steps += 2  # finalize + save
        if "distribution" in sheets:
            steps += 1
        return max(steps, 1)

    def _setup_csv_table(self):
        """list_csv (QTableWidget) 를 '확장자 | 파일 경로' 2열로 구성하고,
        파일 리스트 영역에 한정한 드래그앤드롭(외부 파일)을 활성화한다."""
        t = self.list_csv
        t.setColumnCount(2)
        t.setHorizontalHeaderLabels(["확장자", "파일 경로"])
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(t.NoEditTriggers)
        t.setSelectionBehavior(t.SelectRows)
        t.setSelectionMode(t.SingleSelection)
        hh = t.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # 확장자 좁게
        hh.setSectionResizeMode(1, QHeaderView.Interactive)       # 긴 경로는 가로 스크롤
        hh.setStretchLastSection(False)
        # 드롭은 리스트 영역에서만 받는다 (메인 창엔 setAcceptDrops 를 걸지 않음).
        t.setTextElideMode(Qt.ElideNone)
        t.setWordWrap(False)
        t.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        t.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        t.verticalHeader().setDefaultSectionSize(20)
        t.setAcceptDrops(True)
        t.viewport().installEventFilter(self)

    # ── 드래그앤드롭 (파일 리스트 영역 한정) ─────────────────────────────────
    def eventFilter(self, obj, event):
        if obj is self.list_csv.viewport():
            etype = event.type()
            if etype in (QEvent.DragEnter, QEvent.DragMove):
                if event.mimeData().hasUrls():
                    event.acceptProposedAction()
                    return True
            elif etype == QEvent.Drop:
                self._handle_csv_drop(event)
                return True
        return super().eventFilter(obj, event)

    def _handle_csv_drop(self, event):
        paths = [u.toLocalFile() for u in event.mimeData().urls() if u.isLocalFile()]
        if paths:
            event.acceptProposedAction()
            self._intake(paths)   # 기존 인테이크 흐름 재사용(2개↑면 순서 팝업)

    def _connect_signals(self):
        self.btn_open_local.clicked.connect(self.on_open_local)
        self.btn_pick_csv.clicked.connect(self.on_browse_d1)
        # 입력 파일: 선택 후 ▲▼ 로 순서 변경 (맨 위 파일이 기준)
        self.btn_csv_up.clicked.connect(lambda: self._move_file(-1))
        self.btn_csv_down.clicked.connect(lambda: self._move_file(1))
        # Start: 파일 전처리 후 설정 팝업(Select Items/Option/색/Auto Upload) 열기
        self.btn_start.clicked.connect(self.on_start)
        self.btn_upload_local.clicked.connect(self.on_upload_local)
        # Product Type 선택 변경 시 사용자별 settings.json 에 즉시 저장
        for rb in self._pt_radios.values():
            rb.toggled.connect(self._save_product_type)

    def _disable_engine(self):
        # 분석 관련 기능만 비활성. 로컬 파일 직접 업로드는 엔진 없이도 동작하므로 유지.
        for name in ("btn_open_local", "btn_pick_csv", "btn_start"):
            getattr(self, name).setEnabled(False)
        self.txt_summary.setPlainText(
            "report_generator 모듈을 불러오지 못했습니다 — "
            f"{_RG_IMPORT_ERROR}\n분석/생성에는 pandas / numpy / xlwings + MS Excel 이 필요합니다."
            "\n(로컬 파일 직접 업로드는 가능합니다.)"
        )

    def _status(self, msg):
        self.status.showMessage(msg)

    # ── 입력 선택: 로컬 파일 열기 / d1_storage 검색 ─────────────────────────
    def on_open_local(self):
        # 현재 윈도우(네이티브) 파일 열기 대화상자
        paths, _ = QFileDialog.getOpenFileNames(
            self, "파일 열기 (여러 개 가능)", "",
            "모든 파일 (*.*)")
        self._intake(paths)

    def on_browse_d1(self):
        os.makedirs(D1_STORAGE_DIR, exist_ok=True)
        dlg = D1BrowserDialog(self, D1_STORAGE_DIR)
        if not dlg.exec_():
            return
        paths = dlg.selected_paths()
        if not paths:
            QMessageBox.warning(self, "선택 없음", "가져올 파일을 선택하세요.")
            return
        self._intake(paths)

    def _intake(self, paths):
        """선택된 파일들 → (2개 이상이면) 순서 지정 팝업 → 메인 창에 로드."""
        paths = list(paths or [])
        if not paths:
            return
        if len(paths) > 1:
            dlg = FileOrderDialog(self, paths)
            if not dlg.exec_():
                return
            paths = dlg.ordered_paths()
        self._load_paths(paths)

    def _refill_csv_list(self):
        """self.csv_paths 순서대로 list_csv(테이블) 다시 채우기.
        0열=확장자(좁게), 1열=파일 절대경로."""
        self.list_csv.setRowCount(len(self.csv_paths))
        for r, p in enumerate(self.csv_paths):
            full_path = str(Path(p).resolve())
            ext = Path(full_path).suffix.lstrip(".").lower()
            ext_item = QTableWidgetItem(ext)
            path_item = QTableWidgetItem(full_path)
            path_item.setData(Qt.UserRole, full_path)
            path_item.setToolTip(full_path)
            self.list_csv.setItem(r, 0, ext_item)
            self.list_csv.setItem(r, 1, path_item)
            self.list_csv.setRowHeight(r, 20)
        if self.csv_paths:
            fm = self.list_csv.fontMetrics()
            measure = getattr(fm, "horizontalAdvance", fm.width)
            width = max(measure(str(Path(p).resolve())) for p in self.csv_paths)
            self.list_csv.setColumnWidth(1, max(420, width + 36))

    def _load_paths(self, paths):
        """선택된 입력 파일들 → 리스트 채우기 + 저장 파일명 제안 (전처리는 Start 까지 보류)."""
        self.csv_paths = [str(Path(p).resolve()) for p in paths]
        self._refill_csv_list()
        self.le_outname.setText(_suggest_base_name(self.csv_paths))
        self.group = None
        self.out_path = None
        self.txt_summary.setPlainText("")
        self._status(f"{len(self.csv_paths)}개 파일 선택됨. 순서 확인 후 Start 를 누르세요.")

    def _move_file(self, delta):
        """선택한 입력 파일을 위(-1)/아래(+1)로 이동 (전처리는 Start 까지 보류)."""
        row = self.list_csv.currentRow()
        new = row + delta
        if row < 0 or not (0 <= new < len(self.csv_paths)):
            return
        self.csv_paths[row], self.csv_paths[new] = self.csv_paths[new], self.csv_paths[row]
        self._refill_csv_list()
        self.list_csv.selectRow(new)

    def _rebuild_group(self, warn=False):
        """현재 self.csv_paths 순서로 그룹 재구성 + 항목 갱신.

        맨 위(첫) 파일이 units/항목명/Lower·Upper limit 의 기준이 된다 — 서로 다른
        유형의 파일이 섞여 들어와도 첫 파일 스키마를 기준으로 데이터가 처리된다.
        """
        with _flow_time("_rebuild_group.total"):
            paths = self.csv_paths
            if not paths:
                return False

            n_files = len(paths)
            # CSV 로딩을 백그라운드 스레드(1개)에서 파일 단위로 수행한다. 동기로 돌리면
            # 무거운 pandas 읽기 동안 Qt 이벤트 루프가 멈춰 Windows 가 창을 "응답 없음"
            # 으로 표시한다. 메인 스레드는 짧게 폴링하며 processEvents() 로 UI 를 살려
            # "(진행중)" 을 보여주고, 한 파일이 60초를 넘기면 라벨만 바꾼다(중단 없음).
            _SLOW_FILE_SEC = 60
            progress = _ElapsedProgress(
                self.progress_status, "파일 로딩 준비 중...", self._status,
                busy=True, minimum=0, maximum=n_files)
            QApplication.processEvents()

            results = []
            try:
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    for i, p in enumerate(paths):
                        filename = Path(p).name
                        file_start_perf = time.perf_counter()
                        fut = ex.submit(rg.df_honey.from_csv, p)
                        file_start = time.monotonic()
                        while True:
                            done_set, _ = concurrent.futures.wait([fut], timeout=0.1)
                            elapsed = int(time.monotonic() - file_start)
                            if elapsed >= _SLOW_FILE_SEC:
                                label = (
                                    f"({i + 1}/{n_files})  {filename}  "
                                    f"전처리 계속해서 진행중입니다."
                                )
                            else:
                                label = f"({i + 1}/{n_files})  {filename}"
                            progress.set(label, value=i)
                            if done_set:
                                break
                        results.append(fut.result())  # 로드 실패 시 여기서 예외 전파
                        if _FLOW_PROFILE_ON:
                            print(
                                f"[flow-profile] honey_main.load_file[{filename}]: "
                                f"{time.perf_counter() - file_start_perf:.3f}s",
                                file=sys.stderr,
                                flush=True,
                            )
                with _flow_time("df_honey_group.construct"):
                    self.group = rg.df_honey_group(results)
            except Exception as exc:
                progress.fail(f"실패: 파일 로드 실패 - {exc}")
                QMessageBox.critical(self, "파일 로드 실패", str(exc))
                self._status("파일 로드 실패")
                self.group = None
                return False

            progress.success(f"완료: {n_files}개 파일 전처리 완료", value=n_files)

            if warn:
                with _flow_time("group.validate"):
                    issues = {name: v for name, v in self.group.validate().items() if v}
                if issues:
                    msg = "\n".join(f"- {name}: {', '.join(v)}" for name, v in issues.items())
                    QMessageBox.warning(self, "스키마 경고", f"일부 파일에 문제가 있습니다:\n{msg}")

            self.out_path = None
            self._status(f"{len(paths)}개 파일 전처리 완료 (기준: {Path(paths[0]).name}).")
            return True

    def _apply_modes(self, group, mode_bin1, mode_dut):
        """선택된 데이터 정리 모드를 그룹에 적용. 문제 시 ValueError."""
        work = group
        if mode_bin1:
            work = work.filter_rows_by_bin("1")
            if not work.subjects() or all(len(md.scores) == 0
                                          for md in work.mass_data_map.values()):
                raise ValueError("Bin1 Only: Bin 이 1(Pass)인 데이터가 없습니다.")
        if mode_dut:
            if len(self.csv_paths) != 1:
                raise ValueError("DUT 정리는 입력 파일이 1개일 때만 가능합니다.")
            work = work.split_by_dut()
        return work

    # ── Start: 전처리 → 설정 팝업 → Confirm 시 분석 실행 ─────────────────────
    def on_start(self):
        if not self.csv_paths:
            QMessageBox.warning(self, "입력 누락", "먼저 파일을 가져오세요.")
            return
        # 파일 전처리(그룹 로드/검증) 를 이 시점에 수행
        if not self._rebuild_group(warn=True) or self.group is None:
            return

        dlg = ReportSettingsDialog(self, self.group, len(self.csv_paths))
        if not dlg.exec_():
            self._status("설정 취소됨 — 다시 Start 로 진행할 수 있습니다.")
            return

        selected = dlg.selected_items()
        sheets = dlg.selected_sheets()
        # Filename(legend) 사용자 지정 시 source 명 교체 (DUT 정리 모드는 자체 명명 사용)
        overrides = dlg.filename_overrides()
        if overrides is not None and not dlg.mode_dut():
            self.group.rename_sources(overrides)
        # 데이터 정리 모드 적용 (Bin1 Only → DUT 정리 순서로 그룹 변환)
        try:
            work_group = self._apply_modes(self.group, dlg.mode_bin1(), dlg.mode_dut())
        except ValueError as exc:
            QMessageBox.warning(self, "모드 적용 불가", str(exc))
            return
        self._run_analysis(work_group, selected, sheets, dlg.auto_upload(),
                           dlg.raw_data())

    def _run_analysis(self, work_group, selected, sheets, auto_upload, raw_data=False):
        self.btn_start.setEnabled(False)
        overall_t0 = time.perf_counter()
        self._init_run_log("=== Report Generator Log ===")
        self._set_run_log_total(self._estimate_run_log_steps(work_group, sheets, raw_data))
        profile_events = queue.Queue()

        def _profile_cb(event):
            profile_events.put(event)

        def _drain_profile_events():
            while True:
                try:
                    event = profile_events.get_nowait()
                except queue.Empty:
                    break
                self._log_profile_event(event)

        # Raw Data 시트용 원본 프레임 (체크 시) — source별 df_honey 적재 포맷 그대로
        raw = None
        if raw_data:
            try:
                raw_t0 = time.perf_counter()
                with _flow_time("raw_frames"):
                    raw = work_group.raw_frames()
                self._append_run_log(f"raw_frames done: {time.perf_counter() - raw_t0:.2f}s",
                                     advance=True)
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "Raw Data 생략",
                                    f"원본 데이터 시트를 만들지 못해 건너뜁니다:\n{exc}")
                self._append_run_log(f"raw_frames ERROR - {exc}", advance=True)
                raw = None
        # 진행 단계: 준비(1) → 분석(1) → 요약(1) → 시트별(N, +Raw N) → 저장 마무리(1)
        total = len(sheets) + 4 + (len(raw) if raw else 0)
        progress = _ElapsedProgress(
            self.progress_status, "분석 준비 중...", self._status,
            busy=True, minimum=0, maximum=total)
        QApplication.processEvents()

        def _step(value, label):
            progress.set(label, value=value, status=label)

        # 1) 데이터 검증/준비
        _step(1, "데이터 검증/준비 중...")

        # 2) 데이터 분석 (통계 · Bin 집계)
        progress.set("데이터 분석 중... (통계 · Bin 집계)", status="데이터 분석 중...")
        try:
            analyze_t0 = time.perf_counter()
            with _flow_time("rg.analyze.total"):
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    fut = ex.submit(
                        rg.analyze,
                        work_group,
                        meta=rg.ReportMeta(),
                        selector=rg.ItemSelector(selected_items=selected),
                        profile_cb=_profile_cb,
                    )
                    self.last_result = _wait_for_future(fut, progress, poll_cb=_drain_profile_events)
            _drain_profile_events()
            self._append_run_log(f"Analysis total: {time.perf_counter() - analyze_t0:.2f}s")
        except Exception as exc:
            _drain_profile_events()
            self._append_run_log(f"Analysis ERROR - {exc}")
            progress.fail(f"실패: 분석 실패 - {exc}")
            QMessageBox.critical(self, "분석 실패", str(exc))
            self._status("분석 실패")
            self.btn_start.setEnabled(True)
            return
        progress.set("데이터 분석 완료", value=2)

        # 3) 요약 작성
        progress.set("요약 작성 중...", status="요약 작성 중...")
        self._show_summary(self.last_result)
        progress.set("요약 작성 완료", value=3)

        base = self.le_outname.text().strip() or _suggest_base_name(self.csv_paths, self.group)
        out = _build_output_path(Path(self.csv_paths[0]).parent, base)

        # 4) 시트/차트 생성 (시트 1개당 1스텝, offset 3)
        progress_events = queue.Queue()

        def _sheet_progress(done, total_s, name):
            progress_events.put(("sheet", done, total_s, name))

        _dist_state = {"base": 0, "n": 0, "last_log": 0}

        def _dist_progress(done, n_charts):
            progress_events.put(("dist", done, n_charts, None))

        _attach_state = {"base": 0, "last_log": {}}

        def _attach_progress(event, sheet_name, subject, done=None, total=None):
            payload = {
                "event": event,
                "sheet_name": sheet_name,
                "subject": subject,
                "done": done,
                "total": total,
            }
            progress_events.put(("attach", payload, None, None))

        def _drain_progress_events():
            while True:
                try:
                    kind, a, b, c = progress_events.get_nowait()
                except queue.Empty:
                    break
                if kind == "sheet":
                    done, total_s, name = a, b, c
                    if name == "distribution":
                        continue  # _dist_progress 가 처리
                    progress.set(
                        f"시트/차트 생성 중... ({name})   {done}/{total_s}",
                        value=3 + done,
                        status=f"시트 생성 중... ({name})  {done}/{total_s}",
                    )
                elif kind == "dist":
                    done, n_charts = a, b
                    if _dist_state["n"] == 0 and n_charts:
                        _dist_state["base"] = progress.value()
                        _dist_state["n"] = n_charts
                        progress.set_maximum(progress.maximum() + n_charts - 1)
                    pct = done * 100 // n_charts if n_charts else 100
                    value = _dist_state["base"] + done if n_charts else progress.value()
                    progress.set(
                        f"Distribution 차트 생성 중... ({done}/{n_charts} - {pct}%)",
                        value=value,
                        status=f"Distribution {pct}%  ({done}/{n_charts})",
                    )
                    if n_charts:
                        interval = max(1, n_charts // 10)
                        if done == 1 or done == n_charts or done - _dist_state["last_log"] >= interval:
                            _dist_state["last_log"] = done
                            self._append_run_log(f"Distribution chart {done}/{n_charts} ({pct}%)")
                elif kind == "attach":
                    payload = a or {}
                    event = payload.get("event")
                    sheet_name = payload.get("sheet_name") or ""
                    subject = payload.get("subject") or ""
                    done = int(payload.get("done") or 0)
                    total_a = int(payload.get("total") or 0)
                    if event == "start":
                        if total_a:
                            _attach_state["base"] = progress.value()
                            progress.set_maximum(progress.maximum() + total_a)
                        _attach_state["last_log"][sheet_name] = 0
                        continue
                    if event == "progress":
                        pct = done * 100 // total_a if total_a else 100
                        value = _attach_state["base"] + done if total_a else progress.value()
                        msg = f"PNG 붙이는 중... ({sheet_name} {done}/{total_a} - {pct}%)"
                        progress.set(msg, value=value, status=msg)
                        last_log = _attach_state["last_log"].get(sheet_name, 0)
                        if total_a and (done == total_a or done - last_log >= 10):
                            _attach_state["last_log"][sheet_name] = done
                            self._append_run_log(
                                f"PNG attach {sheet_name} {done}/{total_a} ({pct}%)")
                        continue
                    if event == "done":
                        if total_a:
                            self._append_run_log(f"PNG attach {sheet_name} done: {done}/{total_a}")
                        continue
                    if event == "copy_picture":
                        msg = "Chart 복사 붙여넣기 진행중 잠시 기다려주세요"
                        progress.set(f"{msg} ({sheet_name}: {subject})", status=msg)
                        self._append_run_log(f"{msg} ({sheet_name}: {subject})")

        progress.set(
            f"Excel 시트/차트 생성 중...  → {Path(out).name}",
            status=f"xlsx 생성 중... (Excel)  → {Path(out).name}",
        )
        try:
            colors = chart_colors.load_colors()

            def _write_job():
                com_module = _init_com_for_worker()
                try:
                    with _flow_time("xlsx_writer.write.total"):
                        return xlsx_writer.write(
                            self.last_result, out, sheets=sheets,
                            colors=colors,
                            progress_cb=_sheet_progress, raw_sheets=raw,
                            dist_progress_cb=_dist_progress,
                            attach_progress_cb=_attach_progress,
                            profile_cb=_profile_cb,
                        )
                finally:
                    _co_uninitialize(com_module)

            write_t0 = time.perf_counter()
            with _flow_time("xlsx_generation.total_wait"):
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    fut = ex.submit(_write_job)
                    _wait_for_future(
                        fut,
                        progress,
                        poll_cb=lambda: (_drain_profile_events(), _drain_progress_events()),
                    )
            _drain_profile_events()
            self._append_run_log(f"XLSX write total: {time.perf_counter() - write_t0:.2f}s")
        except Exception as exc:
            _drain_profile_events()
            self._append_run_log(f"XLSX write ERROR - {exc}")
            progress.fail(f"실패: xlsx 생성 실패 - {exc}")
            QMessageBox.critical(self, "생성 실패", str(exc))
            self._status("xlsx 생성 실패")
            self.btn_start.setEnabled(True)
            return
        _drain_progress_events()

        # 5) Excel 파일 저장 마무리
        _step(progress.maximum(), "Excel 파일 저장 마무리 중...")
        self.out_path = out
        self.btn_start.setEnabled(True)
        self._append_run_log(f"Overall total: {time.perf_counter() - overall_t0:.2f}s")
        self._append_run_log(f"저장됨: {out}")
        progress.success(f"완료: {Path(out).name} 저장됨", value=progress.maximum())
        self._status(f"완료: {Path(out).name}  ('서버에 업로드' 가능)")

        # 자동 업로드 옵션
        if auto_upload:
            self._do_upload(self.out_path)

    def _show_summary(self, r):
        feat = r.summary_feature()
        lines = [
            "",
            "=== Summary ===",
            f"Sources: {', '.join(r.sources)}",
            f"Total: {r.total_dut}    Pass(Bin1): {feat['Pass (Bin 1)']}  ({r.pass_yield}%)",
            "",
            "[Major Fail Bins]",
        ]
        for i, b in enumerate(r.major_fail_bins(), start=1):
            lines.append(f"  {i}. bin {b.get('bin')}  -  {b.get('Main Fail subject')}  ({b.get('avg')}%)")
        current = self.txt_summary.toPlainText()
        if current.strip():
            self.txt_summary.append("\n".join(lines))
        else:
            self.txt_summary.setPlainText("\n".join(lines))
        bar = self.txt_summary.verticalScrollBar()
        bar.setValue(bar.maximum())

    # ── 서버 업로드 ─────────────────────────────────────────────────────────
    def on_upload_local(self):
        """로컬에 있는 임의의 xlsx 를 직접 업로드 (분석 엔진 불필요).

        최신 Windows 탐색기(네이티브) 파일 열기 대화상자 사용 — DontUseNativeDialog
        를 주지 않아 OS 기본 다이얼로그가 뜬다.
        """
        path, _ = QFileDialog.getOpenFileName(
            self, "업로드할 파일 선택", "",
            "Excel (*.xlsx);;모든 파일 (*.*)")
        if path:
            self._do_upload(path)

    def product_type(self):
        """메인 UI 에서 선택된 Product Type (라디오). 기본 MD."""
        for key, rb in self._pt_radios.items():
            if rb.isChecked():
                return key
        return "MD"

    def _save_product_type(self, *_):
        """Product Type 선택을 사용자별 설정에 저장 (다음 실행 때 복원)."""
        app_settings.set_setting("product_type", self.product_type())

    def _do_upload(self, path):
        """메타 팝업 입력 → xlsx 전처리 → 업로드 → 완료."""
        defaults = dict(self._last_upload or {})
        defaults["product_type"] = self.product_type()
        dlg = UploadDialog(self, defaults=defaults)
        if not dlg.exec_():
            return
        v = dlg.values()
        self._last_upload = v

        self.btn_upload_local.setEnabled(False)

        # ── xlsx 전처리: xlwings 복호화/재구성 (실패 시 openpyxl fallback) ─────
        try:
            upload_path, is_tmp, issue_imgs = _prepare_upload_xlsx(path)
        except ValueError as exc:
            QMessageBox.critical(self, "파일 오류", str(exc))
            self.btn_upload_local.setEnabled(True)
            return
        except Exception as exc:
            QMessageBox.warning(
                self, "전처리 경고",
                f"xlsx 전처리 중 오류가 발생해 원본 파일로 업로드합니다:\n{exc}")
            upload_path, is_tmp, issue_imgs = path, False, []

        # ── 서버 업로드 ───────────────────────────────────────────────────
        progress = _ElapsedProgress(
            self.progress_status, f"서버 업로드 중... {Path(upload_path).name}",
            self._status, busy=True, minimum=0, maximum=0)
        QApplication.processEvents()

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                fut = ex.submit(
                    uploader.post_xlsx,
                    upload_path,
                    product_type=v["product_type"],
                    product=v["product"],
                    lot_id=v["lot_id"],
                    password=v["password"],
                    issue_imgs=issue_imgs,
                )
                result = _wait_for_future(fut, progress)
        except Exception as exc:
            progress.fail(f"실패: 업로드 실패 - {exc}")
            QMessageBox.critical(self, "업로드 실패", str(exc))
            self._status("업로드 실패")
            self.btn_upload_local.setEnabled(True)
            return
        finally:
            if is_tmp:
                try:
                    os.remove(upload_path)
                except OSError:
                    pass

        sid = result.get("session_id", "?")
        issue_saved = result.get("issue_images_saved", 0)
        progress.success(f"업로드 완료: session_id {sid}, Issue 이미지 {issue_saved}장")
        QMessageBox.information(
            self, "업로드 완료",
            f"session_id: {sid}"
            f"\nIssue 이미지: {issue_saved}장"
            + f"\n\n브라우저에서 확인:\n{SERVER_BASE_URL}/pe/report/view/{sid}",
        )
        self._status(f"업로드 완료 (Issue 이미지 {issue_saved}장)")
        self.btn_upload_local.setEnabled(True)

    # ── version check (기존 로직 무변경) ────────────────────────────────────
    def check_for_update(self):
        try:
            manifest = version_check.fetch_latest()
        except requests.exceptions.RequestException:
            # 연결 불가/타임아웃 = 서버 오프라인으로 간주, 상태바에 명확히 표시
            self.status.showMessage(
                f"⚠ 서버 오프라인 — {SERVER_BASE_URL} 에 연결할 수 없습니다")
            return
        except Exception as exc:
            self.status.showMessage(f"버전 체크 실패: {exc}")
            return

        remote = manifest.get("version") or ""
        if not version_check.is_newer(remote, CURRENT_VERSION):
            self.status.showMessage(
                f"버전 체크 OK — 최신 ({CURRENT_VERSION}). Server: {SERVER_BASE_URL}")
            return

        reply = QMessageBox.question(
            self, "업데이트 사용 가능",
            f"신규 버전 {remote} 이(가) 있습니다.\n"
            f"현재: {CURRENT_VERSION}\n\n지금 다운로드 하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes,
        )
        if reply != QMessageBox.Yes:
            return

        url = manifest.get("url") or "/honey/download"
        expected = manifest.get("sha256") or None
        setup_name = manifest.get("file") or f"HoneySetup-{remote}.exe"
        dest = Path(tempfile.gettempdir()) / setup_name

        # 다운로드 진행 상태는 메인 UI Status bar 에 표시한다.
        progress = _ElapsedProgress(
            self.progress_status, "업데이트 다운로드 중...",
            self.status.showMessage, busy=True, minimum=0, maximum=100)
        download_events = queue.Queue()

        def _cb(done, total):
            download_events.put((done, total))
            return True

        def _drain_download_events():
            while True:
                try:
                    done, total = download_events.get_nowait()
                except queue.Empty:
                    break
                label = f"업데이트 다운로드 중... ({done // (1024 * 1024)}MB"
                label += f" / {total // (1024 * 1024)}MB)" if total else ")"
                progress.set(label, value=int(done * 100 / total) if total else 0)

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                fut = ex.submit(
                    version_check.download_to,
                    dest,
                    url,
                    expected_sha256=expected,
                    progress_cb=_cb,
                )
                _wait_for_future(fut, progress, poll_cb=_drain_download_events)
        except version_check.DownloadCancelled:
            progress.fail("실패: 업데이트 다운로드 취소됨")
            self.status.showMessage("업데이트 취소됨")
            return
        except Exception as exc:
            progress.fail(f"실패: 업데이트 다운로드 실패 - {exc}")
            QMessageBox.critical(self, "다운로드 실패", str(exc))
            self.status.showMessage("업데이트 실패")
            return
        progress.success("완료: 업데이트 다운로드 완료", value=100)

        if not updater.is_frozen():
            QMessageBox.information(
                self, "다운로드 완료 (개발 모드)",
                f"스크립트 실행 중이라 설치를 진행하지 않습니다.\n"
                f"설치본만 다운로드 완료:\n{dest}\n\n"
                f"(자동 설치는 빌드된 exe 에서 동작합니다.)",
            )
            progress.success("다운로드 완료 (개발 모드)", value=100)
            self.status.showMessage("다운로드 완료 (개발 모드)")
            return

        QMessageBox.information(
            self, "업데이트 설치",
            f"새 버전 {remote} 을(를) 설치합니다.\n\n"
            "설치하는 동안 앱이 잠시 종료되며, 설치가 끝나면 자동으로 다시 실행됩니다.\n"
            "잠시만 기다려 주세요.",
        )
        try:
            updater.run_installer(dest)
        except Exception as exc:
            progress.fail(f"실패: 업데이트 설치 실행 실패 - {exc}")
            QMessageBox.critical(self, "설치 실행 실패", str(exc))
            self.status.showMessage("업데이트 실패")
            return
        progress.success("업데이트 설치 중... 앱을 종료합니다.", value=100)
        self.status.showMessage("업데이트 설치 중... 앱을 종료합니다.")
        QApplication.quit()


def _install_excepthook():
    """슬롯에서 발생한 미처리 예외로 앱이 조용히 죽지 않도록, 메시지로 표시.

    PyQt5 는 슬롯의 미처리 예외 시 기본 excepthook 이면 abort 한다. 후킹하면
    앱을 유지하면서 오류를 보여줄 수 있다.
    """
    import traceback

    def hook(etype, value, tb):
        text = "".join(traceback.format_exception(etype, value, tb))
        try:
            QMessageBox.critical(None, "오류가 발생했습니다", text[-3000:])
        except Exception:
            pass
        sys.__excepthook__(etype, value, tb)

    sys.excepthook = hook


def _apply_cute_font(app):
    """앱 전역 글씨체를 귀여운(둥근) 느낌으로. 설치된 첫 후보를 사용."""
    from PyQt5.QtGui import QFontDatabase
    available = set(QFontDatabase().families())
    # 귀여운/둥근 계열 우선순위 (설치돼 있는 첫 폰트 선택)
    candidates = ["Comic Sans MS", "Segoe Print", "Comic Neue",
                  "HY엽서L", "HY견고딕", "맑은 고딕"]
    family = next((c for c in candidates if c in available), None)
    font = QFont(family) if family else app.font()
    font.setPointSize(10)
    app.setFont(font)


def main():
    app = QApplication(sys.argv)
    _apply_cute_font(app)
    _install_excepthook()
    win = HoneyMainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
