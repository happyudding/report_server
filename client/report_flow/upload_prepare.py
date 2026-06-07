"""Upload xlsx preprocessing boundary for Honey.

This module keeps Excel COM reconstruction and issue-table image extraction
separate from the main window.  The behavior intentionally matches the old
``honey_main.py`` helpers.
"""
import datetime
import tempfile
import time
import traceback
import zipfile


def _png_from_com_shape(shape):
    """Extract PNG bytes from an Excel COM Shape via clipboard."""
    import win32clipboard
    fmt = win32clipboard.RegisterClipboardFormat("PNG")
    for _ in range(2):
        try:
            shape.Copy()
            time.sleep(0.05)
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(fmt):
                    data = win32clipboard.GetClipboardData(fmt)
                    if data:
                        return bytes(data)
            finally:
                win32clipboard.CloseClipboard()
        except Exception:  # noqa: BLE001
            pass
    return None


def _normalize_grid(value):
    """Normalize win32com Range.Value to a 2D list."""
    def _conv(v):
        if hasattr(v, "year") and not isinstance(v, (int, float, str)):
            try:
                return datetime.datetime(
                    int(v.year), int(v.month), int(v.day),
                    int(v.hour), int(v.minute), int(v.second))
            except Exception:  # noqa: BLE001
                return v
        return v

    if not isinstance(value, (tuple, list)):
        return [[_conv(value)]]
    if not value or not isinstance(value[0], (tuple, list)):
        return [[_conv(x) for x in value]]
    return [[_conv(x) for x in row] for row in value]


def _extract_via_excel_com(src_path, header_row=3):
    """Rebuild xlsx via Excel COM, excluding distribution sheets."""
    import openpyxl
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(src_path, UpdateLinks=0, ReadOnly=True)

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        issue_imgs = []
        for sht in wb.Worksheets:
            name = sht.Name
            low = name.lower()
            if low == "distribution" or low.startswith("_dist"):
                continue
            ws = out_wb.create_sheet(title=name)
            ur = sht.UsedRange
            r0, c0 = ur.Row, ur.Column
            values = _normalize_grid(ur.Value)
            for i, row_vals in enumerate(values):
                for j, val in enumerate(row_vals):
                    if val is not None:
                        ws.cell(row=r0 + i, column=c0 + j, value=val)
            if low == "issue_table":
                for shape in sht.Shapes:
                    ri = int(shape.TopLeftCell.Row) - (header_row + 1)
                    if ri < 0:
                        continue
                    png = _png_from_com_shape(shape)
                    if png:
                        issue_imgs.append({"row": ri, "png": png})
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        out_wb.save(tmp.name)
        return tmp.name, sorted(issue_imgs, key=lambda x: x["row"])
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:  # noqa: BLE001
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:  # noqa: BLE001
            pass
        pythoncom.CoUninitialize()


def prepare_upload_xlsx(src_path: str) -> tuple:
    """Prepare xlsx for upload.

    Returns ``(upload_path, is_tmp, issue_imgs)``.  A temporary path must be
    removed by the caller after upload.
    """
    com_error = None
    try:
        tmp_path, issue_imgs = _extract_via_excel_com(src_path)
        return tmp_path, True, issue_imgs
    except Exception:
        com_error = traceback.format_exc()

    try:
        with zipfile.ZipFile(src_path):
            pass
    except zipfile.BadZipFile:
        raise ValueError(
            "선택한 파일을 처리할 수 없습니다.\n"
            "DRM(NASCA) 파일은 Excel 이 설치된 PC 에서만 업로드할 수 있고,\n"
            "그 외에는 Excel 에서 일반 xlsx 로 다시 저장한 뒤 시도하세요.\n\n"
            f"[Excel 처리 실패 원인]\n{com_error}")

    try:
        import openpyxl
    except ImportError:
        return src_path, False, []

    wb = openpyxl.load_workbook(src_path)
    dist_names = [s for s in wb.sheetnames if s.lower() == "distribution"]
    if not dist_names:
        wb.close()
        return src_path, False, []

    for name in dist_names:
        del wb[name]
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    wb.close()
    return tmp.name, True, []
