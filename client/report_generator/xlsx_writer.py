"""단일 xlwings(Excel COM) 세션 xlsx 리포트 생성.

- 모든 시트(raw / summary / yield / cpk / fail_item / issue_table / distribution)를
  하나의 xw.App 세션에서 생성·스타일링·저장한다(openpyxl 미사용).
- 셀 기입은 **범위 단위 일괄(bulk range)**, 스타일은 **Range 단위 COM** 적용으로
  셀 단위 왕복을 피한다. raw data 는 임시 CSV 를 Excel 네이티브 파싱으로 복사한다.
- distribution 차트는 같은 세션에서 그린다(차트 옵션 정밀 제어 목적).

스타일 변경은 모듈 상단 상수(_HDR_FONT, _HDR_FILL_RGB 등)만 수정하면 된다.
계산은 analyzer/_builders 에서 끝났고, 이 모듈은 출력만 담당한다. Excel/xlwings 가
없으면 전체 실패한다(openpyxl fallback 없음).
"""
from __future__ import annotations

import contextlib
import contextvars
import linecache
import math
import os
import shutil
import sys
import tempfile
import time
import zipfile
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter  # 순수 열문자 util (차트 _a1 등에서 사용)

from . import DEBUG_CHART_LINE_TRACE, _profile

# ── 차트 생성 병목 측정 프로파일러 (HONEY_CHART_PROFILE set 시에만 동작) ───────
# unset 이면 _prof 는 즉시 통과 → 평상시 동작·출력 불변. 측정 결과는 stderr 로.
_PROF_ON = bool(os.environ.get("HONEY_CHART_PROFILE"))
_FLOW_PROFILE_ON = bool(os.environ.get("HONEY_FLOW_PROFILE"))
_CURRENT_PROFILE_CB = contextvars.ContextVar("xlsx_writer_profile_cb", default=None)
_CURRENT_DIST_STATS = contextvars.ContextVar("xlsx_writer_dist_stats", default=None)
_PROF = defaultdict(float)
_PROF_CNT = defaultdict(int)
_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"
_EXPORT_MOVE_RETRIES = 2
_EXPORT_RETRY_SLEEP = 0.08
_EXCEL_QUIT_FILE_READY_RETRIES = 10
_EXCEL_QUIT_FILE_READY_SLEEP = 1.0
_PNG_ATTACH_MODE = os.environ.get("HONEY_PNG_ATTACH_MODE", "export").strip().lower()
if _PNG_ATTACH_MODE not in {"export", "move_first_export", "copy_picture"}:
    _PNG_ATTACH_MODE = "export"
_PNG_SUBJECT_CACHE = os.environ.get("HONEY_PNG_SUBJECT_CACHE", "").strip().lower() in {
    "1", "true", "yes", "on",
}


def _parse_int_set(raw):
    """\"100,101\" → {100, 101} (콤마 split + int, 잘못된 토큰은 무시)."""
    out = set()
    for tok in str(raw or "").split(","):
        tok = tok.strip()
        if not tok:
            continue
        try:
            out.add(int(tok))
        except ValueError:
            pass
    return out


# Line-by-line chart debug tracer. Toggle DEBUG_CHART_LINE_TRACE in __init__.py.
# When off, sys.settrace is never called, so normal runtime is unchanged.
# 지정한 차트 순번(_LINE_TRACE_TARGETS, 1-based, skip 제외)에서 _chart_draw_at_position
# 호출 1건의 모든 실행 줄을 줄 단위 소요시간과 함께 stderr 로 출력한다.
_LINE_TRACE_ON = bool(DEBUG_CHART_LINE_TRACE)
_LINE_TRACE_TARGETS = _parse_int_set(os.environ.get("HONEY_CHART_LINE_TRACE_TARGETS", "10,11"))
_LINE_TRACE_FILE = os.path.abspath(__file__)
_LINE_TRACE_FILE_KEY = os.path.normcase(_LINE_TRACE_FILE)


@contextlib.contextmanager
def _prof(bucket):
    if not _PROF_ON:
        yield
        return
    t = time.perf_counter()
    try:
        yield
    finally:
        _PROF[bucket] += time.perf_counter() - t
        _PROF_CNT[bucket] += 1


def _prof_count(bucket, n=1):
    """시간 측정 없이 카운터만 증가 (차트/시리즈/PNG 개수 등)."""
    if _PROF_ON:
        _PROF_CNT[bucket] += n


def _emit_profile_event(profile_cb, label, status, elapsed=None, error=None, message=None):
    if profile_cb is None:
        return
    event = {
        "module": "xlsx_writer",
        "label": label,
        "status": status,
    }
    if elapsed is not None:
        event["elapsed"] = elapsed
    if error:
        event["error"] = error
    if message:
        event["message"] = message
    try:
        profile_cb(event)
    except Exception:
        pass


def _emit_profile_info(message):
    _emit_profile_event(_CURRENT_PROFILE_CB.get(), "distribution_profile", "info",
                        message=message)


@contextlib.contextmanager
def _profile_info_time(label):
    profile_cb = _CURRENT_PROFILE_CB.get()
    if profile_cb is None:
        yield
        return
    t = time.perf_counter()
    try:
        yield
    except Exception as exc:
        _emit_profile_info(f"{label} ERROR after {time.perf_counter() - t:.2f}s - {exc}")
        raise
    else:
        _emit_profile_info(f"{label} done: {time.perf_counter() - t:.2f}s")


def _new_dist_stats():
    return {
        "timings": defaultdict(lambda: {"total": 0.0, "count": 0, "max": 0.0}),
        "png": defaultdict(lambda: {
            "count": 0, "direct": 0, "moved": 0, "copy_picture": 0,
            "cache": 0, "failed": 0, "bytes": 0,
        }),
    }


def _dist_add_time(bucket, elapsed):
    stats = _CURRENT_DIST_STATS.get()
    if stats is None:
        return
    rec = stats["timings"][bucket]
    rec["total"] += elapsed
    rec["count"] += 1
    rec["max"] = max(rec["max"], elapsed)


@contextlib.contextmanager
def _dist_time(bucket):
    if _CURRENT_DIST_STATS.get() is None:
        yield
        return
    t = time.perf_counter()
    try:
        yield
    finally:
        _dist_add_time(bucket, time.perf_counter() - t)


def _dist_count_png(sheet_name, method, png_path=None):
    stats = _CURRENT_DIST_STATS.get()
    if stats is None:
        return
    rec = stats["png"][str(sheet_name)]
    rec["count"] += 1
    if method == "direct":
        rec["direct"] += 1
    elif str(method).startswith("moved"):
        rec["moved"] += 1
    elif method == "copy_picture":
        rec["copy_picture"] += 1
    elif method == "cache":
        rec["cache"] += 1
    else:
        rec["failed"] += 1
    if png_path and method != "cache":
        try:
            rec["bytes"] += os.path.getsize(png_path)
        except OSError:
            pass


def _dist_format_time(rec):
    count = rec["count"]
    avg_ms = rec["total"] * 1000.0 / count if count else 0.0
    max_ms = rec["max"] * 1000.0
    return f"total={rec['total']:.2f}s avg={avg_ms:.1f}ms max={max_ms:.1f}ms x{count}"


def _dist_emit_summary():
    stats = _CURRENT_DIST_STATS.get()
    if not stats:
        return
    timings = stats["timings"]
    _emit_profile_info(
        f"Distribution debug: png_mode={_PNG_ATTACH_MODE} "
        f"png_cache={'on' if _PNG_SUBJECT_CACHE else 'off'}"
    )
    loop_order = [
        "dist.loop.finite_scan", "dist.loop.axis_range",
        "dist.loop.chart_add", "dist.loop.series_limits", "dist.loop.series_sources",
        "dist.loop.chart_type", "dist.loop.style_limits", "dist.loop.style_sources",
        "dist.loop.axis_format", "dist.loop.title_format", "dist.loop.plot_format",
        "dist.loop.legend_format", "dist.loop.fail_bg",
    ]
    for bucket in loop_order:
        rec = timings.get(bucket)
        if rec and rec["count"]:
            _emit_profile_info(
                f"Dist loop {bucket.removeprefix('dist.loop.')}: {_dist_format_time(rec)}"
            )
    for sheet_name in ("fail_item", "issue_table"):
        rec = stats["png"].get(sheet_name)
        if not rec or not rec["count"]:
            continue
        total_mb = rec["bytes"] / (1024.0 * 1024.0)
        avg_kb = rec["bytes"] / 1024.0 / rec["count"] if rec["count"] else 0.0
        _emit_profile_info(
            f"PNG stats {sheet_name}: count={rec['count']} direct={rec['direct']} "
            f"moved={rec['moved']} copy={rec['copy_picture']} cache={rec['cache']} "
            f"failed={rec['failed']} "
            f"total={total_mb:.1f}MB avg={avg_kb:.0f}KB"
        )
        parts = []
        for name in (
            "export.direct", "export.moved", "export.failed", "picture_add",
            "copy_picture.copy", "copy_picture.paste", "copy_picture.position",
            "copy_picture.total",
        ):
            trec = timings.get(f"png.{sheet_name}.{name}")
            if trec and trec["count"]:
                parts.append(f"{name} {_dist_format_time(trec)}")
        if parts:
            _emit_profile_info(f"PNG timing {sheet_name}: " + "; ".join(parts))


@contextlib.contextmanager
def _flow_prof(bucket):
    profile_cb = _CURRENT_PROFILE_CB.get()
    if not (_FLOW_PROFILE_ON or _profile.collecting() or profile_cb is not None):
        yield
        return
    _emit_profile_event(profile_cb, bucket, "start")
    depth = _profile.push()
    t = time.perf_counter()
    try:
        yield
    except Exception as exc:
        elapsed = time.perf_counter() - t
        _profile.pop("xlsx_writer", bucket, elapsed, depth)
        _emit_profile_event(profile_cb, bucket, "error", elapsed, str(exc))
        if _FLOW_PROFILE_ON:
            print(f"[flow-profile] xlsx_writer.{bucket}: ERROR after {elapsed:.3f}s ({exc})",
                  file=sys.stderr, flush=True)
        raise
    finally:
        if sys.exc_info()[0] is None:
            elapsed = time.perf_counter() - t
            _profile.pop("xlsx_writer", bucket, elapsed, depth)
            _emit_profile_event(profile_cb, bucket, "done", elapsed)
            if _FLOW_PROFILE_ON:
                print(f"[flow-profile] xlsx_writer.{bucket}: {elapsed:.3f}s",
                      file=sys.stderr, flush=True)


def _prof_report():
    if not _PROF_ON or not _PROF:
        return
    total = sum(_PROF.values())
    print("\n[chart-profile] phase breakdown (s):", file=sys.stderr)
    for k, v in sorted(_PROF.items(), key=lambda kv: -kv[1]):
        pct = (100 * v / total) if total else 0.0
        print(f"  {k:18s} {v:8.3f}  ({pct:5.1f}%)  x{_PROF_CNT[k]}", file=sys.stderr)
    print(f"  {'TOTAL':18s} {total:8.3f}", file=sys.stderr)
    extra = {k: _PROF_CNT[k] for k in ("charts", "series", "pngs") if k in _PROF_CNT}
    if extra:
        print(f"  counts: {extra}", file=sys.stderr)
    _PROF.clear()
    _PROF_CNT.clear()

_CHARTS_PER_ROW = 5
# 차트 크기 — gap 없이 밀착 배치 (사용자 사양 324x198)
_CHART_W, _CHART_H = 324, 198
_PLOT_W, _PLOT_TOP, _PLOT_H = 280, 30, 167
# distribution 찾기(Ctrl+F)용 item 인덱스: 차트 그리드 오른쪽 열, 차트 한 행당 행 수
_INDEX_COL = 33  # AG열
_ROWS_PER_CHART = 12
# distribution 차트 그리드를 제목 배너 아래로 내리는 픽셀 오프셋
_DIST_TITLE_PX = 30

# Excel COM 상수 (distribution 차트 서식)
_XL_VALUE, _XL_CATEGORY, _XL_PRIMARY = 2, 1, 1
_XL_LOW = -4134               # xlLow (y축 TickLabelPosition)
_XL_MARKER_NONE   = -4142       # xlMarkerStyleNone
_XL_MARKER_CIRCLE = 8           # xlMarkerStyleCircle (dot 방식 series)
_MARKER_SIZE = 4             # data 점 크기(pt)
# distribution 차트 제목: item 명(subject) / 둘째줄 limit 캡션(Lo~Hi)
_CHART_TITLE_ITEM_FONT = 11   # item 명
_CHART_TITLE_CAP_FONT = 9     # Lo~Hi limit 캡션
_MSO_TRUE  = -1               # msoTrue  (LineFormat.Visible — 선 활성화)
_MSO_FALSE = 0                # msoFalse (LineFormat.Visible — 선 숨김)
_MSO_LINE_SYSDASH = 10        # msoLineSysDash (limit line)
_RGB_RED = 255               # RGB(255,0,0)
_RGB_FAIL_BG = 255 + 255 * 256 + 204 * 65536  # RGB(255,255,204) 연노랑 (fail 차트 배경)

_CPK_THRESHOLD = 1.33
_CPK_WARN_FILL_RGB = "FFFFFF00"  # 노란색 ARGB — CPK < 1.33 행 하이라이트
_CPK_WARN_FONT_RGB = "FF000000"
_CPK_TOTAL_FILL_RGB = "FF0070C0"
_CPK_TOTAL_FONT_RGB = "FFFFFFFF"
_CPK_TOTAL_CHUNK_SIZE = 100

ALL_SHEETS = ["summary", "yield", "cpk", "fail_item", "issue_table", "distribution"]

# ── 셀 스타일 상수 (색=ARGB 8자리 문자열, 폰트=dict{name,size,bold}) ─────────────
# 색상은 ARGB 8자리(뒤 6자리 RRGGBB 만 _rgb_int 가 사용). 스타일 변경 시 이 상수만 수정.
_HDR_FILL_RGB   = "FFD9E1F2"   # 헤더 연청색
_DATA_FILL_RGB  = "FFFFFFFF"   # 데이터 흰색
_TITLE_FILL_RGB = "FFBDD7EE"   # 제목 연파랑
_HDR_FONT    = {"name": "Calibri", "bold": False, "size": 11}
_DATA_FONT   = {"name": "Calibri", "size": 10}
_TITLE_FONT  = {"name": "Calibri", "bold": False, "size": 20}
_TITLE_ROW_MAX_COL = 26
_SUMMARY_TITLE_FILL_RGB = "FFBFE3FF"
_SUMMARY_HDR_FILL_RGB = "FFE2E8F0"
_SUMMARY_TITLE_FONT = {"name": "Tahoma", "bold": False, "size": 22}
_SUMMARY_SECTION_FONT = {"name": "Tahoma", "bold": False, "size": 20}
_SUMMARY_HDR_FONT = {"name": "Tahoma", "bold": False, "size": 10}
_SUMMARY_DATA_FONT = {"name": "Tahoma", "size": 10}

# ── Excel COM 상수 (서식) ─────────────────────────────────────────────────────
_XL_CENTER = -4108     # xlCenter (H/V align)
_XL_LEFT   = -4131     # xlLeft
_XL_CALC_MANUAL = -4135
_XL_CALC_AUTO   = -4105
_XL_BORDERS = (7, 8, 9, 10, 11, 12)   # xlEdge* + xlInsideVertical/Horizontal
_XL_CONTINUOUS = 1     # xlContinuous
_XL_THIN = 2           # xlThin (Borders.Weight)


def _rgb_int(argb):
    """ARGB 8자리(또는 RRGGBB) → Excel COM 색 정수 (R + G*256 + B*65536)."""
    h = str(argb)[-6:]
    return int(h[0:2], 16) + int(h[2:4], 16) * 256 + int(h[4:6], 16) * 65536


def _apply_font(api, font):
    """COM Range/Cell .api 에 폰트 dict 적용."""
    if not font:
        return
    if font.get("name"):
        api.Font.Name = font["name"]
    if font.get("size") is not None:
        api.Font.Size = font["size"]
    if font.get("bold") is not None:
        api.Font.Bold = bool(font["bold"])
    if font.get("color") is not None:
        api.Font.Color = _rgb_int(font["color"])


def _style_range(rng, *, fill=None, font=None, halign=None, valign=None,
                 wrap=None, border=False):
    """xlwings Range 에 스타일을 **범위 단위 1회** 적용 (셀 단위 COM 왕복 회피)."""
    api = rng.api
    if fill is not None:
        api.Interior.Color = _rgb_int(fill)
    _apply_font(api, font)
    if halign is not None:
        api.HorizontalAlignment = halign
    if valign is not None:
        api.VerticalAlignment = valign
    if wrap is not None:
        api.WrapText = wrap
    if border:
        for e in _XL_BORDERS:
            b = api.Borders(e)
            b.LineStyle = _XL_CONTINUOUS
            b.Weight = _XL_THIN


def _rng(ws, r1, c1, r2=None, c2=None):
    """(r1,c1)[~(r2,c2)] → xlwings Range. 단일 셀이면 r2/c2 생략."""
    if r2 is None:
        return ws.range((r1, c1))
    return ws.range((r1, c1), (r2, c2))


def _hdr_range(ws, r, c1, c2):
    """헤더 스타일(연청색 fill, _HDR_FONT, 중앙, wrap, thin border)을 범위 1회 적용."""
    rng = ws.range((r, c1), (r, c2))
    _style_range(rng, fill=_HDR_FILL_RGB, font=_HDR_FONT, halign=_XL_CENTER,
                 valign=_XL_CENTER, wrap=True, border=True)
    return rng


def _data_range(ws, r1, c1, r2, c2):
    """데이터 스타일(흰 fill, _DATA_FONT, 중앙, wrap, thin border)을 범위 1회 적용."""
    rng = ws.range((r1, c1), (r2, c2))
    _style_range(rng, fill=_DATA_FILL_RGB, font=_DATA_FONT, halign=_XL_CENTER,
                 valign=_XL_CENTER, wrap=True, border=True)
    return rng

# ── table 시트의 표 시작 위치 (A열 비움, 제목 A1, 헤더 3행, 데이터 4행~)
_HEADER_ROW = 3
_START_COL = 2  # B열
_FAIL_ITEM_ROW_HEIGHT = 135   # fail_item 데이터 행 높이(pt) — Distribution 차트 셀 맞춤
_ISSUE_TABLE_ROW_HEIGHT = 78  # issue_table 데이터 행 높이(pt) — Distribution 차트 셀 맞춤
_YIELD_TABLE_ROW_HEIGHT = 22
_YIELD_HEADER_ROW_HEIGHT = 40
_NARROW_COL_WIDTH = 6.5    # bin / count / yield / avg / comment 등 짧은 데이터
_DIST_COL_WIDTH   = 27.1   # Distribution 열 (썸네일 이미지 크기 기준)
_ITEM_COL_WIDTH   = 20.0   # Item / Category 열 (긴 텍스트)
_CPK_TEST_NAME_COL_WIDTH = 60
_CPK_SERIES_COL_WIDTH = 15
_CPK_N_COL_WIDTH = _NARROW_COL_WIDTH * 1.05
_FAIL_VALUES_COLS  = ["DUT", "XCoord", "YCoord", "Bin", "Item", "Value"]
_FAIL_VALUES_NCOLS = 6     # source 블록당 열 수
_FAIL_VALUES_GAP   = 1     # source 블록 간 빈 열 수


# ── write ────────────────────────────────────────────────────────────────────

def write(result, out_path, sheets=None, colors=None, progress_cb=None,
          raw_sheets=None, dist_progress_cb=None, attach_progress_cb=None,
          profile_cb=None) -> str:
    """AnalysisResult 를 xlsx 로 저장. 반환: 저장 경로(str).

    sheets: 출력할 시트명 리스트/집합 (None 이면 전체). 알 수 없는 이름은 무시.
    colors: distribution Legend(소스)별 '#RRGGBB' 색 리스트 (None 이면 Excel 기본색).
    progress_cb: 시트 1개 생성 후 progress_cb(done, total, name) 호출 (선택).
    raw_sheets: [(sheet명, df_honey 포맷 DataFrame), ...]. 주어지면 source(input
        file)별로 df_honey 적재 포맷 그대로의 시트를 맨 앞에 추가한다.

    단일 xlwings(Excel COM) 세션에서 모든 시트(raw/table/diff/distribution)를 생성·
    스타일링·저장한다. Excel/xlwings 가 없으면 전체 실패한다(openpyxl fallback 없음).
    """
    _CURRENT_PROFILE_CB.set(profile_cb)
    out_path = str(Path(out_path).resolve())
    sel = [s for s in ALL_SHEETS if (sheets is None or s in set(sheets))]
    if not sel:
        sel = ["summary"]

    table_writers = {
        "summary": _fill_summary,
        "yield": _fill_yield,
        "cpk": _fill_cpk,
        "fail_item": _fill_fail_item,
        "issue_table": _fill_issue_table,
    }
    want_dist = "distribution" in sel and bool(result.distributions)

    # diff compare: a_only / b_only 의 CPK·Distribution 추가 시트 스펙 준비
    dc = getattr(result, "diff_classification", None)
    diff_cpk_specs = []   # [(시트명, cpk_rows)]
    diff_dist_specs = []  # [(시트명, dists, sources)]
    if dc:
        name_a, name_b = dc["name_a"], dc["name_b"]
        if "cpk" in sel:
            if result.cpk_rows_a_only:
                diff_cpk_specs.append((f"CPK_{name_a}", result.cpk_rows_a_only))
            if result.cpk_rows_b_only:
                diff_cpk_specs.append((f"CPK_{name_b}", result.cpk_rows_b_only))
        if "distribution" in sel:
            if result.distributions_a_only:
                diff_dist_specs.append((f"Distribution_{name_a}",
                                        result.distributions_a_only, [name_a],
                                        result.dist_source_data_a_only))
            if result.distributions_b_only:
                diff_dist_specs.append((f"Distribution_{name_b}",
                                        result.distributions_b_only, [name_b],
                                        result.dist_source_data_b_only))
    want_dist_phase = want_dist or bool(diff_dist_specs)

    table_sel = [nm for nm in ALL_SHEETS if nm in table_writers and nm in sel]
    total = len(table_sel) + (len(raw_sheets) if raw_sheets else 0) \
        + (1 if want_dist else 0) + len(diff_cpk_specs) + len(diff_dist_specs)
    done = 0
    tmpdirs = []

    with xw.App(visible=False, add_book=False) as app:
        app.display_alerts = False
        app.screen_updating = False
        try:
            app.api.Calculation = _XL_CALC_MANUAL
            app.api.EnableEvents = False
        except Exception:
            pass

        with _flow_prof("workbook_init"):
            wb = app.books.add()
            base = wb.sheets[0]   # 기본 빈 시트 — 저장 전 삭제
            for nm in table_sel:
                wb.sheets.add(_report_sheet_display_name(nm),
                              after=wb.sheets[wb.sheets.count - 1])

        # raw 를 맨 앞에 삽입하기 위한 앵커(첫 table 시트, 없으면 base)
        anchor = wb.sheets[1] if wb.sheets.count > 1 else base

        # table 시트 채움 (템플릿 순서 유지)
        for nm in table_sel:
            sheet_name = _report_sheet_display_name(nm)
            ws = wb.sheets[sheet_name]
            if nm == "issue_table":
                with _flow_prof(f"fill_{nm}"):
                    _fill_issue_table(ws, result, include_cpk=("cpk" in sel))
            else:
                with _flow_prof(f"fill_{nm}"):
                    table_writers[nm](ws, result)
            done += 1
            _progress(progress_cb, done, total, nm)

        # diff compare: a_only / b_only CPK 시트 (끝에 추가)
        for disp, cpk_rows in diff_cpk_specs:
            title = _unique_sheet_name(wb, disp)
            ws = wb.sheets.add(title, after=wb.sheets[wb.sheets.count - 1])
            with _flow_prof(f"fill_{title}"):
                _fill_cpk_rows(ws, cpk_rows)
            done += 1
            _progress(progress_cb, done, total, title)

        # Raw Data — source별 df_honey 포맷 시트를 앵커 앞(맨 앞)에 순서대로 추가
        raw_titles = []
        if raw_sheets:
            reserved = [_report_sheet_display_name("distribution")] if want_dist else []
            for name, df in raw_sheets:
                title = _unique_sheet_name(wb, name, reserved)
                with _flow_prof(f"fill_raw_data[{title}]"):
                    _copy_df_via_csv(app, wb, df, title, anchor)
                raw_titles.append(title)
                done += 1
                _progress(progress_cb, done, total, title)

        # 이름 정규화 + 제목 배너/중앙정렬 (distribution 시트 생성 전 — 기존 phase1 순서)
        with _flow_prof("normalize_sheet_names"):
            _normalize_report_sheet_names(wb)
        with _flow_prof("finalize_layouts"):
            _finalize_sheet_layouts(wb, skip_title_titles={t.lower() for t in raw_titles})

        # distribution 차트 + PNG 부착 (같은 세션, 앱/워크북 재오픈 없음)
        if want_dist_phase:
            try:
                with _flow_prof("distribution_xlwings_phase"):
                    tmpdirs.extend(_write_distribution_phase(
                        app, wb, result, colors,
                        attach_fail_item=("fail_item" in sel),
                        attach_issue_cpk=("cpk" in sel),
                        dist_progress_cb=dist_progress_cb,
                        attach_progress_cb=attach_progress_cb,
                        write_main=want_dist, extra_dist=diff_dist_specs))
                done += 1 + len(diff_dist_specs)
                _progress(progress_cb, done, total, "distribution")
            except Exception as exc:
                print(f"[xlsx_writer] distribution 차트 생략: {exc}")

        # 모든 시트 Zoom/눈금선 (단일 세션 1회, distribution 포함)
        with _flow_prof("zoom_gridlines"):
            _apply_zoom_gridlines(app, wb, raw_titles)

        if wb.sheets.count > 1:
            base.delete()
        try:
            app.api.Calculation = _XL_CALC_AUTO
        except Exception:
            pass
        with _flow_prof("workbook_save"):
            wb.save(out_path)
        wb.close()

    # 저장 완료 후 파일 안정화 대기 + 임베드 이미지 무결성 검증
    _wait_for_xlsx_ready(out_path)
    try:
        _validate_embedded_images(out_path)
    finally:
        for td in tmpdirs:
            shutil.rmtree(td, ignore_errors=True)
        _prof_report()

    return out_path


def _copy_df_via_csv(app, wb, df, sheet_name, before_sheet):
    """df 를 임시 CSV 로 쓴 뒤 Excel 네이티브 파싱으로 열어 wb 의 before_sheet 앞으로 복사.

    셀 단위 기입 없이 시트를 통째로 가져온다(raw data 대량 기입 가속). 숫자 변환은
    Excel CSV 파싱이 담당(_coerce_number 대체). Raw Data 는 성능을 위해 별도 서식을
    적용하지 않는다.
    """
    # 헤더는 df.columns 로만 존재(row0=Units 불변)하므로 Serial 은 컬럼명으로 탐지
    serial_cols = [c for c in df.columns if str(c).strip() == "Serial"]
    if serial_cols:
        df = df.drop(columns=serial_cols)
    tmpdir = tempfile.mkdtemp(prefix="honey_raw_")
    csv_path = os.path.join(tmpdir, "df_temp.csv")
    df.to_csv(csv_path, index=False)   # row1=컬럼명, row2~=값 (기존 레이아웃 동일)
    before_names = {s.name for s in wb.sheets}
    wb_csv = app.books.open(csv_path)
    try:
        wb_csv.sheets[0].api.Copy(Before=before_sheet.api)
    finally:
        wb_csv.close()
        shutil.rmtree(tmpdir, ignore_errors=True)
    new = [s for s in wb.sheets if s.name not in before_names]
    copied = new[0] if new else wb.sheets.active
    copied.name = sheet_name
    return copied


def _progress(cb, done, total, name):
    if cb is None:
        return
    try:
        cb(done, total, name)
    except Exception:
        pass


# ── 채움 (table 시트, xlwings) ───────────────────────────────────────────────

def _fill_summary(ws, result):
    """Summary 시트 — 고정 좌표 레이아웃 (xlwings)."""
    meta = result.meta
    title = " ".join(x for x in [meta.product_type, meta.product, meta.lot_id] if x).strip()

    _reset_summary_sheet(ws)
    _apply_summary_dimensions(ws)
    _apply_summary_layout_styles(ws)

    _safe_set(ws, "A1", f"{chr(0x25A0)} {title or 'REPORT TITLE'}")
    _safe_set(ws, "B3", "1. Device Feature")
    _safe_set(ws, "B7", "2. Yield")
    _safe_set(ws, "B15", "3. Evaluation Summary")

    _safe_set(ws, "B4", "DEVICE")
    _safe_set(ws, "C4", "Customer")
    _safe_set(ws, "D4", "PKG_Type")
    _safe_set(ws, "E4", "GrossDie")
    _safe_set(ws, "F4", "Process Line")
    _safe_set(ws, "G4", "EVT_Version")
    _safe_set(ws, "B5", meta.product or meta.product_type or "")
    _safe_set(ws, "C5", "")
    _safe_set(ws, "D5", meta.product_type or "")
    _safe_set(ws, "E5", "")
    _safe_set(ws, "F5", meta.process or "")
    _safe_set(ws, "G5", meta.revision or "")

    _safe_set(ws, "B8", "Lot NO")
    _safe_set(ws, "D8", "Yield")
    _safe_set(ws, "E8", "Major Fail Bins")
    _safe_set(ws, "H8", "Comment")
    _safe_set(ws, "B9", meta.lot_id or "-")
    pass_row = next((r for r in result.yield_rows if str(r.get("bin")) == "1"), None)
    pass_avg = pass_row.get("avg") if pass_row else result.pass_yield
    _safe_set(ws, "D9", pass_avg if pass_avg is not None else "-")

    majors = result.major_fail_subjects(5) or result.major_fail_bins(5)
    for i in range(5):
        r = 9 + i
        _safe_set(ws, f"E{r}", _ordinal_fail_label(i + 1))
        if i < len(majors):
            _safe_set(ws, f"F{r}", majors[i].get("subject") or majors[i].get("Main Fail subject"))
            _safe_set(ws, f"G{r}", _summary_fail_percent(majors[i]))

    _safe_set(ws, "B16", "Category")
    _safe_set(ws, "C16", "Condition & Judge Limit")
    _safe_set(ws, "D16", "Result")
    for r, category in enumerate(["Yield", "CPK", "Temp", "ETC"], start=17):
        _safe_set(ws, f"B{r}", category)
    _safe_set(ws, "C17", "-")
    _safe_set(ws, "D17", "-")


def _reset_summary_sheet(ws):
    rng = ws.range("A1:H20")
    try:
        rng.api.UnMerge()
    except Exception:
        pass
    rng.clear_contents()
    _style_range(rng, fill=_DATA_FILL_RGB, font=_SUMMARY_DATA_FONT,
                 halign=_XL_CENTER, valign=_XL_CENTER, wrap=True)


def _apply_summary_dimensions(ws):
    widths = {
        "A": 2.625, "B": 16, "C": 26.125, "D": 10.375,
        "E": 10.5, "F": 12.625, "G": 9, "H": 44.75,
    }
    for col, width in widths.items():
        ws.range(f"{col}:{col}").column_width = width
    row_heights = {
        1: 30, 3: 25.5, 4: 16.5, 5: 16.5, 7: 21.75, 8: 16.5,
        15: 27, 17: 48.75, 18: 48.75, 19: 48.75, 20: 48.75,
    }
    for row, height in row_heights.items():
        ws.range(f"{row}:{row}").row_height = height


def _summary_style_range(ws, cell_range, font=None, fill_rgb=None, halign=None, border=False):
    """summary 전용 범위 스타일. halign=None 이면 center+wrap, _XL_LEFT 면 left(no wrap)."""
    center = halign is None
    _style_range(ws.range(cell_range), fill=fill_rgb or _DATA_FILL_RGB,
                 font=font or _SUMMARY_DATA_FONT,
                 halign=_XL_CENTER if center else halign,
                 valign=_XL_CENTER, wrap=center, border=border)


def _apply_summary_layout_styles(ws):
    _summary_style_range(ws, "A1:H1", _SUMMARY_TITLE_FONT, _SUMMARY_TITLE_FILL_RGB, _XL_LEFT)
    b = ws.range("A1:H1").api.Borders(9)   # xlEdgeBottom
    b.LineStyle = _XL_CONTINUOUS
    b.Weight = _XL_THIN

    for cell_range in ("B3:C3", "B7:C7", "B15:C15"):
        _summary_style_range(ws, cell_range, _SUMMARY_SECTION_FONT, _DATA_FILL_RGB, _XL_LEFT)

    _summary_style_range(ws, "B4:H4", _SUMMARY_HDR_FONT, _SUMMARY_HDR_FILL_RGB, border=True)
    _summary_style_range(ws, "B5:H5", _SUMMARY_DATA_FONT, _DATA_FILL_RGB, border=True)
    _summary_style_range(ws, "B8:H13", _SUMMARY_DATA_FONT, _DATA_FILL_RGB, border=True)
    _summary_style_range(ws, "E8:H8", _SUMMARY_HDR_FONT, _SUMMARY_HDR_FILL_RGB, border=True)
    _summary_style_range(ws, "B16:H16", _SUMMARY_HDR_FONT, _SUMMARY_HDR_FILL_RGB, border=True)
    _summary_style_range(ws, "B17:H20", _SUMMARY_DATA_FONT, _DATA_FILL_RGB, border=True)

    ws.range("D9").number_format = "0.00"
    for row in range(9, 14):
        ws.range(f"G{row}").number_format = "0.00"

    for cell_range in (
        "A1:H1", "B3:C3", "B7:C7", "B8:C8", "B9:C13", "D9:D13",
        "E8:G8", "B15:C15", "D16:H16", "D17:H17", "D18:H18",
        "D19:H19", "D20:H20",
    ):
        ws.range(cell_range).merge()


def _ordinal_fail_label(index):
    suffix = "th"
    if index == 1:
        suffix = "st"
    elif index == 2:
        suffix = "nd"
    elif index == 3:
        suffix = "rd"
    return f"{index}{suffix} Fail"


def _summary_fail_percent(row):
    if "ratio" in row and row.get("ratio") is not None:
        return row.get("ratio") * 100
    return row.get("avg")


def _yield_table(result):
    """yield / fail_item 공용 표 (bin | Item | {src}_count | {src}_yield | avg | comment)."""
    src = result.sources
    # count 들을 먼저 모두, 이어서 yield 들을 모두 (cnt cnt … yield yield …)
    header = ["bin", "Item"]
    header += [f"{s}_count" for s in src]
    header += [f"{s}_yield" for s in src]
    header += ["avg", "comment"]
    rows = []
    for r in result.yield_rows:
        row = [_bin_label(r.get("bin")), r.get("Main Fail subject", "")]
        row += [r.get(f"{s}_count") for s in src]
        row += [r.get(f"{s}_yield") for s in src]
        row += [r.get("avg"), r.get("comment", "")]
        rows.append(row)
    return header, rows


def _fill_yield(ws, result):
    if result.df_yield is not None and not result.df_yield.empty:
        header = list(result.df_yield.columns)
        rows = [list(r) for r in result.df_yield.itertuples(index=False)]
        if "comment" not in header:
            header.append("comment")
            rows = [row + [""] for row in rows]
    else:
        header, rows = _yield_table(result)
    _fill_table(ws, header, rows)
    _apply_table_col_widths(ws, header, custom_widths={"comment": 50, "Item": _ITEM_COL_WIDTH * 2})
    _apply_table_font(ws, header, size=12)
    _apply_small_font_headers(ws, header, ["_count", "_yield"], size=10)
    _set_table_row_heights(ws, len(rows), height=_YIELD_TABLE_ROW_HEIGHT)
    ws.range(f"{_HEADER_ROW}:{_HEADER_ROW}").row_height = _YIELD_HEADER_ROW_HEIGHT


def _fill_fail_item(ws, result):
    src = result.sources
    # count 들을 먼저 모두, 이어서 yield 들을 모두 (cnt cnt … yield yield …)
    header = ["Bin", "Item"]
    header += [f"{s}_count" for s in src]
    header += [f"{s}_yield" for s in src]
    header += ["Distribution"]
    rows = []
    for r in result.yield_rows:
        row = [_bin_label(r.get("bin")), r.get("Main Fail subject", "")]
        row += [r.get(f"{s}_count") for s in src]
        row += [r.get(f"{s}_yield") for s in src]
        row += [""]   # Distribution 열 — 차트는 xlwings 단계에서 삽입
        rows.append(row)
    with _flow_prof("fill_fail_item.top_table"):
        _fill_table(ws, header, rows)
        if rows:
            ws.range(f"{_HEADER_ROW + 1}:{_HEADER_ROW + len(rows)}").row_height = _FAIL_ITEM_ROW_HEIGHT
    with _flow_prof("fill_fail_item.fail_values"):
        _fill_fail_values_section(ws, result)
    with _flow_prof("fill_fail_item.style"):
        _apply_table_col_widths(ws, header, col_multiplier=1.3)
        _apply_used_cell_font(ws, size=15, bold=False)
        _apply_named_columns_font(ws, header, ["Bin", "Item"], size=15, bold=False,
                                  last_row=_HEADER_ROW + len(rows))


def _fill_fail_values_section(ws, result):
    """Fail_item 시트 bin 테이블 아래 FAIL_VALUES 섹션 — source별 fail DUT 레코드를 수평 나열."""
    fvr = getattr(result, "fail_value_rows", {})
    if not fvr:
        return

    n_bin = len(result.yield_rows)
    title_row = _HEADER_ROW + n_bin + 3   # "FAIL_VALUES" 라벨 행 (2행 공백 후)
    src_row   = title_row + 1             # source 이름 행
    hdr_row   = title_row + 2            # 열 헤더 행
    data_row0 = title_row + 3           # 데이터 시작 행

    with _flow_prof("fail_values.title"):
        _hdr_cell(ws, title_row, _START_COL, "FAIL_VALUES")

    for i, (src_name, df) in enumerate(fvr.items()):
        with _flow_prof(f"fail_values.write_source[{src_name}]"):
            ncol = df.shape[1]
            col0 = _START_COL + i * (_FAIL_VALUES_NCOLS + _FAIL_VALUES_GAP)
            c1 = col0 + ncol - 1
            # source 이름 셀 + 열 헤더행(df.columns) 일괄
            _hdr_cell(ws, src_row, col0, src_name)
            ws.range((hdr_row, col0), (hdr_row, c1)).value = list(df.columns)
            _hdr_range(ws, hdr_row, col0, c1)
            # 분석단계 산출 DataFrame 을 블록 단위 1회 write (재계산·행단위 setter 없음)
            if len(df):
                rlast = data_row0 + len(df) - 1
                # DUT/XCoord/YCoord/Bin(식별자 텍스트)은 Excel 숫자 자동변환 방지로 text 유지
                ws.range((data_row0, col0), (rlast, col0 + 3)).number_format = "@"
                ws.range((data_row0, col0), (rlast, c1)).value = df.values.tolist()
                _data_range(ws, data_row0, col0, rlast, c1)

    # 소스 블록별 윤곽선 적용 (src_row 헤더~마지막 데이터행) — 블록 Range 1회
    with _flow_prof("fail_values.borders"):
        for i, (src_name, df) in enumerate(fvr.items()):
            ncol = df.shape[1]
            col0 = _START_COL + i * (_FAIL_VALUES_NCOLS + _FAIL_VALUES_GAP)
            last_row = data_row0 + len(df) - 1 if len(df) else hdr_row
            _style_range(ws.range((src_row, col0), (last_row, col0 + ncol - 1)),
                         border=True)


def _fill_cpk(ws, result):
    _fill_cpk_rows(ws, result.cpk_rows)


def _fill_cpk_rows(ws, cpk_rows):
    header = ["TEST NAME", "LOW SPEC", "HIGH SPEC", "SCALE", "계열", "n",
              "min", "median", "max", "average", "stdev",
              "cpl", "cpu", "cp", "cpk", "comment"]
    rows = []
    total_row_offsets = []
    for r in cpk_rows:
        if str(r.get("source") or "").strip().lower() == "total":
            total_row_offsets.append(len(rows))
        rows.append([
            r.get("subject"), r.get("lower_limit"), r.get("upper_limit"),
            r.get("units"), r.get("source"), r.get("n"), r.get("min"),
            r.get("median"), r.get("max"), r.get("average"), r.get("stdev"),
            r.get("cpl"), r.get("cpu"), r.get("cp"), r.get("cpk"), "",
        ])
    _blank_repeated_cpk_labels(rows)
    _fill_cpk_table(ws, header, rows)
    _apply_cpk_total_fill(ws, total_row_offsets)
    _apply_cpk_warn_fill(ws, header, rows)   # CPK < 1.33 행 노란 하이라이트 (병합 전)
    _apply_table_col_widths(ws, header, custom_widths={
        "TEST NAME": _CPK_TEST_NAME_COL_WIDTH,
        "계열": _CPK_SERIES_COL_WIDTH,
        "n": _CPK_N_COL_WIDTH,
        "comment": 30,
    })
    _apply_font_delta_to_columns(ws, header, ["TEST NAME", "LOW SPEC", "HIGH SPEC", "SCALE"], 2)


def _blank_repeated_cpk_labels(rows):
    prev_key = None
    for row in rows:
        key = tuple(row[:4])
        if key == prev_key:
            row[0:4] = ["", "", "", ""]
        else:
            prev_key = key


def _fill_cpk_table(ws, header, rows):
    with _flow_prof(f"fill_cpk.fill_table[{len(rows)}x{len(header)}]"):
        _fill_table(ws, header, rows)


def _apply_cpk_total_fill(ws, row_offsets, header_row=_HEADER_ROW):
    if not row_offsets:
        return
    with _flow_prof("fill_cpk.total_fill"):
        excel_rows = [header_row + 1 + offset for offset in row_offsets]
        for i in range(0, len(excel_rows), _CPK_TOTAL_CHUNK_SIZE):
            addresses = [f"B{row}:P{row}"
                         for row in excel_rows[i:i + _CPK_TOTAL_CHUNK_SIZE]]
            _style_range(ws.range(",".join(addresses)),
                         fill=_CPK_TOTAL_FILL_RGB,
                         font={"color": _CPK_TOTAL_FONT_RGB})


def _apply_cpk_warn_fill(ws, header, rows, header_row=_HEADER_ROW, start_col=_START_COL):
    with _flow_prof("fill_cpk.warn_fill"):
        return _apply_cpk_warn_fill_inner(ws, header, rows, header_row, start_col)


def _apply_cpk_warn_fill_inner(ws, header, rows, header_row=_HEADER_ROW, start_col=_START_COL):
    """CPK 열 값이 _CPK_THRESHOLD 미만인 행 전체에 노란 배경 적용 (병합 전 호출)."""
    cpk_idx = next((i for i, h in enumerate(header) if h == "cpk"), None)
    if cpk_idx is None:
        return
    source_idx = 4
    ncol = len(header)
    for ri, row in enumerate(rows):
        val = row[cpk_idx] if cpk_idx < len(row) else None
        try:
            f = float(val) if val is not None else None
        except (TypeError, ValueError):
            f = None
        if f is not None and f < _CPK_THRESHOLD:
            excel_row = header_row + 1 + ri
            if len(row) > source_idx and str(row[source_idx]).strip().lower() == "total":
                cpk_col = start_col + cpk_idx
                _style_range(ws.range((excel_row, cpk_col)),
                             fill=_CPK_WARN_FILL_RGB,
                             font={"color": _CPK_WARN_FONT_RGB})
            else:
                _style_range(ws.range((excel_row, start_col), (excel_row, start_col + ncol - 1)),
                             fill=_CPK_WARN_FILL_RGB)


def _cpk_fail_subjects(result):
    """source=='total' 이고 CPK < _CPK_THRESHOLD 인 (subject, cpk_val) 목록. 순서 보존."""
    out = []
    seen = set()
    for r in (getattr(result, "cpk_rows", None) or []):
        if str(r.get("source") or "").strip().lower() != "total":
            continue
        cpk_val = r.get("cpk")
        try:
            cpk_f = float(cpk_val) if cpk_val is not None else None
        except (TypeError, ValueError):
            cpk_f = None
        if cpk_f is not None and cpk_f < _CPK_THRESHOLD:
            subj = r.get("subject")
            if subj and subj not in seen:
                seen.add(subj)
                out.append((subj, cpk_f))
    return out


def _fill_issue_table(ws, result, include_cpk=True):
    """Category 그룹 레이아웃. Yield = yield 데이터, CPK = CPK < 1.33 아이템, ETC = 플레이스홀더."""
    src = result.sources
    header = ["Category", "Bin", "Item", "avg"]
    for s in src:
        header += [f"{s}_yield"]          # count 열 제거, yield 만 유지
    header += ["Distribution", "comment", "개발 1차 comment",
               "PTE 2차 comment", "개발 2차 comment"]
    pad = len(header) - (4 + len(src))    # Distribution + comment 열 수

    rows = []
    for r in result.yield_rows:
        row = ["Yield", _bin_label(r.get("bin")),
               r.get("Main Fail subject", ""), r.get("avg")]
        for s in src:
            row += [r.get(f"{s}_yield")]  # count 제거
        row += [""] * pad
        rows.append(row)

    # CPK Category: CPK < 1.33 아이템 (source='total' 기준).
    # 카테고리 시작 행에 "item name" / "cpk" 서브헤더를 넣어 'avg' 헤더와의 혼동 방지.
    cpk_fails = _cpk_fail_subjects(result) if include_cpk else []
    n_cpk = max(1, len(cpk_fails))
    cpk_subheader = ["item name", "cpk"] if include_cpk else ["", ""]
    rows.append(["CPK", "", cpk_subheader[0], cpk_subheader[1]] + [""] * len(src) + [""] * pad)
    if cpk_fails:
        for subj, cpk_val in cpk_fails:
            row = ["", "", subj, _sanitize_cell(cpk_val)]
            row += [""] * len(src)       # _yield 열: CPK 에 해당 없음
            row += [""] * pad
            rows.append(row)
    else:
        rows.append([""] * len(header))  # CPK 없으면 빈 데이터행 유지

    rows.append(["ETC"] + [""] * (len(header) - 1))

    _fill_table(ws, header, rows)
    n_yield = len(result.yield_rows)

    _merge_issue_category(ws, n_yield)  # Yield 병합 (기존)

    # CPK Category 병합: 서브헤더 + 데이터 행 전체에 걸쳐 B열 "CPK" 세로 표시
    cpk_start = _HEADER_ROW + 1 + n_yield   # 서브헤더 행
    cpk_block = 1 + n_cpk                    # 서브헤더 + 데이터 행 수
    rng = ws.range((cpk_start, _START_COL), (cpk_start + cpk_block - 1, _START_COL))
    rng.merge()
    rng.api.VerticalAlignment = _XL_CENTER

    if n_yield:
        ws.range(f"{_HEADER_ROW + 1}:{_HEADER_ROW + n_yield}").row_height = _ISSUE_TABLE_ROW_HEIGHT
    if n_cpk:
        ws.range(f"{cpk_start + 1}:{cpk_start + n_cpk}").row_height = _ISSUE_TABLE_ROW_HEIGHT

    _apply_table_col_widths(ws, header, custom_widths={
        "Item": _ITEM_COL_WIDTH * 2,
        "Distribution": 22.1,
        "comment": 40,
        "개발 1차 comment": 40,
        "PTE 2차 comment": 40,
        "개발 2차 comment": 40,
    })
    _apply_used_cell_font(ws, size=15, bold=False)
    _apply_named_columns_font(ws, header, ["Bin", "Item"], size=15, bold=False,
                              last_row=_HEADER_ROW + len(rows))
    # CPK 서브헤더(item name / cpk)는 폰트 패스 이후 헤더 스타일 재적용 — 굵게/음영 유지
    if include_cpk:
        _hdr_range(ws, cpk_start, _START_COL + 2, _START_COL + 2)  # Item 열
        _hdr_range(ws, cpk_start, _START_COL + 3, _START_COL + 3)  # avg 열


def _merge_issue_category(ws, n_yield, header_row=_HEADER_ROW, start_col=_START_COL):
    """issue_table Category 열의 Yield 행 전체를 병합 + 세로 중앙 정렬."""
    if n_yield <= 1:
        return
    data_start = header_row + 1
    rng = ws.range((data_start, start_col), (data_start + n_yield - 1, start_col))
    rng.merge()
    rng.api.VerticalAlignment = _XL_CENTER


def _unique_sheet_name(wb, name, reserved=()):
    """Excel 시트명 규칙(≤31자, []:*?/\\ 금지, 중복 불가)으로 정제."""
    existing = {s.name.lower() for s in wb.sheets} | {str(s).lower() for s in reserved}
    return _excel_safe_sheet_name(name, existing)


def _excel_safe_sheet_name(name, existing_lower):
    """Excel 시트명 규칙(≤31자, []:*?/\\ 금지, 중복 불가)으로 정제.

    existing_lower: 이미 사용 중인 시트명(소문자) 집합. 충돌 시 _n 접미사로 회피.
    """
    import re
    base = re.sub(r"[\[\]:*?/\\]", "_", str(name or "Sheet")).strip()[:31] or "Sheet"
    cand, n = base, 2
    existing = {str(s).lower() for s in existing_lower}
    while cand.lower() in existing:
        suffix = f"_{n}"
        cand = base[:31 - len(suffix)] + suffix
        n += 1
    return cand


# ── 표 채움 헬퍼 (xlwings — 범위 단위 일괄) ──────────────────────────────────

def _last_row(ws):
    """used_range 의 마지막 행 (없으면 헤더행)."""
    try:
        return ws.used_range.last_cell.row
    except Exception:
        return _HEADER_ROW


def _hdr_cell(ws, r, c, value):
    """단일 셀에 값 기입 + 헤더 스타일 적용."""
    ws.range((r, c)).value = value
    _hdr_range(ws, r, c, c)


def _fill_table(ws, header, rows, header_row=_HEADER_ROW, start_col=_START_COL):
    """헤더+데이터를 범위 단위 일괄 기입 후 헤더행/데이터블록 스타일을 1회씩 적용."""
    ncol = len(header)
    c2 = start_col + ncol - 1
    ws.range((header_row, start_col), (header_row, c2)).value = list(header)
    _hdr_range(ws, header_row, start_col, c2)
    if not rows:
        return
    nrow = len(rows)
    data = [[_sanitize_cell(v) for v in row] for row in rows]
    ws.range((header_row + 1, start_col), (header_row + nrow, c2)).value = data
    _data_range(ws, header_row + 1, start_col, header_row + nrow, c2)


def _safe_set(ws, coord, value):
    """단일 셀 값 기입 (summary 고정 좌표용)."""
    ws.range(coord).value = value


def _is_cpk_header(header):
    return list(header[:4]) == ["TEST NAME", "LOW SPEC", "HIGH SPEC", "SCALE"] and "cpk" in header


def _apply_table_col_widths(ws, header, start_col=_START_COL, custom_widths=None, col_multiplier=1.0):
    if _is_cpk_header(header):
        with _flow_prof("fill_cpk.col_widths"):
            return _apply_table_col_widths_inner(ws, header, start_col, custom_widths, col_multiplier)
    return _apply_table_col_widths_inner(ws, header, start_col, custom_widths, col_multiplier)


def _apply_table_col_widths_inner(ws, header, start_col=_START_COL, custom_widths=None, col_multiplier=1.0):
    """헤더 이름 기반 열너비 일괄 설정.

    Distribution → _DIST_COL_WIDTH, Item/Category → _ITEM_COL_WIDTH, 나머지 → _NARROW_COL_WIDTH.
    custom_widths dict 에 있는 열은 해당 너비로 우선 적용. col_multiplier 로 전체 배율 조정.
    """
    _WIDE = {"Item", "Category"}
    for i, name in enumerate(header):
        col = start_col + i
        if custom_widths and name in custom_widths:
            w = custom_widths[name] * col_multiplier
        elif name == "Distribution":
            w = _DIST_COL_WIDTH * col_multiplier
        elif name in _WIDE:
            w = _ITEM_COL_WIDTH * col_multiplier
        else:
            w = _NARROW_COL_WIDTH * col_multiplier
        ws.range((1, col)).column_width = w


def _apply_table_font(ws, header, size=None, bold=None,
                      header_row=_HEADER_ROW, start_col=_START_COL):
    last = max(_last_row(ws), header_row)
    c2 = start_col + len(header) - 1
    _style_range(ws.range((header_row, start_col), (last, c2)),
                 font={"size": size, "bold": bold})


def _apply_used_cell_font(ws, size=None, bold=None):
    _style_range(ws.used_range, font={"size": size, "bold": bold})


def _set_table_row_heights(ws, n_rows, height, header_row=_HEADER_ROW):
    ws.range(f"{header_row}:{header_row + n_rows}").row_height = height


def _apply_named_columns_font(ws, header, names, size=None, bold=None,
                              header_row=_HEADER_ROW, start_col=_START_COL,
                              include_header=True, include_data=True,
                              last_row=None):
    name_set = set(names)
    first = header_row if include_header else header_row + 1
    last = last_row if last_row is not None else (_last_row(ws) if include_data else header_row)
    if last < first:
        last = first
    for i, name in enumerate(header):
        if name not in name_set:
            continue
        col = start_col + i
        _style_range(ws.range((first, col), (last, col)), font={"size": size, "bold": bold})


def _apply_font_delta_to_columns(ws, header, names, delta,
                                 header_row=_HEADER_ROW, start_col=_START_COL):
    if _is_cpk_header(header):
        with _flow_prof("fill_cpk.font_delta"):
            return _apply_font_delta_to_columns_inner(
                ws, header, names, delta, header_row, start_col)
    return _apply_font_delta_to_columns_inner(ws, header, names, delta, header_row, start_col)


def _apply_font_delta_to_columns_inner(ws, header, names, delta,
                                       header_row=_HEADER_ROW, start_col=_START_COL):
    """지정 열의 폰트 크기를 헤더(_HDR_FONT)·데이터(_DATA_FONT) 기준 + delta 로 설정."""
    name_set = set(names)
    last = _last_row(ws)
    for i, name in enumerate(header):
        if name not in name_set:
            continue
        col = start_col + i
        _style_range(ws.range((header_row, col), (header_row, col)),
                     font={"size": _HDR_FONT["size"] + delta})
        if last > header_row:
            _style_range(ws.range((header_row + 1, col), (last, col)),
                         font={"size": _DATA_FONT["size"] + delta})


def _normalize_report_sheet_names(wb):
    canonical = {name.lower(): _report_sheet_display_name(name) for name in ALL_SHEETS}
    existing = {s.name.lower() for s in wb.sheets}
    for ws in list(wb.sheets):
        lower = ws.name.lower()
        if not lower.endswith("1"):
            continue
        base = lower[:-1]
        target = canonical.get(base)
        if target and target.lower() not in existing:
            existing.discard(lower)
            ws.name = target
            existing.add(target.lower())


def _center_used_cells(ws):
    """used_range 전체 중앙정렬(+wrap) — 범위 1회 적용."""
    _style_range(ws.used_range, halign=_XL_CENTER, valign=_XL_CENTER, wrap=True)


def _apply_sheet_title(ws, title=None):
    """1행에 제목 배너(연파랑 fill, 좌측정렬) + A1 제목값/폰트 적용."""
    title = title or ws.name
    row1 = ws.range((1, 1), (1, _TITLE_ROW_MAX_COL))
    try:
        row1.api.UnMerge()
    except Exception:
        pass
    ws.range("1:1").row_height = 30
    _style_range(row1, fill=_TITLE_FILL_RGB, halign=_XL_LEFT, valign=_XL_CENTER)
    a1 = ws.range("A1")
    a1.value = title
    _apply_font(a1.api, _TITLE_FONT)


def _finalize_sheet_layouts(wb, skip_title_titles=()):
    """시트 레이아웃 마무리(중앙정렬 + 제목 배너). distribution 시트는 생성 전이라
    여기 대상이 아니며, skip_title_titles(Raw Data) 는 모든 서식을 건너뛴다."""
    skip = {str(t).lower() for t in skip_title_titles}
    for ws in wb.sheets:
        if ws.name.lower() == "summary":
            continue
        if ws.name.lower() in skip:
            continue
        _center_used_cells(ws)
        _apply_sheet_title(ws)


def _apply_small_font_headers(ws, header, suffixes_or_names,
                               header_row=_HEADER_ROW, start_col=_START_COL, size=9):
    """지정 조건에 해당하는 헤더 셀 폰트 크기를 size 로 변경.

    suffixes_or_names 항목이 '_' 로 시작하면 endswith 검사, 그 외엔 == 검사.
    """
    for i, name in enumerate(header):
        match = any(
            (name.endswith(pat) if pat.startswith("_") else name == pat)
            for pat in suffixes_or_names
        )
        if match:
            _style_range(ws.range((header_row, start_col + i), (header_row, start_col + i)),
                         font={"size": size})


def _bin_label(value):
    """bin 표시값: 정수 문자열이면 int, 아니면 원본."""
    s = str(value).strip()
    if s.isdigit():
        return int(s)
    try:
        f = float(s)
        if f.is_integer():
            return int(f)
    except (TypeError, ValueError):
        pass
    return value


def _sanitize_cell(v):
    """Excel 기록용 셀 정제 — NaN/inf 는 빈칸, numpy 스칼라는 native 로."""
    if v is None:
        return None
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return ""
        return v
    if isinstance(v, np.generic):
        try:
            return v.item()
        except Exception:
            return str(v)
    return v


def _report_sheet_display_name(name):
    text = str(name or "")
    return text[:1].upper() + text[1:] if text else text


# ── distribution (xlwings / Excel COM) ───────────────────────────────────────

def _write_distribution_phase(app, wb, result, colors=None, attach_fail_item=False,
                              attach_issue_cpk=True,
                              dist_progress_cb=None, attach_progress_cb=None,
                              write_main=True, extra_dist=None):
    """이미 열린 app/wb 에 distribution 시트 + 차트 + PNG 썸네일을 추가한다.

    attach_fail_item=True 면 distribution 차트를 PNG 로 export 해 fail_item 시트에
    불량율 높은 순으로 1/3 크기 썸네일로 부착한다 (차트 원본 재생성 없이 재활용).

    write_main=False 면 공통 distribution 메인 시트는 건너뛴다 (common 분포가 없는
    diff 케이스). extra_dist=[(시트명, dists, sources), ...] 는 diff compare 의
    a_only/b_only distribution 시트를 추가로 그린다 (fail_item/issue 부착 없음).

    반환: 정리할 임시 PNG 디렉토리 리스트(호출자가 save 후 rmtree).
    """
    token = _CURRENT_DIST_STATS.set(_new_dist_stats())
    tmpdirs = []
    last_sheet = None
    png_cache = {} if (_PNG_SUBJECT_CACHE and _PNG_ATTACH_MODE != "copy_picture") else None
    try:
        if write_main:
            with _prof("clear"):
                names = [s.name for s in wb.sheets]
                dist_name = next((n for n in names if n.lower() == "distribution"), None)
                if dist_name:
                    sh = wb.sheets[dist_name]
                    for c in list(sh.charts):     # 템플릿/이전 차트 제거
                        try:
                            c.delete()
                        except Exception:
                            pass
                    sh.clear()
                else:
                    sh = wb.sheets.add(_report_sheet_display_name("distribution"),
                                       after=wb.sheets[len(wb.sheets) - 1])
            chart_map = _write_distribution(
                wb, sh, result, colors, dist_progress_cb=dist_progress_cb,
                profile_label="main")
            last_sheet = sh
            if attach_fail_item and chart_map:
                with _profile_info_time("distribution.attach_fail_item"):
                    tmpdir = _attach_fail_item_charts(
                        wb, result, chart_map, attach_progress_cb=attach_progress_cb,
                        png_cache=png_cache
                    )
                if tmpdir:
                    tmpdirs.append(tmpdir)
            if chart_map:
                with _profile_info_time("distribution.attach_issue_table"):
                    tmpdir = _attach_issue_table_charts(
                        wb, result, chart_map, include_cpk=attach_issue_cpk,
                        attach_progress_cb=attach_progress_cb, png_cache=png_cache
                    )
                if tmpdir:
                    tmpdirs.append(tmpdir)

        # diff compare: a_only / b_only distribution 시트 추가
        for raw_title, dists, sources, sdata in (extra_dist or []):
            if not dists:
                continue
            existing = {s.name.lower() for s in wb.sheets}
            sheet_title = _excel_safe_sheet_name(raw_title, existing)
            if sheet_title.lower() in existing:
                d_sh = wb.sheets[sheet_title]
                for c in list(d_sh.charts):
                    try:
                        c.delete()
                    except Exception:
                        pass
                d_sh.clear()
            else:
                d_sh = wb.sheets.add(sheet_title, after=wb.sheets[len(wb.sheets) - 1])
            _write_distribution(
                wb, d_sh, result, colors, dists=dists, sources=sources,
                source_data=sdata, title=sheet_title, profile_label=f"diff:{sheet_title}")
            last_sheet = d_sh

        if last_sheet is not None:
            try:
                last_sheet.activate()
            except Exception:
                pass
        return tmpdirs
    finally:
        _dist_emit_summary()
        _CURRENT_DIST_STATS.reset(token)


def _apply_zoom_gridlines(app, wb, raw_gridline_sheets=None):
    """모든 시트 Zoom(fail_item/issue_table/distribution=80, 그 외 100) + 눈금선 숨김.
    raw 시트만 눈금선 표시 (단일 세션 1회 적용)."""
    zoom80 = {"fail_item", "issue_table", "distribution"}
    raw_names = {str(n).lower() for n in (raw_gridline_sheets or [])}
    for s in wb.sheets:
        try:
            s.activate()
            app.api.ActiveWindow.DisplayGridlines = s.name.lower() in raw_names
            nm_key = s.name.lower().replace(" ", "_")
            app.api.ActiveWindow.Zoom = 80 if any(z in nm_key for z in zoom80) else 100
        except Exception:
            pass


def _attach_fail_item_charts(wb, result, chart_map, attach_progress_cb=None, png_cache=None):
    """fail_item 시트의 Distribution 열(각 bin 행)에 fail item 차트 PNG 삽입.

    한 bin 에 fail item 이 여럿일 수 있으므로, 해당 행 fail_subjects 전체를 불량율
    (portion %) 내림차순으로 Distribution 칸에서 오른쪽으로 나열한다 (fail_subjects 는
    이미 정렬됨). Distribution 은 마지막 열이라 우측 빈 공간으로 확장된다.
    """
    import os
    import tempfile

    names = [s.name for s in wb.sheets]
    fi_name = next((n for n in names if n.lower() == "fail_item"), None)
    if fi_name is None:
        return None
    fi = wb.sheets[fi_name]

    # Distribution 열 = B(2) + bin + Item + (count+yield) × sources
    dist_col = _START_COL + 2 + 2 * len(result.sources)

    tmpdir = tempfile.mkdtemp(prefix="honey_fi_")
    seq = 0
    total = sum(
        1
        for r in result.fail_item_rows
        for fs in (r.get("fail_subjects") or [])
        if fs.get("subject") in chart_map
    )
    done = 0
    _notify_attach_progress(attach_progress_cb, "start", "fail_item", "", done, total)

    for i, r in enumerate(result.fail_item_rows):
        fail_subjects = r.get("fail_subjects") or []
        if not fail_subjects:
            continue
        row_excel = _HEADER_ROW + 1 + i
        try:
            cell = fi.range((row_excel, dist_col))
            base_left = cell.left
            top = cell.top
            w = cell.width
            h = cell.height
        except Exception:
            base_left, top = 700.0, 60.0 + i * _FAIL_ITEM_ROW_HEIGHT
            w, h = 200.0, float(_FAIL_ITEM_ROW_HEIGHT)
        k = 0   # 실제 부착된 차트 수 — 가로 위치 인덱스
        for fs in fail_subjects:
            subj = fs.get("subject")
            if not subj or subj not in chart_map:
                continue
            ch = chart_map[subj]   # COM Chart (Pass2 가 COM Chart 를 저장)
            left = base_left + k * w
            png = os.path.join(tmpdir, f"fi_{seq}.png")
            seq += 1
            if _attach_chart_picture(fi, ch, png, f"fi_chart_{seq}", left, top, w, h,
                                     "fail_item", subj, attach_progress_cb,
                                     png_cache=png_cache):
                _prof_count("pngs")
            done += 1
            _notify_attach_progress(
                attach_progress_cb, "progress", "fail_item", subj, done, total)
            k += 1

    _notify_attach_progress(attach_progress_cb, "done", "fail_item", "", done, total)
    return tmpdir if seq > 0 else None


def _attach_issue_table_charts(wb, result, chart_map, include_cpk=True,
                               attach_progress_cb=None, png_cache=None):
    """issue_table 시트의 Distribution 열(각 데이터 행)에 해당 subject 차트 PNG 삽입.

    fail_item 과 동일한 COM Export 방식. dist_col 계산만 issue_table header 기준으로 다름.
    header: ["Category","bin","Item","avg", {src}_yield×N, "Distribution", ...]
    → dist_col = _START_COL + 4 + len(sources)
    """
    import os
    import tempfile

    names = [s.name for s in wb.sheets]
    it_name = next((n for n in names if n.lower() == "issue_table"), None)
    if it_name is None:
        return None
    it = wb.sheets[it_name]

    dist_col = _START_COL + 4 + len(result.sources)

    tmpdir = tempfile.mkdtemp(prefix="honey_it_")
    seq = 0
    cpk_subjects = _cpk_fail_subjects(result) if include_cpk else []
    total = sum(
        1
        for r in result.yield_rows
        if r.get("Main Fail subject") in chart_map
    ) + sum(1 for subj, _cpk_val in cpk_subjects if subj in chart_map)
    done = 0
    _notify_attach_progress(attach_progress_cb, "start", "issue_table", "", done, total)

    for i, r in enumerate(result.yield_rows):
        subj = r.get("Main Fail subject")
        if not subj or subj not in chart_map:
            continue
        ch = chart_map[subj]
        row_excel = _HEADER_ROW + 1 + i
        try:
            cell = it.range((row_excel, dist_col))
            left = cell.left
            top = cell.top
            w = cell.width
            h = cell.height
        except Exception:
            left, top = 700.0, 60.0 + i * _ISSUE_TABLE_ROW_HEIGHT
            w, h = 200.0, float(_ISSUE_TABLE_ROW_HEIGHT)
        png = os.path.join(tmpdir, f"it_{seq}.png")
        seq += 1
        if _attach_chart_picture(it, ch, png, f"it_chart_{seq}", left, top, w, h,
                                 "issue_table", subj, attach_progress_cb,
                                 png_cache=png_cache):
            _prof_count("pngs")
        done += 1
        _notify_attach_progress(
            attach_progress_cb, "progress", "issue_table", subj, done, total)

    if not include_cpk:
        _notify_attach_progress(attach_progress_cb, "done", "issue_table", "", done, total)
        return tmpdir if seq > 0 else None

    # CPK < 1.33 행 distribution 차트 부착 (+1: CPK 카테고리 서브헤더 행 보정)
    n_yield = len(result.yield_rows)
    for j, (subj, _cpk_val) in enumerate(cpk_subjects):
        if subj not in chart_map:
            continue
        ch = chart_map[subj]
        row_excel = _HEADER_ROW + 1 + n_yield + 1 + j
        try:
            cell = it.range((row_excel, dist_col))
            left = cell.left
            top = cell.top
            w = cell.width
            h = cell.height
        except Exception:
            left = 700.0
            top = 60.0 + (n_yield + 1 + j) * _ISSUE_TABLE_ROW_HEIGHT
            w, h = 200.0, float(_ISSUE_TABLE_ROW_HEIGHT)
        png = os.path.join(tmpdir, f"it_{seq}.png")
        seq += 1
        if _attach_chart_picture(it, ch, png, f"it_chart_{seq}", left, top, w, h,
                                 "issue_table", subj, attach_progress_cb,
                                 png_cache=png_cache):
            _prof_count("pngs")
        done += 1
        _notify_attach_progress(
            attach_progress_cb, "progress", "issue_table", subj, done, total)

    _notify_attach_progress(attach_progress_cb, "done", "issue_table", "", done, total)
    return tmpdir if seq > 0 else None


def _attach_chart_picture(sheet, chart, png_path, name, left, top, width, height,
                          sheet_name, subject, attach_progress_cb=None, png_cache=None):
    """Export a COM chart as PNG, then embed it as a picture on target sheet."""
    try:
        if _PNG_ATTACH_MODE == "copy_picture":
            _notify_attach_progress(attach_progress_cb, "copy_picture", sheet_name, subject)
            with _prof(f"{sheet_name}.copy_picture"):
                with _dist_time(f"png.{sheet_name}.copy_picture.total"):
                    _copy_chart_picture_to_sheet(
                        chart, sheet, name, left, top, width, height, sheet_name=sheet_name)
            _dist_count_png(sheet_name, "copy_picture")
            return True

        cached_png = png_cache.get(subject) if png_cache is not None else None
        if cached_png and os.path.exists(cached_png):
            with _prof(f"{sheet_name}.picadd"):
                with _dist_time(f"png.{sheet_name}.picture_add"):
                    _add_picture_from_file(sheet, cached_png, name, left, top, width, height)
            _dist_count_png(sheet_name, "cache", cached_png)
            return True

        export_t0 = time.perf_counter()
        with _prof(f"{sheet_name}.export"):
            if _PNG_ATTACH_MODE == "move_first_export":
                method = _export_chart_png_move_first(chart, png_path)
            else:
                method = _export_chart_png_stable(chart, png_path)
        export_elapsed = time.perf_counter() - export_t0
        if method:
            export_bucket = "export.direct" if method == "direct" else "export.moved"
            _dist_add_time(f"png.{sheet_name}.{export_bucket}", export_elapsed)
            with _prof(f"{sheet_name}.picadd"):
                with _dist_time(f"png.{sheet_name}.picture_add"):
                    _add_picture_from_file(sheet, png_path, name, left, top, width, height)
            if png_cache is not None:
                png_cache[subject] = png_path
            _dist_count_png(sheet_name, method, png_path)
            return True
        _dist_add_time(f"png.{sheet_name}.export.failed", export_elapsed)
        _notify_attach_progress(attach_progress_cb, "copy_picture", sheet_name, subject)
        with _prof(f"{sheet_name}.copy_picture"):
            with _dist_time(f"png.{sheet_name}.copy_picture.total"):
                _copy_chart_picture_to_sheet(
                    chart, sheet, name, left, top, width, height, sheet_name=sheet_name)
        _dist_count_png(sheet_name, "copy_picture")
        _log_chart_attach(f"{sheet_name}:{subject} used CopyPicture fallback")
        return True
    except Exception as exc:
        _dist_count_png(sheet_name, "failed")
        _log_chart_attach(f"{sheet_name}:{subject} attach failed: {exc!r}")
        return False


def _add_picture_from_file(sheet, png_path, name, left, top, width, height):
    sheet.pictures.add(
        png_path,
        link_to_file=False,
        save_with_document=True,
        name=name,
        left=left,
        top=top,
        width=width,
        height=height,
    )


def _export_chart_png_stable(chart, png_path):
    """Keep COM Chart.Export, but retry after moving off-screen charts into view."""
    if _export_chart_png_once(chart, png_path):
        return "direct"

    chart_object = _chart_object(chart)
    if chart_object is None:
        return None

    old_left = old_top = None
    try:
        old_left, old_top = chart_object.Left, chart_object.Top
        for attempt in range(1, _EXPORT_MOVE_RETRIES + 1):
            chart_object.Left = 0
            chart_object.Top = _DIST_TITLE_PX + (attempt - 1) * (_CHART_H + 6)
            time.sleep(_EXPORT_RETRY_SLEEP * attempt)
            if _export_chart_png_once(chart, png_path):
                return f"moved{attempt}"
    except Exception as exc:
        _log_chart_attach(f"Chart.Export move retry failed: {exc!r}")
    finally:
        if old_left is not None and old_top is not None:
            try:
                chart_object.Left = old_left
                chart_object.Top = old_top
            except Exception:
                pass
    return None


def _export_chart_png_move_first(chart, png_path):
    """Move the chart into Excel's renderable area before the first PNG export."""
    chart_object = _chart_object(chart)
    if chart_object is None:
        return None

    old_left = old_top = None
    try:
        old_left, old_top = chart_object.Left, chart_object.Top
        for attempt in range(1, _EXPORT_MOVE_RETRIES + 2):
            chart_object.Left = 0
            chart_object.Top = _DIST_TITLE_PX + (attempt - 1) * (_CHART_H + 6)
            time.sleep(_EXPORT_RETRY_SLEEP * attempt)
            if _export_chart_png_once(chart, png_path):
                return f"moved{attempt}"
    except Exception as exc:
        _log_chart_attach(f"Chart.Export move-first failed: {exc!r}")
    finally:
        if old_left is not None and old_top is not None:
            try:
                chart_object.Left = old_left
                chart_object.Top = old_top
            except Exception:
                pass
    return None


def _export_chart_png_once(chart, png_path):
    try:
        if os.path.exists(png_path):
            os.remove(png_path)
        chart.Export(png_path, "PNG")
    except Exception as exc:
        _log_chart_attach(f"Chart.Export failed: {exc!r}")
        return False
    return _is_valid_png(png_path)


def _is_valid_png(png_path):
    try:
        if os.path.getsize(png_path) <= len(_PNG_MAGIC):
            return False
        with open(png_path, "rb") as fh:
            return fh.read(len(_PNG_MAGIC)) == _PNG_MAGIC
    except OSError:
        return False


def _copy_chart_picture_to_sheet(chart, sheet, name, left, top, width, height, sheet_name=None):
    chart_object = _chart_object(chart)
    if chart_object is None:
        raise RuntimeError("chart object not found for CopyPicture fallback")
    prefix = f"png.{sheet_name}." if sheet_name else "png."
    with _dist_time(prefix + "copy_picture.copy"):
        chart_object.CopyPicture(Appearance=1, Format=-4147)
    with _dist_time(prefix + "copy_picture.paste"):
        sheet_api = sheet.api
        shapes = sheet_api.Shapes
        before = int(shapes.Count)
        try:
            sheet_api.Activate()
        except Exception:
            pass
        sheet_api.Paste()
        after = int(shapes.Count)
        if after <= before:
            raise RuntimeError("CopyPicture paste did not create a shape")
        shape = shapes.Item(after)
    with _dist_time(prefix + "copy_picture.position"):
        shape.Name = name
        shape.Left = float(left)
        shape.Top = float(top)
        shape.Width = float(width)
        shape.Height = float(height)
    return shape


def _notify_attach_progress(cb, event, sheet_name, subject, done=None, total=None):
    if cb is None:
        return
    try:
        cb(event, sheet_name, subject, done, total)
    except Exception:
        pass


def _chart_object(chart):
    try:
        return chart.Parent
    except Exception:
        return None


def _log_chart_attach(message):
    print(f"[xlsx_writer] chart attach: {message}", file=sys.stderr)


def _validate_embedded_images(xlsx_path):
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = set(zf.namelist())
            rel_names = [n for n in names if n.startswith("xl/drawings/_rels/")
                         and n.endswith(".rels")]
            for rel_name in rel_names:
                rel_xml = zf.read(rel_name).decode("utf-8", errors="replace")
                if 'Target="NULL"' in rel_xml or "Target='NULL'" in rel_xml:
                    raise RuntimeError(f"broken image relationship in {rel_name}: Target=NULL")
                for target in _image_rel_targets(rel_xml):
                    part = _resolve_xlsx_part(rel_name, target)
                    if part not in names:
                        raise RuntimeError(
                            f"broken image relationship in {rel_name}: missing {part}"
                        )
    except zipfile.BadZipFile as exc:
        raise RuntimeError(f"invalid xlsx package: {xlsx_path}") from exc


def _wait_for_xlsx_ready(xlsx_path):
    last_exc = None
    for attempt in range(1, _EXCEL_QUIT_FILE_READY_RETRIES + 1):
        try:
            with zipfile.ZipFile(xlsx_path) as zf:
                zf.namelist()
            return
        except (zipfile.BadZipFile, PermissionError, OSError) as exc:
            last_exc = exc
            if attempt >= _EXCEL_QUIT_FILE_READY_RETRIES:
                break
            time.sleep(_EXCEL_QUIT_FILE_READY_SLEEP)
    raise RuntimeError(
        f"xlsx file is not ready after Excel quit: {xlsx_path}"
    ) from last_exc


def _is_package_integrity_error(exc):
    msg = str(exc)
    return ("broken image relationship" in msg
            or "invalid xlsx package" in msg)


def _image_rel_targets(rel_xml):
    import xml.etree.ElementTree as ET

    root = ET.fromstring(rel_xml)
    for rel in root:
        typ = rel.attrib.get("Type", "")
        if typ.endswith("/image"):
            yield rel.attrib.get("Target", "")


def _resolve_xlsx_part(rel_name, target):
    base = Path(rel_name).parent.parent
    part = (base / target).as_posix()
    while "/../" in part:
        left, right = part.split("/../", 1)
        part = left.rsplit("/", 1)[0] + "/" + right
    return part.lstrip("/")


def _hex_to_excel_rgb(hex_color):
    """'#RRGGBB' → Excel COM RGB 정수 (R + G*256 + B*65536). 실패 시 None."""
    try:
        s = str(hex_color).strip().lstrip("#")
        r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
        return r + (g << 8) + (b << 16)
    except Exception:
        return None


def _chart_pos(i):
    """차트 grid 좌상단 픽셀 (gap 없이 밀착, 제목 배너 아래)."""
    col = i % _CHARTS_PER_ROW
    grow = i // _CHARTS_PER_ROW
    return col * _CHART_W, _DIST_TITLE_PX + grow * _CHART_H


def sort_alldata(df, ascending=True):
    """각 열을 독립적으로 오름차순 정렬(NaN 말미). 행=정렬된 값, 열=subject."""
    out = df.copy()
    for c in df.columns:
        out[c] = df[c].sort_values(ascending=ascending, na_position="last").to_numpy()
    return out


def sort_data_to_percent(df):
    """열별 ECDF(rank/count). y = 행순위/notna개수, >1 또는 count=0 → NaN (모든 DUT 1점)."""
    inds = np.arange(df.shape[0]).reshape(-1, 1) + 1            # (N,1)
    counts = df.notna().sum().to_numpy()                        # (n_cols,)
    vals = np.divide(inds, counts, out=np.full(df.shape, np.nan, dtype=float),
                     where=counts != 0)
    vals[vals > 1] = np.nan
    return pd.DataFrame(vals, columns=df.columns)


def _build_compact_dist_y(df_x_list):
    """Build ECDF Y helper columns keyed by non-NaN count.

    X data still uses one column per subject. Y only depends on the number of
    valid points, so subjects with the same count can share one Y column.
    """
    if not df_x_list:
        return pd.DataFrame(), [], [], {
            "full_y_cells": 0, "compact_y_cells": 0, "unique_counts": 0,
            "unique_count_cases": 0, "valid_refs": 0, "trimmed_nan_tail": 0,
        }

    count_arrs = [df.notna().sum().astype(int).to_numpy() for df in df_x_list]
    unique_counts_by_source = []
    for counts in count_arrs:
        unique_counts = sorted({int(c) for c in counts})
        unique_counts_by_source.append(unique_counts or [0])
    max_unique_counts = max(len(counts) for counts in unique_counts_by_source)

    columns = [f"count_slot_{i + 1}" for i in range(max_unique_counts)]
    y_blocks = []
    y_cols_by_source = []
    for df, unique_counts, counts in zip(df_x_list, unique_counts_by_source, count_arrs):
        n_rows = df.shape[0]
        block = np.full((n_rows, max_unique_counts), np.nan, dtype=float)
        count_to_pos = {count: pos for pos, count in enumerate(unique_counts)}
        for count, pos in count_to_pos.items():
            if count <= 0:
                continue
            upto = min(count, n_rows)
            block[:upto, pos] = np.arange(1, upto + 1, dtype=float) / float(count)
        y_blocks.append(pd.DataFrame(block, columns=columns))
        y_cols_by_source.append([
            get_column_letter(count_to_pos[int(count)] + 2) for count in counts
        ])

    df_y = pd.concat(y_blocks, ignore_index=True) if y_blocks else pd.DataFrame(columns=columns)
    full_y_cells = sum(df.shape[0] * df.shape[1] for df in df_x_list)
    compact_y_cells = df_y.shape[0] * df_y.shape[1]
    valid_refs = sum(int(counts.sum()) for counts in count_arrs)
    return df_y, y_cols_by_source, count_arrs, {
        "full_y_cells": full_y_cells,
        "compact_y_cells": compact_y_cells,
        "unique_counts": max_unique_counts,
        "unique_count_cases": sum(len(counts) for counts in unique_counts_by_source),
        "valid_refs": valid_refs,
        "trimmed_nan_tail": full_y_cells - valid_refs,
    }


def _unique_helper_name(base, existing):
    """존재 시트명과 충돌 회피한 헬퍼 시트명(정리/정리_Y)."""
    name, n = base, 2
    while name in existing:
        name = f"{base}{n}"
        n += 1
    return name


# ── distribution 차트 그리기 (subject별 순회 + 매번 새로 생성) ────────────────
# HONEY 원본 구조 참조: draw_all_chart → chart_draw_at_position →
#   {chart_title_limit_set, chart_data_set, chart_layout_setting}.
# template/duplicate 미사용. 차트당 COM 객체(api[1]/SeriesCollection/Axes/Legend 등)는
# 한 번만 바인딩해 호출 수를 최소화한다.

def _draw_all_chart(sh, dists, x_arrs, x_name, y_name, cnt_list, src_names, colors,
                    y_cols_by_source, count_matrix, n_charts, dist_progress_cb):
    """전체 subject 를 순회하며 차트를 1개씩 새로 생성.

    유한 데이터가 없는 subject 는 건너뛰되 grid 인덱스 i 는 유지해 칸 gap 을 보존한다.
    반환: (chart_map{subject: COM Chart}, index_entries[(row, subject)]).
    """
    chart_map = {}
    index_entries = []
    done = 0
    chart_seq = 0   # skip 제외, 실제 생성 시도한 차트의 1-based 순번 (line-trace 타깃)
    for i, d in enumerate(dists):
        with _dist_time("dist.loop.finite_scan"):
            cols = [arr[:, i] for arr in x_arrs]
            finite = np.concatenate([c[np.isfinite(c)] for c in cols]) if cols else np.empty(0)
            if finite.size:
                data_min, data_max = float(finite.min()), float(finite.max())
        if finite.size == 0:
            done += 1
            if dist_progress_cb:
                dist_progress_cb(done, n_charts)
            continue
        with _dist_time("dist.loop.axis_range"):
            lo, hi = d.lower_limit, d.upper_limit
            is_fail = (_isnum(lo) and data_min < float(lo)) or (
                _isnum(hi) and data_max > float(hi))
            x_min, x_max = _x_axis_range(lo, hi, data_min, data_max, is_fail)

        chart_seq += 1
        try:
            with _trace_chart_lines(chart_seq, d.subject):
                chart_map[d.subject] = _chart_draw_at_position(
                    sh, i, d, x_name, y_name, i, cnt_list, src_names, colors,
                    x_min, x_max, is_fail, y_cols_by_source, count_matrix)
        except Exception as _e:
            print(f"[dist-chart] skip subject={d.subject!r}: {_e}", file=sys.stderr)
        col = i % _CHARTS_PER_ROW
        grow = i // _CHARTS_PER_ROW
        index_entries.append((2 + grow * _ROWS_PER_CHART + col, d.subject))
        done += 1
        if dist_progress_cb:
            dist_progress_cb(done, n_charts)
    return chart_map, index_entries


def _chart_draw_at_position(sh, i, d, x_sheet, y_sheet, col_idx, cnt_list, src_names,
                            colors, x_min, x_max, is_fail, y_cols_by_source, count_matrix):
    """차트 1개를 새로 생성·서식 적용 (정리/정리_Y range 참조). 반환: COM Chart."""
    left, top = _chart_pos(i)
    with _prof("dist.series_add"):
        with _dist_time("dist.loop.chart_add"):
            ch = sh.charts.add(left, top, _CHART_W, _CHART_H)
            chart_api = _chart_com(ch)   # api[1] 1회 바인딩, 이하 하위 함수에 전달
        with _dist_time("dist.loop.chart_type"):
            ch.chart_type = "xy_scatter_lines_no_markers"
    _chart_title_limit_set(chart_api, d)
    with _prof("dist.style"):
        limit_count = _chart_data_set(
            chart_api, d, x_sheet, y_sheet, col_idx, cnt_list, src_names, x_min,
            colors, y_cols_by_source, count_matrix)
    with _prof("dist.format"):
        _chart_layout_setting(chart_api, x_min, x_max, is_fail, limit_count)
    return chart_api


def _chart_title_limit_set(chart_api, d):
    """차트 제목: item 명 + 둘째 줄 (LO~HI unit) 캡션 (현재 서식 유지)."""
    with _dist_time("dist.loop.title_format"):
        try:
            chart_api.HasTitle = True
            title = chart_api.ChartTitle
            cap = _limit_caption(d)          # item 명 아래 줄: (LO ~ HI units)
            title.Text = d.subject + "\n" + cap
            tf = title.Font
            tf.Name = "Arial Black"
            tf.Size = _CHART_TITLE_ITEM_FONT
            try:                             # 둘째 줄(캡션)은 작게
                title.Characters(len(d.subject) + 2, len(cap)).Font.Size = _CHART_TITLE_CAP_FONT
            except Exception:
                pass
            title.Top = 0
        except Exception:
            pass


def _chart_data_set(chart_api, d, x_sheet, y_sheet, col_idx, cnt_list, src_names, x_min,
                    colors, y_cols_by_source, count_matrix):
    """LSL/USL + source series 를 생성하고 같은 객체에 스타일까지 한 번에 적용.

    series 1=LSL, 2=USL(없으면 차트 밖 -2,-2), 3+=source. 정리/정리_Y 시트의 subject
    열을 source별 행구간으로 참조(compact-Y). SeriesCollection 은 1회 바인딩, 각 series
    는 NewSeries 직후 같은 객체에 스타일 적용 — COM 재조회 없음. limit line 스타일도
    series 객체를 보유한 여기서 적용(COM 최소화). 반환: limit series 개수(범례 삭제 수).
    """
    col = get_column_letter(col_idx + 2)   # A=index, B=subject0
    lo = float(d.lower_limit) if _isnum(d.lower_limit) else None
    hi = float(d.upper_limit) if _isnum(d.upper_limit) else None
    xv0 = x_min if x_min is not None else 0.0
    sc = chart_api.SeriesCollection()
    limit_count = 0
    with _dist_time("dist.loop.series_limits"):
        for lim, nm in ((lo, "LSL"), (hi, "USL")):
            s = sc.NewSeries()
            if lim is not None:
                s.XValues = (lim, lim)
                s.Values = (-1.0, 1.0)            # x=lim 세로선(Y 0~1 덮음)
            else:
                s.XValues = (xv0, xv0)
                s.Values = (-2.0, -2.0)           # 차트 밖(안 보임) — series 인덱스 안정용
            s.Name = nm
            _style_limit_series(s)
            limit_count += 1
    y = 0
    with _dist_time("dist.loop.series_sources"):
        for k, name in enumerate(src_names):
            n = cnt_list[k]
            valid_count = min(int(count_matrix[k][col_idx]), n)
            r1 = y + 2
            y += n
            s = sc.NewSeries()
            if valid_count > 0:
                r2 = r1 + valid_count - 1
                x_ref = f"='{x_sheet}'!${col}${r1}:${col}${r2}"
                y_col = y_cols_by_source[k][col_idx]
                y_ref = f"='{y_sheet}'!${y_col}${r1}:${y_col}${r2}"
                s.XValues = x_ref
                s.Values = y_ref
            else:
                s.XValues = (x_min if x_min is not None else 0.0,
                             x_min if x_min is not None else 0.0)
                s.Values = (-2.0, -2.0)
            s.Name = str(name)
            rgb = _hex_to_excel_rgb(colors[k % len(colors)]) if colors else None
            _style_data_series(s, rgb)
    return limit_count


def _chart_layout_setting(chart_api, x_min, x_max, is_fail, limit_count):
    """축·gridline·plotarea·fail 배경·범례 삭제 (COM 객체별 1회 바인딩, 현재 서식 유지).

    range 재바인딩이 없는 신규 차트라 fail 이 아니면 ChartArea 는 기본(흰색) 그대로 둔다.
    범례는 맨 마지막에 앞 limit_count(=LSL/USL)개를 인덱스로 삭제.
    """
    with _dist_time("dist.loop.axis_format"):
        try:
            yax = chart_api.Axes(_XL_VALUE, _XL_PRIMARY)
            yax.MinimumScale = 0
            yax.MaximumScale = 1
            yax.MajorUnit = 0.2
            yax.HasMinorGridlines = True
            ytl = yax.TickLabels
            ytl.NumberFormatLocal = "0%"
            ytl.Font.Size = 8
            yax.TickLabelPosition = _XL_LOW
        except Exception:
            pass
        try:
            xax = chart_api.Axes(_XL_CATEGORY, _XL_PRIMARY)
            if x_min is not None and x_max is not None and x_min < x_max:
                xax.MinimumScale = x_min
                xax.MaximumScale = x_max
            xax.HasMinorGridlines = True
            xax.TickLabels.Font.Size = 8
        except Exception:
            pass
    with _dist_time("dist.loop.plot_format"):
        try:
            pa = chart_api.PlotArea
            pa.Width = _PLOT_W
            pa.Top = _PLOT_TOP
            pa.Height = _PLOT_H
        except Exception:
            pass
    if is_fail:
        with _dist_time("dist.loop.fail_bg"):
            try:
                chart_api.ChartArea.Interior.Color = _RGB_FAIL_BG
            except Exception:
                pass
    with _dist_time("dist.loop.legend_format"):
        try:
            chart_api.HasLegend = True
            leg = chart_api.Legend
            leg.Font.Size = 8
            for _ in range(limit_count):
                try:
                    leg.LegendEntries(1).Delete()
                except Exception:
                    pass
        except Exception:
            pass


# ── 차트 생성 함수 line-by-line 디버그 트레이서 ──────────────────────────────
# _chart_draw_at_position 호출 1건(그 호출 스택 전체 — title/data/layout 등 하위 함수
# + COM 조작 줄 포함)의 모든 실행 줄을 줄 단위 소요시간과 함께 stderr 로 출력한다.
# DEBUG_CHART_LINE_TRACE is off by default, so sys.settrace is not called in normal runs.
# 동작·성능에 영향 없음.

def _chart_line_tracer(state):
    """sys.settrace 콜백. 이 모듈 파일에 속한 프레임의 'line' 이벤트만 기록한다.

    다른 파일(numpy/pandas/get_column_letter 등 외부 라이브러리)로 진입하는 프레임은
    None 을 반환해 추적을 차단 — 소스가 없어 줄 단위 추적이 불가능한 컴파일 확장
    (pywin32 COM 등) 및 무관한 노이즈를 거른다. 활성 구간 자체가
    _chart_draw_at_position 호출 1건으로 한정되므로 실질적으로는 그 호출 스택만
    보게 된다.

    'line' 이벤트마다 "직전 줄"의 (filename, lineno, func, elapsed) 를
    state['entries'] 에 push 하고 현재 줄을 state['last'] 로 갱신한다 — 호출 스택을
    가로지른 단일 시간순 로그(실제 실행 순서 그대로).
    """
    def tracer(frame, event, _arg):
        if os.path.normcase(frame.f_code.co_filename) != _LINE_TRACE_FILE_KEY:
            return None
        if event == "call":
            return tracer
        if event == "line":
            now = time.perf_counter()
            last = state["last"]
            if last is not None:
                state["entries"].append((last[0], last[1], now - last[2]))
            state["last"] = (frame.f_lineno, frame.f_code.co_name, now)
            return tracer
        return tracer
    return tracer


def _dump_chart_line_trace(seq, subject, entries, total_elapsed):
    sep = "=" * 88
    print(sep, file=sys.stderr)
    print(f"[chart-line-trace] chart_seq={seq} subject={subject!r} "
          f"lines={len(entries)} total={total_elapsed * 1000:.3f}ms", file=sys.stderr)
    print(sep, file=sys.stderr)
    for lineno, func, elapsed in entries:
        src = linecache.getline(_LINE_TRACE_FILE, lineno).rstrip()
        print(f"  {func}:{lineno} | {elapsed * 1000:8.3f} ms | {src}", file=sys.stderr)
    print(sep, file=sys.stderr, flush=True)


@contextlib.contextmanager
def _trace_chart_lines(seq, subject):
    """대상 차트(seq)일 때만 _chart_draw_at_position 호출 1건을 줄 단위로 추적.

    대상이 아니면 sys.settrace 호출 없이 즉시 통과 → 다른 모든 차트·코드 경로는
    평소와 100% 동일하게 실행된다.
    """
    if not _LINE_TRACE_ON or seq not in _LINE_TRACE_TARGETS:
        yield
        return
    state = {"entries": [], "last": None}
    old_trace = sys.gettrace()
    sys.settrace(_chart_line_tracer(state))
    t0 = time.perf_counter()
    try:
        yield
    finally:
        sys.settrace(old_trace)
        last = state["last"]
        if last is not None:
            state["entries"].append((last[0], last[1], time.perf_counter() - last[2]))
        _dump_chart_line_trace(seq, subject, state["entries"], time.perf_counter() - t0)


def _write_distribution(wb, sh, result, colors=None, dist_progress_cb=None,
                        dists=None, sources=None, source_data=None, title="Distribution",
                        profile_label="main"):
    """각 subject 의 누적분포(ECDF) 차트 — 모든 DUT 가 1점(중복 제거 없음).

    source(input file)별 데이터 series + LSL/USL 세로 한계선(series 1,2, COM 배열리터럴).
    데이터는 정리(X)/정리_Y(Y) 두 시트에 통째 1회 bulk write, 차트 series 는 그 시트의
    subject 열을 source별 행구간으로 참조한다. 차트는 gap 없이 밀착 배치·독립 생성.

    dists/sources/source_data: diff compare 의 a_only/b_only 시트용 override (None 이면
    result.distributions / result.sources / result.dist_source_data 사용).
    """
    dists = result.distributions if dists is None else dists
    sources = result.sources if sources is None else sources
    source_data = result.dist_source_data if source_data is None else source_data
    if not dists or not source_data:
        sh.range("A1").value = "선택된 항목에 분포 데이터가 없습니다."
        return {}

    subj_names = [d.subject for d in dists]
    sd_map = dict(source_data)
    src_dfs, src_names = [], []
    for s in sources:
        df = sd_map.get(s)
        if df is None:
            continue
        src_dfs.append(df.reindex(columns=subj_names))   # 없는 subject 열 → NaN
        src_names.append(s)
    if not src_dfs:
        sh.range("A1").value = "선택된 항목에 분포 데이터가 없습니다."
        return {}

    # ── 데이터 변환: 열별 정렬(X) + rank/count(Y), source별 블록 concat ──────────
    with _profile_info_time("distribution.data_write"):
        with _prof("dist.data_write"):
            with _dist_time("dist.data_prepare"):
                df_x_list = [sort_alldata(df, True).reset_index(drop=True) for df in src_dfs]
                df_x = pd.concat(df_x_list, ignore_index=True)
                df_y, y_cols_by_source, count_matrix, y_stats = _build_compact_dist_y(df_x_list)
                cnt_list = [df.shape[0] for df in df_x_list]   # source별 DUT 수(모든 subject 균일)
                reduction = 0.0
                if y_stats["full_y_cells"]:
                    reduction = 100.0 * (1.0 - (
                        y_stats["compact_y_cells"] / y_stats["full_y_cells"]))
                _emit_profile_info(
                    "Dist helper Y mode: sheet count_cases "
                    f"x_cells={df_x.shape[0] * df_x.shape[1]:,} "
                    f"full_y_cells={y_stats['full_y_cells']:,} "
                    f"compact_y_cells={y_stats['compact_y_cells']:,} "
                    f"unique_counts={y_stats['unique_counts']} "
                    f"unique_count_cases={y_stats['unique_count_cases']} "
                    f"valid_refs={y_stats['valid_refs']:,} "
                    f"trimmed_nan_tail={y_stats['trimmed_nan_tail']:,} "
                    f"reduction={reduction:.1f}%"
                )

            with _dist_time("dist.helper_sheet_write"):
                existing = {s.name for s in wb.sheets}
                x_name = _unique_helper_name("정리", existing)
                y_name = _unique_helper_name("정리_Y", existing | {x_name})
                ws_x = wb.sheets.add(x_name, after=sh)
                ws_y = wb.sheets.add(y_name, after=ws_x)
                ws_x.range("A1").options(index=True, header=True).value = df_x
                ws_y.range("A1").options(index=True, header=True).value = df_y
                for w in (ws_x, ws_y):
                    try:
                        w.api.Visible = False
                    except Exception:
                        pass

    _put_title(sh, 8, title)
    sh.range((1, _INDEX_COL)).value = "Item Index (Ctrl+F)"
    sh.range((1, _INDEX_COL)).column_width = 26

    # 정렬값 numpy 캐시(데이터 min/max·step 판정용) — 셀 읽기 없이
    x_arrs = [df.to_numpy(dtype=float) for df in df_x_list]   # [source](N, n_subj)

    n_charts = len(dists)
    with _profile_info_time(f"distribution.chart_create[{profile_label}]"):
        chart_map, index_entries = _draw_all_chart(
            sh, dists, x_arrs, x_name, y_name, cnt_list, src_names, colors,
            y_cols_by_source, count_matrix, n_charts, dist_progress_cb)

    # Item Index 열 일괄 기입
    if index_entries:
        max_idx = max(r for r, _ in index_entries)
        col_vals = [[None] for _ in range(2, max_idx + 1)]
        for r, subj in index_entries:
            col_vals[r - 2] = [subj]
        sh.range((2, _INDEX_COL), (max_idx, _INDEX_COL)).value = col_vals

    _prof_count("charts", len(chart_map))
    _finalize_title_row(sh)
    return chart_map


def _chart_com(ch):
    """xlwings Chart → COM Chart 객체. ch.api 가 (ChartObject, Chart) 튜플일 수 있음."""
    api = ch.api
    if isinstance(api, tuple):
        return api[1]
    return getattr(api, "Chart", api)


def _style_limit_series(s):
    """limit line series: 빨강 system dash + 마커 제거 (선만)."""
    try:
        line = s.Format.Line
        line.DashStyle = _MSO_LINE_SYSDASH
        line.ForeColor.RGB = _RGB_RED
    except Exception:
        pass
    try:
        s.MarkerStyle = _XL_MARKER_NONE
    except Exception:
        pass


def _style_data_series(s, rgb=None):
    """Source data series: always show all DUT values as dot markers."""
    try:
        s.Format.Line.Visible = _MSO_FALSE
    except Exception:
        pass
    try:
        s.MarkerStyle = _XL_MARKER_CIRCLE
        s.MarkerSize = _MARKER_SIZE
    except Exception:
        pass
    if rgb is not None:
        try:
            s.MarkerBackgroundColor = rgb
            s.MarkerForegroundColor = rgb
        except Exception:
            pass


# ── x축 범위 계산 헬퍼 ───────────────────────────────────────────────────────

def _isnum(v):
    if v is None:
        return False
    try:
        f = float(v)
        return not math.isnan(f) and not math.isinf(f)
    except (TypeError, ValueError):
        return False


def _fmt_lim(v):
    """limit 표시값: nan/None → '-', 정수 → int, 그 외 → 간결한 실수."""
    if not _isnum(v):
        return "-"
    f = float(v)
    return str(int(f)) if f.is_integer() else f"{f:g}"


def _limit_caption(d):
    """차트 item 명 아래 줄: '(LO ~ HI units)'."""
    unit = (d.unit or "").strip()
    body = f"{_fmt_lim(d.lower_limit)} ~ {_fmt_lim(d.upper_limit)}"
    return f"({body} {unit})" if unit else f"({body})"


def _decimals(v):
    """숫자의 유효 소수 자릿수 (정수면 0)."""
    if v is None:
        return 0
    s = repr(float(v))
    if "e" in s or "E" in s or "." not in s:
        return 0
    return len(s.split(".")[1].rstrip("0"))


def _floor_dec(x, dec):
    f = 10 ** dec
    return math.floor(x * f) / f


def _ceil_dec(x, dec):
    f = 10 ** dec
    return math.ceil(x * f) / f


def _x_axis_range(lo, hi, dmin, dmax, is_fail):
    """x축 [min,max]. Pass=LIM 그대로, Fail=±5% 가드밴드 후 LIM 자릿수로 floor/ceil.
    LIM None/nan 이면 data min/max 사용."""
    lo_n = float(lo) if _isnum(lo) else None
    hi_n = float(hi) if _isnum(hi) else None
    xmin = lo_n if lo_n is not None else dmin
    xmax = hi_n if hi_n is not None else dmax
    if not is_fail:
        return xmin, xmax
    if lo_n is not None and dmin < lo_n:
        xmin = dmin - (lo_n - dmin) * 0.05
    if hi_n is not None and dmax > hi_n:
        xmax = dmax + (dmax - hi_n) * 0.05
    dec = max(_decimals(lo_n), _decimals(hi_n))
    return _floor_dec(xmin, dec), _ceil_dec(xmax, dec)


# ── distribution 시트 제목 배너 (xlwings) ────────────────────────────────────

_XL_CENTER = -4108        # xlCenter
_XL_LEFT = -4131          # xlLeft
_TITLE_FILL = (191, 227, 255)
_TITLE_FONT_SIZE = 20
_TITLE_ROW_HEIGHT = 30


def _put_title(sh, ncols, text):
    """1행에 시트 제목 배너 — 하늘색 배경, 검은 bold, 큰 글씨, 가로 병합."""
    span = max(_TITLE_ROW_MAX_COL, ncols)
    try:
        sh.range((1, 1), (1, span)).merge()
    except Exception:
        pass
    c = sh.range((1, 1))
    c.value = text
    try:
        c.color = _TITLE_FILL
        f = c.api.Font
        f.Bold = True
        f.Size = _TITLE_FONT_SIZE
        f.Color = 0  # black
        c.api.HorizontalAlignment = _XL_LEFT
        c.api.VerticalAlignment = _XL_CENTER
        sh.range((1, 1), (1, span)).color = _TITLE_FILL
    except Exception:
        pass


def _finalize_title_row(sh):
    """제목 행 높이 보정."""
    try:
        sh.range((1, 1)).row_height = _TITLE_ROW_HEIGHT
    except Exception:
        pass
