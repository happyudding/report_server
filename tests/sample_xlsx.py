"""더미 .xlsx 생성기.

외부 report generator 가 만드는 실제 산출물 형식을 모방한다.
xlsx_parser.parse_report_xlsx() 가 정상 동작하는지 검증하는 입력 샘플로 사용.

실행:
    python tests/sample_xlsx.py [output_path]
기본 출력: tests/sample.xlsx
"""
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl 미설치. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_HDR_FILL = PatternFill("solid", fgColor="E2E8F0")
_HDR2_FILL = PatternFill("solid", fgColor="D9E1F2")
_TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style(ws, coord, *, sz=None, bold=False, fill=None, center=False):
    c = ws[coord]
    c.font = Font(size=sz or 11, bold=bold)
    if fill:
        c.fill = fill
    if center:
        c.alignment = _CENTER
    return c


def build_sample_xlsx(out_path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb.create_sheet("summary"))
    _write_yield(wb.create_sheet("yield"))
    _write_cpk(wb.create_sheet("cpk"))
    _write_fail_data(wb.create_sheet("fail_data"))
    _write_fail_values(wb.create_sheet("fail_values"))
    _write_issue_table(wb.create_sheet("issue_table"))
    _write_distribution(wb.create_sheet("distribution"))
    _write_histogram(wb.create_sheet("histogram"))

    wb.save(str(out_path))
    print(f"[ok] wrote {out_path}")


def _write_summary(ws):
    """실측 레이아웃: B4='DEVICE' 앵커. Device Feature + 2.Yield + Major Fail Bins."""
    # 제목 배너 (A1:H1 병합, 22pt)
    ws["A1"] = "📊 REPORT TITLE — sample-001"
    ws.merge_cells("A1:H1")
    _style(ws, "A1", sz=22, bold=True, fill=PatternFill("solid", fgColor="BFE3FF"))

    # 1. Device Feature (B3:C3 병합 제목, B4 헤더, B5 값)
    ws["B3"] = "1. Device Feature"; ws.merge_cells("B3:C3")
    _style(ws, "B3", sz=20, bold=True)
    feat_hdr = ["DEVICE", "Customer", "PKG_Type", "GrossDie", "Process Line", "EVT_Version"]
    for i, h in enumerate(feat_hdr):
        col = chr(ord("B") + i)
        ws[f"{col}4"] = h
        _style(ws, f"{col}4", sz=10, bold=True, fill=_HDR_FILL, center=True)
    feat_val = ["S5E_TEST", "SLSI", "FCBGA", 280, "L1", "EVT0"]
    for i, v in enumerate(feat_val):
        col = chr(ord("B") + i)
        ws[f"{col}5"] = v
        _style(ws, f"{col}5", sz=10, center=True)

    # 2. Yield + Major Fail Bins
    ws["B7"] = "2. Yield"; _style(ws, "B7", sz=20, bold=True)
    for coord, txt in [("B8", "Lot NO"), ("D8", "Yield"), ("E8", "Major Fail Bins"), ("H8", "Comment")]:
        ws[coord] = txt
        _style(ws, coord, sz=10, bold=True, center=True)
    ws.merge_cells("E8:G8")
    ws["B9"] = "-"; ws["D9"] = 93.3
    fails = [("1st Fail", "subject_14", 0.4), ("2nd Fail", "subject_25", 0.4),
             ("3rd Fail", "subject_33", 0.4), ("4th Fail", "subject_03", 0.3),
             ("5th Fail", "subject_09", 0.3)]
    for i, (rank, subj, ratio) in enumerate(fails):
        r = 9 + i
        ws[f"E{r}"] = rank; ws[f"F{r}"] = subj; ws[f"G{r}"] = ratio
        _style(ws, f"E{r}", sz=10, center=True)

    # 열너비·행높이 (원형 재현 검증용)
    for col, w in {"A": 2.6, "B": 16.0, "C": 26.1, "D": 10.4,
                   "E": 10.5, "F": 12.6, "G": 9.0, "H": 44.8}.items():
        ws.column_dimensions[col].width = w
    for r, h in {1: 30.0, 3: 25.5, 4: 16.5, 7: 21.75, 8: 16.5}.items():
        ws.row_dimensions[r].height = h


def _write_yield(ws):
    """실측 레이아웃: A1 배너, B3='bin' 헤더행, 데이터 4행~."""
    ws["A1"] = "Yield"; ws.merge_cells("A1:J1")
    _style(ws, "A1", sz=20, bold=True, fill=_TITLE_FILL, center=True)

    header = ["bin", "Item", "mass_data_a_count", "mass_data_a_yield",
              "mass_data_b_count", "mass_data_b_yield",
              "mass_data_c_count", "mass_data_c_yield", "avg", "comment"]
    rows = [
        [1, "Pass", 280, 93.33, 327, 93.43, 326, 93.14, 93.3, ""],
        [10, "subject_03", 1, 0.33, 3, 0.86, 3, 0.86, 0.68, ""],
        [11, "subject_09", 0, 0.0, 3, 0.86, 3, 0.86, 0.57, "전압 측정 실패"],
        [12, "subject_14", 0, 0.0, 3, 0.86, 3, 0.86, 0.57, "온도 한계 초과"],
    ]
    for i, h in enumerate(header):
        ws.cell(3, 2 + i, h)
        _style(ws, f"{chr(ord('B')+i)}3", sz=12, bold=True, fill=_HDR2_FILL, center=True)
    for ri, row in enumerate(rows):
        for ci, v in enumerate(row):
            ws.cell(4 + ri, 2 + ci, v)
            _style(ws, f"{chr(ord('B')+ci)}{4+ri}", sz=12, center=True)

    for col, w in {"B": 6.5, "C": 20.0, "D": 6.5, "K": 50.0}.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30.0
    ws.row_dimensions[3].height = 39.95
    for r in range(4, 8):
        ws.row_dimensions[r].height = 21.95


def _write_cpk(ws):
    ws.append(["subject", "lower_limit", "upper_limit", "units", "n", "average", "stdev", "cpk"])
    ws.append(["VDD_LOW", 0.9, 1.1, "V", 1200, 1.005, 0.012, 1.85])
    ws.append(["TEMP_HI", -40, 85, "C", 1200, 32.5, 12.3, 1.42])


def _write_fail_data(ws):
    ws["A1"] = "see fail_values sheet for detail"


def _write_fail_values(ws):
    ws.append(["source", "dut", "x_coord", "y_coord", "bin",
               "subject", "value", "lower_limit", "upper_limit", "fail"])
    ws.append(["A.csv", 101, 3, 5, 2, "VDD_LOW", 0.82, 0.9, 1.1, "< lo"])
    ws.append(["A.csv", 102, 4, 6, 3, "TEMP_HI", 92.0, -40, 85, "> hi"])


def _write_issue_table(ws):
    """실측 레이아웃: A1 배너, B3='Category'/C3='Bin' 헤더행, I열=Distribution."""
    ws["A1"] = "Issue_table"; ws.merge_cells("A1:J1")
    _style(ws, "A1", sz=20, bold=True, fill=_TITLE_FILL, center=True)

    header = ["Category", "Bin", "Item", "avg",
              "mass_data_a_yield", "mass_data_b_yield", "mass_data_c_yield",
              "Distribution", "comment", "개발팀 1차 comment"]
    rows = [
        ["Yield", 1, "Pass", 93.3, 93.33, 93.43, 93.14, "", "", ""],
        ["", 10, "subject_03", 0.68, 0.33, 0.86, 0.86, "", "재현 가능", "PMU 확인 필요"],
        ["", 11, "subject_09", 0.57, 0.0, 0.86, 0.86, "", "고온 조건", "센서 보정 필요"],
    ]
    for i, h in enumerate(header):
        ws.cell(3, 2 + i, h)
        _style(ws, f"{chr(ord('B')+i)}3", sz=15, fill=_HDR2_FILL, center=True)
    for ri, row in enumerate(rows):
        for ci, v in enumerate(row):
            ws.cell(4 + ri, 2 + ci, v)
            _style(ws, f"{chr(ord('B')+ci)}{4+ri}", sz=15, center=True)
    # Category 세로 병합 (병합은 grid 에 저장하지 않지만 실측 구조 모방)
    ws.merge_cells("B4:B6")

    for col, w in {"B": 20.0, "C": 6.5, "D": 20.0, "E": 6.5, "I": 17.0, "J": 40.0}.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30.0
    for r in range(3, 7):
        ws.row_dimensions[r].height = 78.0


def _write_distribution(ws):
    ws["A1"] = "(image placeholder)"


def _write_histogram(ws):
    ws["A1"] = "(image placeholder)"


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else str(Path(__file__).parent / "sample.xlsx")
    build_sample_xlsx(out)
